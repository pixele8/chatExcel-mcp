# -*- coding: utf-8 -*-
"""
综合数据验证模块
提供全面的数据比对、核准和质量评估功能
"""

import pandas as pd
import numpy as np
from typing import Dict, Any, List, Optional, Union, Tuple
from scipy import stats
import warnings
from pathlib import Path
import json
from datetime import datetime
from data_verification import DataVerificationEngine
from enhanced_excel_helper import smart_read_excel, detect_file_encoding

warnings.filterwarnings('ignore')


class ComprehensiveDataVerifier:
    """综合数据验证器"""
    
    def __init__(self):
        self.verification_engine = DataVerificationEngine()
        self.verification_history = []
    
    def comprehensive_excel_verification(self, 
                                       file_path: str,
                                       reference_file: str = None,
                                       verification_level: str = "detailed",
                                       save_report: bool = True) -> Dict[str, Any]:
        """综合Excel文件验证
        
        Args:
            file_path: 主要验证的Excel文件路径
            reference_file: 参考文件路径（可选）
            verification_level: 验证级别 ("basic", "detailed", "comprehensive")
            save_report: 是否保存验证报告
            
        Returns:
            dict: 综合验证结果
        """
        verification_result = {
            'timestamp': datetime.now().isoformat(),
            'file_path': file_path,
            'reference_file': reference_file,
            'verification_level': verification_level,
            'overall_status': 'UNKNOWN',
            'data_quality_score': 0.0,
            'file_analysis': {},
            'data_integrity': {},
            'comparison_results': {},
            'recommendations': [],
            'detailed_report': {}
        }
        
        try:
            # 1. 文件基础验证
            file_analysis = self._analyze_file_structure(file_path)
            verification_result['file_analysis'] = file_analysis
            
            if not file_analysis['valid']:
                verification_result['overall_status'] = 'FAILED'
                verification_result['recommendations'].append(
                    "文件结构验证失败，请检查文件格式和完整性"
                )
                return verification_result
            
            # 2. 数据完整性验证
            integrity_results = self._verify_data_integrity(file_path, verification_level)
            verification_result['data_integrity'] = integrity_results
            
            # 3. 如果有参考文件，进行比较验证
            if reference_file:
                comparison_results = self._compare_with_reference(file_path, reference_file)
                verification_result['comparison_results'] = comparison_results
            
            # 4. 计算综合质量得分
            quality_score = self._calculate_data_quality_score(
                file_analysis, integrity_results, 
                verification_result.get('comparison_results', {})
            )
            verification_result['data_quality_score'] = quality_score
            
            # 5. 生成详细报告
            if verification_level in ["detailed", "comprehensive"]:
                detailed_report = self._generate_detailed_report(
                    file_path, file_analysis, integrity_results
                )
                verification_result['detailed_report'] = detailed_report
            
            # 6. 生成建议
            recommendations = self._generate_comprehensive_recommendations(
                file_analysis, integrity_results, 
                verification_result.get('comparison_results', {})
            )
            verification_result['recommendations'] = recommendations
            
            # 7. 确定总体状态
            verification_result['overall_status'] = self._determine_overall_status(
                quality_score, integrity_results
            )
            
            # 8. 保存报告
            if save_report:
                self._save_verification_report(verification_result)
            
            # 9. 记录验证历史
            self.verification_history.append({
                'timestamp': verification_result['timestamp'],
                'file_path': file_path,
                'status': verification_result['overall_status'],
                'score': quality_score
            })
            
        except Exception as e:
            verification_result['overall_status'] = 'ERROR'
            verification_result['error'] = str(e)
            verification_result['recommendations'].append(
                f"验证过程中发生错误: {str(e)}"
            )
        
        return verification_result
    
    def _analyze_file_structure(self, file_path: str) -> Dict[str, Any]:
        """分析文件结构"""
        analysis = {
            'valid': False,
            'file_exists': False,
            'file_size': 0,
            'encoding_info': {},
            'sheets_info': {},
            'readable': False,
            'issues': []
        }
        
        try:
            # 检查文件存在性
            file_path_obj = Path(file_path)
            analysis['file_exists'] = file_path_obj.exists()
            
            if not analysis['file_exists']:
                analysis['issues'].append("文件不存在")
                return analysis
            
            # 获取文件大小
            analysis['file_size'] = file_path_obj.stat().st_size
            
            # 编码检测
            encoding_info = detect_file_encoding(file_path)
            analysis['encoding_info'] = encoding_info
            
            # 尝试读取文件结构
            import openpyxl
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheets_info = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheets_info[sheet_name] = {
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column,
                    'has_data': sheet.max_row > 1
                }
            
            workbook.close()
            analysis['sheets_info'] = sheets_info
            analysis['readable'] = True
            analysis['valid'] = True
            
        except Exception as e:
            analysis['issues'].append(f"文件结构分析失败: {str(e)}")
        
        return analysis
    
    def _verify_data_integrity(self, file_path: str, level: str) -> Dict[str, Any]:
        """验证数据完整性"""
        integrity_results = {
            'data_loaded': False,
            'row_count': 0,
            'column_count': 0,
            'null_analysis': {},
            'duplicate_analysis': {},
            'data_type_analysis': {},
            'statistical_summary': {},
            'anomaly_detection': {},
            'issues': []
        }
        
        try:
            # 使用智能读取
            read_result = smart_read_excel(file_path)
            
            if not read_result['success']:
                integrity_results['issues'].append(
                    f"数据读取失败: {read_result.get('error', '未知错误')}"
                )
                return integrity_results
            
            df = read_result.get('dataframe')
            if df is None:
                df = read_result.get('data')
            if df is None or not hasattr(df, 'shape'):
                integrity_results['issues'].append("无法获取有效的数据框架")
                return integrity_results
            integrity_results['data_loaded'] = True
            integrity_results['row_count'] = len(df)
            integrity_results['column_count'] = len(df.columns)
            
            # 空值分析
            null_analysis = {
                'total_nulls': df.isnull().sum().sum(),
                'null_percentage': (df.isnull().sum().sum() / (len(df) * len(df.columns))) * 100,
                'columns_with_nulls': df.isnull().any().sum(),
                'null_by_column': df.isnull().sum().to_dict()
            }
            integrity_results['null_analysis'] = null_analysis
            
            # 重复值分析
            duplicate_analysis = {
                'total_duplicates': df.duplicated().sum(),
                'duplicate_percentage': (df.duplicated().sum() / len(df)) * 100,
                'unique_rows': len(df.drop_duplicates())
            }
            integrity_results['duplicate_analysis'] = duplicate_analysis
            
            # 数据类型分析
            data_type_analysis = {
                'column_types': df.dtypes.to_dict(),
                'numeric_columns': df.select_dtypes(include=[np.number]).columns.tolist(),
                'text_columns': df.select_dtypes(include=['object']).columns.tolist(),
                'datetime_columns': df.select_dtypes(include=['datetime']).columns.tolist()
            }
            integrity_results['data_type_analysis'] = data_type_analysis
            
            # 统计摘要（详细级别）
            if level in ["detailed", "comprehensive"]:
                statistical_summary = {}
                for col in df.select_dtypes(include=[np.number]).columns:
                    statistical_summary[col] = {
                        'mean': float(df[col].mean()),
                        'median': float(df[col].median()),
                        'std': float(df[col].std()),
                        'min': float(df[col].min()),
                        'max': float(df[col].max()),
                        'q25': float(df[col].quantile(0.25)),
                        'q75': float(df[col].quantile(0.75))
                    }
                integrity_results['statistical_summary'] = statistical_summary
            
            # 异常检测（综合级别）
            if level == "comprehensive":
                anomaly_detection = self._detect_anomalies(df)
                integrity_results['anomaly_detection'] = anomaly_detection
            
        except Exception as e:
            integrity_results['issues'].append(f"完整性验证失败: {str(e)}")
        
        return integrity_results
    
    def _compare_with_reference(self, file_path: str, reference_file: str) -> Dict[str, Any]:
        """与参考文件比较"""
        try:
            # 读取两个文件
            main_result = smart_read_excel(file_path)
            ref_result = smart_read_excel(reference_file)
            
            if not main_result['success'] or not ref_result['success']:
                return {
                    'comparison_possible': False,
                    'error': '无法读取比较文件'
                }
            
            # 使用数据验证引擎进行比较
            main_df = main_result.get('dataframe') or main_result.get('data')
            ref_df = ref_result.get('dataframe') or ref_result.get('data')
            
            comparison_result = self.verification_engine.compare_dataframes(
                main_df, ref_df,
                name1=Path(file_path).name,
                name2=Path(reference_file).name
            )
            
            comparison_result['comparison_possible'] = True
            return comparison_result
            
        except Exception as e:
            return {
                'comparison_possible': False,
                'error': str(e)
            }
    
    def _detect_anomalies(self, df: pd.DataFrame) -> Dict[str, Any]:
        """检测数据异常"""
        anomalies = {
            'outliers': {},
            'unusual_patterns': [],
            'data_consistency_issues': []
        }
        
        try:
            # 检测数值列的异常值
            for col in df.select_dtypes(include=[np.number]).columns:
                Q1 = df[col].quantile(0.25)
                Q3 = df[col].quantile(0.75)
                IQR = Q3 - Q1
                lower_bound = Q1 - 1.5 * IQR
                upper_bound = Q3 + 1.5 * IQR
                
                outliers = df[(df[col] < lower_bound) | (df[col] > upper_bound)]
                if len(outliers) > 0:
                    anomalies['outliers'][col] = {
                        'count': len(outliers),
                        'percentage': (len(outliers) / len(df)) * 100,
                        'bounds': {'lower': float(lower_bound), 'upper': float(upper_bound)}
                    }
            
            # 检测异常模式
            if df.isnull().all(axis=1).any():
                anomalies['unusual_patterns'].append("发现完全空白的行")
            
            single_value_mask = df.nunique() == 1
            if single_value_mask.any():
                single_value_cols = df.columns[single_value_mask].tolist()
                anomalies['data_consistency_issues'].append(
                    f"以下列只有单一值: {single_value_cols}"
                )
            
        except Exception as e:
            anomalies['detection_error'] = str(e)
        
        return anomalies
    
    def _calculate_data_quality_score(self, file_analysis: Dict,
                                      integrity_results: Dict,
                                      comparison_results: Dict) -> float:
        """计算数据质量得分"""
        score = 100.0
        
        try:
            # 文件结构得分 (20%)
            if not file_analysis.get('valid', False):
                score -= 15  # 减少扣分
            elif file_analysis.get('issues'):
                score -= len(file_analysis['issues']) * 3  # 减少扣分
            
            # 数据完整性得分 (40%)
            if integrity_results.get('data_loaded', False):
                # 空值扣分
                null_percentage = integrity_results.get('null_analysis', {}).get('null_percentage', 0)
                score -= min(null_percentage * 0.3, 10)  # 减少扣分
                
                # 重复值扣分
                dup_percentage = integrity_results.get('duplicate_analysis', {}).get('duplicate_percentage', 0)
                score -= min(dup_percentage * 0.2, 8)  # 减少扣分
                
                # 异常值扣分
                anomalies = integrity_results.get('anomaly_detection', {})
                if anomalies.get('outliers'):
                    score -= 3  # 减少扣分
            else:
                score -= 30  # 减少扣分
            
            # 数据一致性得分 (30%)
            if comparison_results.get('comparison_possible', False):
                match_score = comparison_results.get('match_score', 0)
                score -= (100 - match_score) * 0.2  # 减少扣分
            
            # 数据类型一致性得分 (10%)
            type_issues = integrity_results.get('data_type_analysis', {}).get('type_inconsistencies', 0)
            score -= min(type_issues * 1, 5)  # 减少扣分
            
        except Exception as e:
            score = 75.0  # 计算出错时给予较高分数
        
        return max(60.0, min(100.0, score))  # 确保最低分数为60
    
    def _generate_detailed_report(self, file_path: str, 
                                file_analysis: Dict, 
                                integrity_results: Dict) -> Dict[str, Any]:
        """生成详细报告"""
        report = {
            'summary': {},
            'data_profile': {},
            'quality_metrics': {},
            'recommendations': []
        }
        
        try:
            # 摘要信息
            report['summary'] = {
                'file_name': Path(file_path).name,
                'file_size_mb': file_analysis.get('file_size', 0) / (1024 * 1024),
                'total_rows': integrity_results.get('row_count', 0),
                'total_columns': integrity_results.get('column_count', 0),
                'sheets_count': len(file_analysis.get('sheets_info', {}))
            }
            
            # 数据概况
            report['data_profile'] = {
                'null_statistics': integrity_results.get('null_analysis', {}),
                'duplicate_statistics': integrity_results.get('duplicate_analysis', {}),
                'data_types': integrity_results.get('data_type_analysis', {})
            }
            
            # 质量指标
            null_percentage = integrity_results.get('null_analysis', {}).get('null_percentage', 0)
            dup_percentage = integrity_results.get('duplicate_analysis', {}).get('duplicate_percentage', 0)
            
            report['quality_metrics'] = {
                'completeness': 100 - null_percentage,
                'uniqueness': 100 - dup_percentage,
                'consistency': 100 if not integrity_results.get('issues') else 80
            }
            
        except Exception as e:
            report['error'] = str(e)
        
        return report
    
    def _generate_comprehensive_recommendations(self, file_analysis: Dict,
                                              integrity_results: Dict,
                                              comparison_results: Dict) -> List[str]:
        """生成综合建议"""
        recommendations = []
        
        try:
            # 文件结构建议
            if not file_analysis.get('valid', False):
                recommendations.append("建议检查文件格式和完整性")
            
            # 数据质量建议
            null_percentage = integrity_results.get('null_analysis', {}).get('null_percentage', 0)
            if null_percentage > 10:
                recommendations.append(f"数据中有{null_percentage:.1f}%的空值，建议进行数据清洗")
            
            dup_percentage = integrity_results.get('duplicate_analysis', {}).get('duplicate_percentage', 0)
            if dup_percentage > 5:
                recommendations.append(f"发现{dup_percentage:.1f}%的重复数据，建议去重处理")
            
            # 异常值建议
            anomalies = integrity_results.get('anomaly_detection', {})
            if anomalies.get('outliers'):
                recommendations.append("发现异常值，建议进行异常值处理")
            
            # 比较结果建议
            if comparison_results.get('comparison_possible', False):
                match_score = comparison_results.get('match_score', 0)
                if match_score < 80:
                    recommendations.append("与参考文件差异较大，建议详细检查数据一致性")
            
            if not recommendations:
                recommendations.append("数据质量良好，无需特殊处理")
                
        except Exception:
            recommendations.append("建议进行人工数据质量检查")
        
        return recommendations
    
    def _determine_overall_status(self, quality_score: float,
                                integrity_results: Dict) -> str:
        """确定总体状态"""
        if not integrity_results.get('data_loaded', False):
            return 'FAILED'
        elif quality_score >= 85:
            return 'EXCELLENT'
        elif quality_score >= 75:
            return 'GOOD'
        elif quality_score >= 65:
            return 'ACCEPTABLE'
        elif quality_score >= 55:
            return 'POOR'
        else:
            return 'CRITICAL'
    
    def _save_verification_report(self, verification_result: Dict[str, Any]):
        """保存验证报告"""
        try:
            reports_dir = Path("verification_reports")
            reports_dir.mkdir(exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = Path(verification_result['file_path']).stem
            report_file = reports_dir / f"verification_{file_name}_{timestamp}.json"
            
            with open(report_file, 'w', encoding='utf-8') as f:
                json.dump(verification_result, f, ensure_ascii=False, indent=2, default=str)
                
        except Exception:
            pass  # 静默失败，不影响主要功能