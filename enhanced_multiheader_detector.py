#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
增强的多级列头检测器
提供更准确的多级列头识别和最优参数推荐
系统性强化版本 - 提升复杂表格数据的列头确认准确性和效率
"""

import pandas as pd
import openpyxl
from typing import Dict, Any, List, Optional, Tuple, Union
import re
from collections import Counter
import logging
import hashlib
import os
import time
from datetime import datetime

class EnhancedMultiHeaderDetector:
    """增强的多级列头检测器 - 强化版"""
    
    # 缓存目录
    CACHE_DIR = ".excel_analysis_cache"
    
    # 常见列名模式（语义分析）
    SEMANTIC_PATTERNS = {
        'person_info': [r'姓名', r'name', r'员工', r'人员', r'客户'],
        'time_info': [r'日期', r'时间', r'date', r'time', r'年', r'月', r'day'],
        'financial': [r'金额', r'价格', r'费用', r'成本', r'收入', r'amount', r'price', r'cost'],
        'location': [r'地址', r'城市', r'省份', r'区域', r'address', r'city', r'region'],
        'contact': [r'电话', r'邮箱', r'联系', r'phone', r'email', r'contact'],
        'category': [r'类型', r'分类', r'类别', r'category', r'type', r'class'],
        'quantity': [r'数量', r'个数', r'count', r'quantity', r'num'],
        'status': [r'状态', r'情况', r'status', r'state', r'condition']
    }
    
    # 表格结构模式
    TABLE_PATTERNS = {
        'summary_table': {'has_totals': True, 'numeric_heavy': True},
        'detail_table': {'has_totals': False, 'text_heavy': True},
        'cross_table': {'has_categories': True, 'matrix_like': True},
        'report_table': {'has_headers': True, 'structured': True}
    }
    
    def __init__(self, file_path: str, sheet_name: Optional[str] = None):
        """
        初始化检测器
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = None
        self.sheet = None
        self.raw_data = []
        self.merged_cells_info = []
        self.cache_key = self._generate_cache_key()
        
        # 确保缓存目录存在
        if not os.path.exists(self.CACHE_DIR):
            os.makedirs(self.CACHE_DIR)
    
    def _generate_cache_key(self) -> str:
        """生成缓存键"""
        try:
            stat = os.stat(self.file_path)
            key_data = f"{self.file_path}_{stat.st_mtime}_{stat.st_size}_{self.sheet_name}"
            return hashlib.md5(key_data.encode()).hexdigest()
        except:
            return hashlib.md5(f"{self.file_path}_{self.sheet_name}_{time.time()}".encode()).hexdigest()
    
    def _load_cache(self) -> Optional[Dict[str, Any]]:
        """加载缓存结果"""
        cache_file = os.path.join(self.CACHE_DIR, f"{self.cache_key}.json")
        try:
            if os.path.exists(cache_file):
                import json
                with open(cache_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except:
            pass
        return None
    
    def _save_cache(self, result: Dict[str, Any]) -> None:
        """保存缓存结果"""
        cache_file = os.path.join(self.CACHE_DIR, f"{self.cache_key}.json")
        try:
            import json
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2)
        except:
            pass
        
    def __enter__(self):
        """上下文管理器入口"""
        self.workbook = openpyxl.load_workbook(self.file_path, read_only=True)
        self.sheet = self.workbook[self.sheet_name] if self.sheet_name else self.workbook.active
        return self
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        if self.workbook:
            self.workbook.close()
    
    def analyze_semantic_patterns(self, text: str) -> Dict[str, float]:
        """
        分析文本的语义模式
        
        Args:
            text: 要分析的文本
            
        Returns:
            语义模式匹配分数
        """
        if not text or not isinstance(text, str):
            return {}
        
        text_lower = text.lower().strip()
        pattern_scores = {}
        
        for pattern_name, patterns in self.SEMANTIC_PATTERNS.items():
            score = 0.0
            for pattern in patterns:
                if re.search(pattern, text_lower, re.IGNORECASE):
                    score = max(score, 1.0)
                elif pattern.lower() in text_lower:
                    score = max(score, 0.8)
            pattern_scores[pattern_name] = score
        
        return pattern_scores
    
    def detect_table_structure_type(self, data: List[List[Any]]) -> Dict[str, Any]:
        """
        检测表格结构类型
        
        Args:
            data: 原始数据矩阵
            
        Returns:
            表格结构类型分析结果
        """
        if not data or len(data) < 2:
            return {'type': 'unknown', 'confidence': 0.0}
        
        # 分析数据特征
        total_cells = sum(len(row) for row in data)
        numeric_cells = 0
        text_cells = 0
        empty_cells = 0
        
        for row in data:
            for cell in row:
                if cell is None or str(cell).strip() == '':
                    empty_cells += 1
                elif isinstance(cell, (int, float)) or (isinstance(cell, str) and cell.replace('.', '').replace('-', '').isdigit()):
                    numeric_cells += 1
                else:
                    text_cells += 1
        
        if total_cells == 0:
            return {'type': 'empty', 'confidence': 1.0}
        
        numeric_ratio = numeric_cells / total_cells
        text_ratio = text_cells / total_cells
        
        # 检测是否有汇总行（最后几行数字较多）
        has_totals = False
        if len(data) >= 3:
            last_rows = data[-2:]
            last_numeric = sum(1 for row in last_rows for cell in row 
                             if isinstance(cell, (int, float)) or 
                             (isinstance(cell, str) and cell.replace('.', '').replace('-', '').isdigit()))
            last_total = sum(len(row) for row in last_rows)
            if last_total > 0 and last_numeric / last_total > 0.7:
                has_totals = True
        
        # 判断表格类型
        if numeric_ratio > 0.6 and has_totals:
            return {'type': 'summary_table', 'confidence': 0.8, 'numeric_ratio': numeric_ratio}
        elif text_ratio > 0.6:
            return {'type': 'detail_table', 'confidence': 0.7, 'text_ratio': text_ratio}
        elif 0.3 < numeric_ratio < 0.7:
            return {'type': 'mixed_table', 'confidence': 0.6, 'mixed_ratio': min(numeric_ratio, text_ratio)}
        else:
            return {'type': 'unknown', 'confidence': 0.3}

    def __exit__(self, exc_type, exc_val, exc_tb):
        """上下文管理器出口"""
        if self.workbook:
            self.workbook.close()
    
    def analyze_merged_cells(self) -> List[Dict[str, Any]]:
        """
        分析合并单元格信息
        
        Returns:
            合并单元格信息列表
        """
        merged_info = []
        
        # 重新打开工作簿以获取合并单元格信息
        temp_workbook = openpyxl.load_workbook(self.file_path)
        temp_sheet = temp_workbook[self.sheet_name] if self.sheet_name else temp_workbook.active
        
        for merged_range in temp_sheet.merged_cells.ranges:
            merged_info.append({
                'range': str(merged_range),
                'min_row': merged_range.min_row,
                'max_row': merged_range.max_row,
                'min_col': merged_range.min_col,
                'max_col': merged_range.max_col,
                'span_rows': merged_range.max_row - merged_range.min_row + 1,
                'span_cols': merged_range.max_col - merged_range.min_col + 1
            })
        
        temp_workbook.close()
        return merged_info
    
    def extract_raw_data(self, max_rows: int = 20, max_cols: int = 30) -> List[List[Any]]:
        """
        提取原始数据（优化版 - 智能采样）
        
        Args:
            max_rows: 最大行数（增加到20行以获得更好的分析效果）
            max_cols: 最大列数
            
        Returns:
            原始数据矩阵
        """
        raw_data = []
        actual_max_row = min(max_rows, self.sheet.max_row)
        actual_max_col = min(max_cols, self.sheet.max_column)
        
        for row_idx in range(1, actual_max_row + 1):
            row_data = []
            for col_idx in range(1, actual_max_col + 1):
                cell = self.sheet.cell(row=row_idx, column=col_idx)
                row_data.append(cell.value)
            raw_data.append(row_data)
        
        return raw_data
    
    def find_empty_rows(self, data: List[List[Any]]) -> List[int]:
        """
        查找空行（强化版）
        
        Args:
            data: 原始数据矩阵
            
        Returns:
            空行索引列表
        """
        empty_rows = []
        
        for i, row in enumerate(data):
            if not row:  # 完全空的行
                empty_rows.append(i)
                continue
                
            # 检查是否所有单元格都为空或只包含空白字符
            non_empty_count = 0
            for cell in row:
                if cell is not None and str(cell).strip():
                    non_empty_count += 1
            
            # 如果非空单元格少于总单元格的10%，认为是空行
            if non_empty_count / len(row) < 0.1:
                empty_rows.append(i)
        
        return empty_rows
    
    def analyze_row_content_enhanced(self, row: List[Any]) -> Dict[str, Any]:
        """
        增强的行内容分析
        
        Args:
            row: 行数据
            
        Returns:
            行内容特征分析结果
        """
        if not row:
            return {
                'non_empty_count': 0,
                'unique_count': 0,
                'numeric_count': 0,
                'text_count': 0,
                'pattern_diversity': 0.0,
                'semantic_scores': {},
                'is_likely_header': False,
                'header_confidence': 0.0
            }
        
        non_empty_cells = [cell for cell in row if cell is not None and str(cell).strip()]
        non_empty_count = len(non_empty_cells)
        
        if non_empty_count == 0:
            return {
                'non_empty_count': 0,
                'unique_count': 0,
                'numeric_count': 0,
                'text_count': 0,
                'pattern_diversity': 0.0,
                'semantic_scores': {},
                'is_likely_header': False,
                'header_confidence': 0.0
            }
        
        # 基本统计
        unique_values = list(set(str(cell).strip() for cell in non_empty_cells))
        unique_count = len(unique_values)
        
        numeric_count = 0
        text_count = 0
        
        for cell in non_empty_cells:
            if isinstance(cell, (int, float)):
                numeric_count += 1
            elif isinstance(cell, str):
                if cell.replace('.', '').replace('-', '').replace(',', '').isdigit():
                    numeric_count += 1
                else:
                    text_count += 1
            else:
                text_count += 1
        
        # 模式多样性（唯一值比例）
        pattern_diversity = unique_count / non_empty_count if non_empty_count > 0 else 0.0
        
        # 语义分析
        combined_text = ' '.join(str(cell) for cell in non_empty_cells)
        semantic_scores = self.analyze_semantic_patterns(combined_text)
        
        # 判断是否可能是标题行
        is_likely_header = False
        header_confidence = 0.0
        
        # 标题行特征：
        # 1. 文本比例高
        # 2. 唯一值比例高
        # 3. 包含语义模式
        # 4. 长度适中的文本
        
        text_ratio = text_count / non_empty_count if non_empty_count > 0 else 0.0
        avg_text_length = sum(len(str(cell)) for cell in non_empty_cells) / non_empty_count
        
        # 计算标题置信度
        confidence_factors = []
        
        # 文本比例因子
        if text_ratio > 0.7:
            confidence_factors.append(0.3)
        elif text_ratio > 0.5:
            confidence_factors.append(0.2)
        
        # 唯一性因子
        if pattern_diversity > 0.8:
            confidence_factors.append(0.25)
        elif pattern_diversity > 0.6:
            confidence_factors.append(0.15)
        
        # 语义因子
        max_semantic_score = max(semantic_scores.values()) if semantic_scores else 0.0
        if max_semantic_score > 0.8:
            confidence_factors.append(0.2)
        elif max_semantic_score > 0.5:
            confidence_factors.append(0.1)
        
        # 文本长度因子（标题通常不会太长也不会太短）
        if 3 <= avg_text_length <= 15:
            confidence_factors.append(0.15)
        elif 2 <= avg_text_length <= 20:
            confidence_factors.append(0.1)
        
        # 非空比例因子
        non_empty_ratio = non_empty_count / len(row)
        if non_empty_ratio > 0.5:
            confidence_factors.append(0.1)
        
        header_confidence = sum(confidence_factors)
        is_likely_header = header_confidence > 0.4
        
        return {
            'non_empty_count': non_empty_count,
            'unique_count': unique_count,
            'numeric_count': numeric_count,
            'text_count': text_count,
            'pattern_diversity': pattern_diversity,
            'semantic_scores': semantic_scores,
            'is_likely_header': is_likely_header,
            'header_confidence': header_confidence,
            'text_ratio': text_ratio,
            'avg_text_length': avg_text_length,
            'non_empty_ratio': non_empty_ratio
        }
    
    def detect_header_candidates_enhanced(self, raw_data: List[List[Any]]) -> List[Dict[str, Any]]:
        """
        增强的标题行候选检测
        
        Args:
            raw_data: 原始数据
            
        Returns:
            标题行候选列表
        """
        candidates = []
        
        for i, row in enumerate(raw_data[:15]):  # 检查前15行
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue  # 跳过空行
            
            # 使用增强的行内容分析
            analysis = self.analyze_row_content_enhanced(row, i)
            
            # 计算标题行置信度
            title_confidence = analysis['title_confidence']
            
            if title_confidence > 0.3:  # 降低阈值以包含更多候选
                candidates.append({
                    'row_index': i,
                    'confidence': title_confidence,
                    'analysis': analysis
                })
        
        # 按置信度排序
        candidates.sort(key=lambda x: x['confidence'], reverse=True)
        
        return candidates[:5]  # 返回前5个最佳候选
    
    def _calculate_enhanced_hierarchy_confidence(self, upper_candidate: Dict[str, Any], 
                                               lower_candidate: Dict[str, Any], 
                                               raw_data: List[List[Any]]) -> float:
        """
        计算增强的层次关系置信度
        
        Args:
            upper_candidate: 上级标题候选
            lower_candidate: 下级标题候选
            raw_data: 原始数据
            
        Returns:
            层次关系置信度
        """
        upper_analysis = upper_candidate['analysis']
        lower_analysis = lower_candidate['analysis']
        
        # 多维度置信度计算
        confidence_factors = {
            'text_length_hierarchy': 0.0,
            'unique_count_hierarchy': 0.0,
            'semantic_relationship': 0.0,
            'position_appropriateness': 0.0,
            'content_complexity_progression': 0.0
        }
        
        # 1. 文本长度层次关系
        upper_avg_len = upper_analysis.get('avg_text_length', 0)
        lower_avg_len = lower_analysis.get('avg_text_length', 0)
        
        if 2 <= upper_avg_len <= 15 and lower_avg_len > upper_avg_len:
            confidence_factors['text_length_hierarchy'] = min(0.8, (lower_avg_len - upper_avg_len) / 10)
        
        # 2. 唯一值数量层次关系
        upper_unique = upper_analysis.get('unique_count', 0)
        lower_unique = lower_analysis.get('unique_count', 0)
        
        if lower_unique > upper_unique and upper_unique > 0:
            ratio = min(3.0, lower_unique / upper_unique)
            confidence_factors['unique_count_hierarchy'] = min(0.9, ratio / 3.0)
        
        # 3. 语义关系分析
        upper_semantic = upper_analysis.get('semantic_scores', {})
        lower_semantic = lower_analysis.get('semantic_scores', {})
        
        semantic_correlation = 0.0
        common_patterns = 0
        
        for pattern in upper_semantic:
            if pattern in lower_semantic:
                upper_score = upper_semantic[pattern]
                lower_score = lower_semantic[pattern]
                if upper_score > 0.3 and lower_score > 0.3:
                    common_patterns += 1
                    semantic_correlation += abs(upper_score - lower_score)
        
        if common_patterns > 0:
            # 相似但有差异的语义模式表明层次关系
            avg_diff = semantic_correlation / common_patterns
            confidence_factors['semantic_relationship'] = min(0.7, avg_diff)
        
        # 4. 位置合适性
        upper_row = upper_candidate['row_index']
        lower_row = lower_candidate['row_index']
        row_gap = lower_row - upper_row
        
        if 1 <= row_gap <= 3:
            confidence_factors['position_appropriateness'] = 0.8 - (row_gap - 1) * 0.2
        
        # 5. 内容复杂度递进
        upper_complexity = upper_analysis.get('pattern_diversity', 0)
        lower_complexity = lower_analysis.get('pattern_diversity', 0)
        
        if lower_complexity > upper_complexity:
            complexity_diff = lower_complexity - upper_complexity
            confidence_factors['content_complexity_progression'] = min(0.6, complexity_diff)
        
        # 加权计算最终置信度
        weights = {
            'text_length_hierarchy': 0.25,
            'unique_count_hierarchy': 0.30,
            'semantic_relationship': 0.20,
            'position_appropriateness': 0.15,
            'content_complexity_progression': 0.10
        }
        
        final_confidence = sum(
            confidence_factors[factor] * weights[factor] 
            for factor in confidence_factors
        )
        
        return min(1.0, final_confidence)
    
    def detect_empty_rows(self, raw_data: List[List[Any]]) -> List[int]:
        """
        检测空行（保持向后兼容）
        
        Args:
            raw_data: 原始数据
            
        Returns:
            空行索引列表
        """
        return self.find_empty_rows(raw_data)
    
    def detect_header_candidates(self, raw_data: List[List[Any]]) -> List[Dict[str, Any]]:
        """
        检测标题行候选（保持向后兼容）
        
        Args:
            raw_data: 原始数据
            
        Returns:
            标题行候选列表
        """
        return self.detect_header_candidates_enhanced(raw_data)
    
    def detect_multi_level_structure(self, header_candidates: List[Dict[str, Any]], 
                                   merged_cells: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        检测多级结构（保持向后兼容）
        
        Args:
            header_candidates: 标题行候选
            merged_cells: 合并单元格信息
            
        Returns:
            多级结构检测结果
        """
        enhanced_result = self.detect_multi_level_structure_enhanced(header_candidates, merged_cells)
        
        # 转换为旧格式以保持兼容性
        legacy_result = {
            'is_multi_level': enhanced_result['is_multi_level'],
            'confidence': enhanced_result['confidence'],
            'structure_type': enhanced_result['structure_type'],
            'recommended_header': enhanced_result['recommended_header']
        }
        
        # 添加旧格式的header_rows字段
        if enhanced_result['is_multi_level'] and isinstance(enhanced_result['recommended_header'], list):
            legacy_result['header_rows'] = enhanced_result['recommended_header']
        elif enhanced_result['is_multi_level']:
            legacy_result['header_rows'] = [0, enhanced_result['recommended_header']]
        else:
            legacy_result['header_rows'] = [enhanced_result['recommended_header']]
        
        return legacy_result

    def suggest_optimal_parameters(self) -> Dict[str, Any]:
        """
        建议最优的Excel读取参数（强化版）
        
        Returns:
            包含建议参数的字典
        """
        # 检查缓存
        cached_result = self._load_cache()
        if cached_result:
            return cached_result
        
        try:
            # 提取数据和分析合并单元格
            self.raw_data = self.extract_raw_data()
            self.merged_cells_info = self.analyze_merged_cells()
            
            # 检测表格结构类型
            table_structure = self.detect_table_structure_type(self.raw_data)
            
            # 查找空行
            empty_rows = self.find_empty_rows(self.raw_data)
            
            # 检测标题行候选
            header_candidates = self.detect_header_candidates_enhanced(self.raw_data)
            
            # 检测多级结构
            multi_level_result = self.detect_multi_level_structure_enhanced(
                header_candidates, self.merged_cells_info
            )
            
            # 构建建议参数
            suggestions = {
                'skiprows': None,
                'header': None,
                'is_multi_level': multi_level_result['is_multi_level'],
                'confidence': multi_level_result['confidence'],
                'structure_type': multi_level_result['structure_type'],
                'table_structure': table_structure,
                'empty_rows': empty_rows,
                'header_candidates': header_candidates,
                'merged_cells_count': len(self.merged_cells_info),
                'analysis_details': multi_level_result.get('analysis_details', ''),
                'warnings': [],
                'tips': []
            }
            
            # 根据检测结果调整参数
            if multi_level_result['is_multi_level']:
                header_rows = multi_level_result['recommended_header']
                if isinstance(header_rows, list) and len(header_rows) >= 2:
                    suggestions['header'] = header_rows
                    suggestions['tips'].append("检测到多级列头结构，将创建MultiIndex")
                else:
                    suggestions['header'] = 0
                    suggestions['warnings'].append("多级列头检测不确定，回退到单级模式")
            else:
                # 单级列头
                if isinstance(multi_level_result['recommended_header'], list):
                    suggestions['header'] = multi_level_result['recommended_header'][-1]
                else:
                    suggestions['header'] = multi_level_result['recommended_header']
            
            # 处理空行
            if empty_rows:
                first_data_row = None
                for i, row in enumerate(self.raw_data):
                    if i not in empty_rows:
                        first_data_row = i
                        break
                
                if first_data_row is not None and first_data_row > 0:
                    suggestions['skiprows'] = list(range(first_data_row))
                    suggestions['tips'].append(f"建议跳过前{first_data_row}行空行")
            
            # 合并单元格警告
            if self.merged_cells_info:
                suggestions['warnings'].append(
                    f"检测到{len(self.merged_cells_info)}个合并单元格，可能影响数据读取"
                )
            
            # 表格结构提示
            if table_structure['type'] != 'unknown':
                suggestions['tips'].append(
                    f"检测到表格类型：{table_structure['type']}，置信度：{table_structure['confidence']:.2f}"
                )
            
            # 验证参数
            validation_result = self._validate_parameters_enhanced(suggestions)
            suggestions.update(validation_result)
            
            # 保存缓存
            self._save_cache(suggestions)
            
            return suggestions
            
        except Exception as e:
            logging.error(f"Enhanced parameter suggestion failed: {e}")
            # 回退到基本建议
            return self._fallback_suggest_parameters()
    
    def _validate_parameters_enhanced(self, suggestions: Dict[str, Any]) -> Dict[str, Any]:
        """
        增强的参数验证
        
        Args:
            suggestions: 建议的参数
            
        Returns:
            验证结果和调整后的参数
        """
        validation_result = {
            'validation_passed': False,
            'validation_details': [],
            'adjusted_params': {}
        }
        
        try:
            # 尝试使用建议的参数读取少量数据
            test_params = {
                'skiprows': suggestions.get('skiprows'),
                'header': suggestions.get('header'),
                'nrows': 5  # 只读取前5行进行验证
            }
            
            # 移除None值
            test_params = {k: v for k, v in test_params.items() if v is not None}
            
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, **test_params)
            
            # 验证结果质量
            validation_checks = {
                'has_data': len(df) > 0,
                'has_columns': len(df.columns) > 0,
                'no_unnamed_columns': not any('Unnamed' in str(col) for col in df.columns),
                'reasonable_column_count': 1 <= len(df.columns) <= 50
            }
            
            passed_checks = sum(validation_checks.values())
            total_checks = len(validation_checks)
            
            validation_result['validation_passed'] = passed_checks >= total_checks * 0.75
            validation_result['validation_details'] = validation_checks
            validation_result['validation_score'] = passed_checks / total_checks
            
            # 如果验证失败，尝试调整参数
            if not validation_result['validation_passed']:
                adjusted_params = self._adjust_parameters_based_on_validation(
                    suggestions, validation_checks
                )
                validation_result['adjusted_params'] = adjusted_params
            
        except Exception as e:
            validation_result['validation_details'].append(f"Validation error: {str(e)}")
            validation_result['validation_passed'] = False
        
        return validation_result
    
    def _adjust_parameters_based_on_validation(self, original_params: Dict[str, Any], 
                                             validation_checks: Dict[str, bool]) -> Dict[str, Any]:
        """基于验证结果调整参数"""
        adjusted = original_params.copy()
        
        # 如果有未命名列，尝试调整header参数
        if not validation_checks.get('no_unnamed_columns', True):
            current_header = adjusted.get('header', 0)
            if isinstance(current_header, int):
                adjusted['header'] = current_header + 1
            elif isinstance(current_header, list):
                adjusted['header'] = [h + 1 for h in current_header]
        
        # 如果没有数据，尝试减少skiprows
        if not validation_checks.get('has_data', True):
            current_skiprows = adjusted.get('skiprows')
            if current_skiprows:
                if isinstance(current_skiprows, list) and len(current_skiprows) > 1:
                    adjusted['skiprows'] = current_skiprows[:-1]
                elif isinstance(current_skiprows, int) and current_skiprows > 0:
                    adjusted['skiprows'] = current_skiprows - 1
                else:
                    adjusted['skiprows'] = None
        
        return adjusted
    
    def _fallback_suggest_parameters(self) -> Dict[str, Any]:
        """回退参数建议"""
        return {
            'skiprows': None,
            'header': 0,
            'is_multi_level': False,
            'confidence': 0.0,
            'structure_type': 'fallback',
            'warnings': ['使用回退参数建议'],
            'tips': ['建议手动检查Excel文件结构']
        }

def enhanced_suggest_excel_read_parameters(file_path: str, sheet_name: str = None) -> Dict[str, Any]:
    """
    增强的Excel读取参数建议函数
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        
    Returns:
        参数建议结果
    """
    try:
        with EnhancedMultiHeaderDetector(file_path, sheet_name) as detector:
            return detector.suggest_optimal_parameters()
    except Exception as e:
        # 回退到基本参数
        return {
            "recommended_params": {"header": 0},
            "analysis": {
                "multi_level_header_detected": False,
                "error": str(e)
            },
            "warnings": [f"增强检测失败，使用默认参数: {str(e)}"],
            "tips": ["建议手动检查文件格式"]
        }