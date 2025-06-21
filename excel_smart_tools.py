#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel智能工具模块
提供Excel文件结构检测、参数建议和读取模板创建功能
"""

import pandas as pd
import numpy as np
import openpyxl
from typing import Dict, List, Any, Optional, Tuple, Union
import os
from pathlib import Path
import chardet
import warnings
from openpyxl import load_workbook

class ExcelStructureDetector:
    """Excel文件结构检测器"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.structure_info = {}
    
    def detect_structure(self) -> Dict[str, Any]:
        """检测Excel文件结构"""
        try:
            self.workbook = load_workbook(self.file_path, read_only=True, data_only=True)
            
            structure = {
                'file_path': self.file_path,
                'file_size': os.path.getsize(self.file_path),
                'sheet_count': len(self.workbook.sheetnames),
                'sheet_names': self.workbook.sheetnames,
                'sheets_info': {},
                'recommended_params': {}
            }
            
            # 分析每个工作表
            for sheet_name in self.workbook.sheetnames:
                sheet_info = self._analyze_sheet(sheet_name)
                structure['sheets_info'][sheet_name] = sheet_info
            
            # 生成推荐参数
            structure['recommended_params'] = self._generate_recommended_params(structure)
            
            self.structure_info = structure
            return structure
            
        except Exception as e:
            return {
                'error': f'文件结构检测失败: {str(e)}',
                'file_path': self.file_path
            }
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _analyze_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """分析单个工作表"""
        try:
            sheet = self.workbook[sheet_name]
            
            # 获取数据范围
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            # 检测标题行
            header_row = self._detect_header_row(sheet)
            
            # 检测数据类型
            data_types = self._detect_data_types(sheet, header_row)
            
            # 检测合并单元格
            merged_cells = list(sheet.merged_cells.ranges)
            
            return {
                'max_row': max_row,
                'max_column': max_col,
                'header_row': header_row,
                'data_start_row': header_row + 1 if header_row else 1,
                'data_types': data_types,
                'merged_cells_count': len(merged_cells),
                'has_merged_cells': len(merged_cells) > 0,
                'estimated_data_rows': max_row - (header_row or 0),
                'column_names': self._extract_column_names(sheet, header_row)
            }
            
        except Exception as e:
            return {'error': f'工作表分析失败: {str(e)}'}
    
    def _detect_header_row(self, sheet, max_check_rows: int = 10) -> Optional[int]:
        """检测标题行位置"""
        for row_idx in range(1, min(max_check_rows + 1, sheet.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(sheet.max_column + 1, 20)):  # 检查前20列
                cell_value = sheet.cell(row_idx, col_idx).value
                if cell_value is not None:
                    row_values.append(str(cell_value).strip())
            
            # 如果这一行有多个非空值且看起来像标题
            if len(row_values) >= 2:
                # 检查是否包含常见的标题特征
                header_indicators = ['名称', '姓名', 'name', 'id', '编号', '日期', 'date', '金额', 'amount']
                if any(indicator in ''.join(row_values).lower() for indicator in header_indicators):
                    return row_idx
                
                # 如果下一行的数据类型与当前行不同，当前行可能是标题
                if row_idx < sheet.max_row:
                    next_row_values = []
                    for col_idx in range(1, min(sheet.max_column + 1, len(row_values) + 1)):
                        next_cell_value = sheet.cell(row_idx + 1, col_idx).value
                        if next_cell_value is not None:
                            next_row_values.append(next_cell_value)
                    
                    if self._looks_like_header_vs_data(row_values, next_row_values):
                        return row_idx
        
        return None
    
    def _looks_like_header_vs_data(self, header_row: List[str], data_row: List[Any]) -> bool:
        """判断是否像标题行vs数据行"""
        if not header_row or not data_row:
            return False
        
        # 标题行通常是字符串，数据行可能包含数字
        header_strings = sum(1 for val in header_row if isinstance(val, str) and not val.isdigit())
        data_numbers = sum(1 for val in data_row if isinstance(val, (int, float)) or 
                          (isinstance(val, str) and val.replace('.', '').replace('-', '').isdigit()))
        
        return header_strings >= len(header_row) * 0.7 and data_numbers >= len(data_row) * 0.3
    
    def _detect_data_types(self, sheet, header_row: Optional[int]) -> Dict[int, str]:
        """检测每列的数据类型"""
        data_types = {}
        start_row = (header_row + 1) if header_row else 1
        sample_size = min(100, sheet.max_row - start_row + 1)  # 采样前100行
        
        for col_idx in range(1, sheet.max_column + 1):
            col_values = []
            for row_idx in range(start_row, start_row + sample_size):
                if row_idx <= sheet.max_row:
                    cell_value = sheet.cell(row_idx, col_idx).value
                    if cell_value is not None:
                        col_values.append(cell_value)
            
            if col_values:
                data_types[col_idx] = self._infer_column_type(col_values)
            else:
                data_types[col_idx] = 'empty'
        
        return data_types
    
    def _infer_column_type(self, values: List[Any]) -> str:
        """推断列的数据类型"""
        if not values:
            return 'empty'
        
        type_counts = {
            'int': 0,
            'float': 0,
            'datetime': 0,
            'string': 0,
            'boolean': 0
        }
        
        for value in values:
            if isinstance(value, bool):
                type_counts['boolean'] += 1
            elif isinstance(value, int):
                type_counts['int'] += 1
            elif isinstance(value, float):
                type_counts['float'] += 1
            elif hasattr(value, 'date'):  # datetime objects
                type_counts['datetime'] += 1
            else:
                # 尝试解析字符串
                str_val = str(value).strip()
                if str_val.replace('.', '').replace('-', '').isdigit():
                    if '.' in str_val:
                        type_counts['float'] += 1
                    else:
                        type_counts['int'] += 1
                else:
                    type_counts['string'] += 1
        
        # 返回最常见的类型
        return max(type_counts, key=type_counts.get)
    
    def _extract_column_names(self, sheet, header_row: Optional[int]) -> List[str]:
        """提取列名"""
        if not header_row:
            return [f'Column_{i}' for i in range(1, sheet.max_column + 1)]
        
        column_names = []
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(header_row, col_idx).value
            if cell_value is not None:
                column_names.append(str(cell_value).strip())
            else:
                column_names.append(f'Column_{col_idx}')
        
        return column_names
    
    def _generate_recommended_params(self, structure: Dict[str, Any]) -> Dict[str, Any]:
        """生成推荐的读取参数"""
        if not structure.get('sheets_info'):
            return {}
        
        # 选择最可能的主要工作表（通常是第一个或数据最多的）
        main_sheet = None
        max_data_rows = 0
        
        for sheet_name, sheet_info in structure['sheets_info'].items():
            if 'error' not in sheet_info:
                data_rows = sheet_info.get('estimated_data_rows', 0)
                if data_rows > max_data_rows:
                    max_data_rows = data_rows
                    main_sheet = sheet_name
        
        if not main_sheet:
            main_sheet = structure['sheet_names'][0]
        
        sheet_info = structure['sheets_info'][main_sheet]
        
        params = {
            'sheet_name': main_sheet,
            'header': sheet_info.get('header_row', 0) - 1 if sheet_info.get('header_row') else None,
            'skiprows': None,
            'usecols': None,
            'dtype': None,
            'parse_dates': [],
            'na_values': ['', ' ', 'NULL', 'null', 'N/A', 'n/a', '#N/A']
        }
        
        # 如果有合并单元格，建议跳过
        if sheet_info.get('has_merged_cells'):
            params['skiprows'] = list(range(0, sheet_info.get('header_row', 1) - 1))
        
        # 根据数据类型设置dtype和parse_dates
        data_types = sheet_info.get('data_types', {})
        column_names = sheet_info.get('column_names', [])
        
        dtype_mapping = {}
        parse_dates_cols = []
        
        for col_idx, col_type in data_types.items():
            col_name = column_names[col_idx - 1] if col_idx <= len(column_names) else f'Column_{col_idx}'
            
            if col_type == 'string':
                dtype_mapping[col_name] = 'str'
            elif col_type == 'int':
                dtype_mapping[col_name] = 'Int64'  # 可空整数类型
            elif col_type == 'float':
                dtype_mapping[col_name] = 'float64'
            elif col_type == 'datetime':
                parse_dates_cols.append(col_name)
            elif col_type == 'boolean':
                dtype_mapping[col_name] = 'boolean'
        
        if dtype_mapping:
            params['dtype'] = dtype_mapping
        if parse_dates_cols:
            params['parse_dates'] = parse_dates_cols
        
        return params

def suggest_excel_read_parameters(file_path: str) -> Dict[str, Any]:
    """建议Excel读取参数"""
    try:
        detector = ExcelStructureDetector(file_path)
        structure = detector.detect_structure()
        
        if 'error' in structure:
            return structure
        
        return {
            'success': True,
            'file_info': {
                'path': file_path,
                'size': structure['file_size'],
                'sheet_count': structure['sheet_count'],
                'sheet_names': structure['sheet_names']
            },
            'recommended_params': structure['recommended_params'],
            'detailed_analysis': structure['sheets_info'],
            'usage_example': _generate_usage_example(structure['recommended_params'])
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f'参数建议生成失败: {str(e)}',
            'file_path': file_path
        }

def detect_excel_file_structure(file_path: str) -> Dict[str, Any]:
    """检测Excel文件结构（别名函数）"""
    return suggest_excel_read_parameters(file_path)

def create_excel_read_template(file_path: str, output_path: Optional[str] = None) -> Dict[str, Any]:
    """创建Excel读取模板代码"""
    try:
        suggestion_result = suggest_excel_read_parameters(file_path)
        
        if not suggestion_result.get('success'):
            return suggestion_result
        
        params = suggestion_result['recommended_params']
        file_info = suggestion_result['file_info']
        
        # 生成模板代码
        template_code = _generate_template_code(file_path, params, file_info)
        
        result = {
            'success': True,
            'template_code': template_code,
            'file_info': file_info,
            'parameters_used': params
        }
        
        # 如果指定了输出路径，保存模板文件
        if output_path:
            try:
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(template_code)
                result['template_saved'] = output_path
            except Exception as e:
                result['save_error'] = f'模板保存失败: {str(e)}'
        
        return result
        
    except Exception as e:
        return {
            'success': False,
            'error': f'模板创建失败: {str(e)}',
            'file_path': file_path
        }

def _generate_usage_example(params: Dict[str, Any]) -> str:
    """生成使用示例代码"""
    code_lines = ["import pandas as pd", ""]
    
    # 构建pd.read_excel调用
    read_excel_params = []
    
    if params.get('sheet_name'):
        read_excel_params.append(f"sheet_name='{params['sheet_name']}'")
    
    if params.get('header') is not None:
        read_excel_params.append(f"header={params['header']}")
    
    if params.get('skiprows'):
        read_excel_params.append(f"skiprows={params['skiprows']}")
    
    if params.get('dtype'):
        read_excel_params.append(f"dtype={params['dtype']}")
    
    if params.get('parse_dates'):
        read_excel_params.append(f"parse_dates={params['parse_dates']}")
    
    if params.get('na_values'):
        read_excel_params.append(f"na_values={params['na_values']}")
    
    params_str = ',\n    '.join(read_excel_params)
    
    code_lines.extend([
        "# 读取Excel文件",
        f"df = pd.read_excel(",
        f"    'your_file_path.xlsx',",
        f"    {params_str}",
        ")",
        "",
        "# 查看数据基本信息",
        "print(df.info())",
        "print(df.head())"
    ])
    
    return '\n'.join(code_lines)

def _generate_template_code(file_path: str, params: Dict[str, Any], file_info: Dict[str, Any]) -> str:
    """生成完整的模板代码"""
    template = f'''#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文件读取模板
自动生成于: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}

文件信息:
- 路径: {file_path}
- 大小: {file_info.get('size', 0)} bytes
- 工作表数量: {file_info.get('sheet_count', 0)}
- 工作表名称: {file_info.get('sheet_names', [])}
"""

import pandas as pd
import numpy as np
from pathlib import Path

def load_excel_data(file_path: str = r"{file_path}"):
    """
    加载Excel数据
    
    Args:
        file_path (str): Excel文件路径
    
    Returns:
        pd.DataFrame: 加载的数据
    """
    try:
        # 检查文件是否存在
        if not Path(file_path).exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        # 读取Excel文件
        df = pd.read_excel(
            file_path,
'''
    
    # 添加参数
    for key, value in params.items():
        if value is not None and key != 'na_values':  # na_values单独处理
            if isinstance(value, str):
                template += f"            {key}='{value}',\n"
            else:
                template += f"            {key}={value},\n"
    
    # 添加na_values
    if params.get('na_values'):
        template += f"            na_values={params['na_values']},\n"
    
    template += '''        )
        
        print(f"数据加载成功: {df.shape[0]} 行, {df.shape[1]} 列")
        return df
        
    except Exception as e:
        print(f"数据加载失败: {str(e)}")
        raise

def analyze_data(df: pd.DataFrame):
    """
    分析数据基本信息
    
    Args:
        df (pd.DataFrame): 要分析的数据
    """
    print("=== 数据基本信息 ===")
    print(f"数据形状: {df.shape}")
    print(f"列名: {list(df.columns)}")
    print("\n=== 数据类型 ===")
    print(df.dtypes)
    print("\n=== 缺失值统计 ===")
    print(df.isnull().sum())
    print("\n=== 前5行数据 ===")
    print(df.head())
    print("\n=== 数据描述统计 ===")
    print(df.describe())

if __name__ == "__main__":
    # 加载数据
    data = load_excel_data()
    
    # 分析数据
    analyze_data(data)
    
    # 在这里添加你的数据处理代码
    # ...
'''
    
    return template

# 向后兼容的别名
detect_excel_structure = detect_excel_file_structure