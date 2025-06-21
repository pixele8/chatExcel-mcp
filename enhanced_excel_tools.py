#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
增强Excel工具模块
提供高级Excel操作和分析功能
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Tuple, Union, Callable
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import xlsxwriter
from datetime import datetime, date
import warnings
from dataclasses import dataclass
from enum import Enum
import io
import base64

class ExcelFormat(Enum):
    """Excel格式类型"""
    XLSX = "xlsx"
    XLS = "xls"
    CSV = "csv"
    TSV = "tsv"

@dataclass
class ExcelStyle:
    """Excel样式配置"""
    font_name: str = "Arial"
    font_size: int = 11
    font_bold: bool = False
    font_color: str = "000000"
    bg_color: Optional[str] = None
    border: bool = False
    alignment: str = "left"  # left, center, right
    number_format: Optional[str] = None

@dataclass
class ChartConfig:
    """图表配置"""
    chart_type: str  # bar, line, pie, scatter
    title: str
    x_axis_title: Optional[str] = None
    y_axis_title: Optional[str] = None
    width: int = 15
    height: int = 10
    position: str = "A1"

class EnhancedExcelTools:
    """增强Excel工具类"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.writer = None
    
    def create_workbook(self, filename: Optional[str] = None) -> Dict[str, Any]:
        """创建新的工作簿"""
        try:
            if filename:
                self.writer = pd.ExcelWriter(filename, engine='xlsxwriter')
                self.workbook = self.writer.book
            else:
                # 创建内存中的工作簿
                self.workbook = xlsxwriter.Workbook(io.BytesIO(), {'in_memory': True})
            
            return {
                'success': True,
                'message': '工作簿创建成功',
                'filename': filename
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'创建工作簿失败: {str(e)}'
            }
    
    def add_worksheet(self, name: str = 'Sheet1') -> Dict[str, Any]:
        """添加工作表"""
        try:
            if not self.workbook:
                return {
                    'success': False,
                    'error': '请先创建工作簿'
                }
            
            self.worksheet = self.workbook.add_worksheet(name)
            
            return {
                'success': True,
                'message': f'工作表 "{name}" 创建成功',
                'worksheet_name': name
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'创建工作表失败: {str(e)}'
            }
    
    def write_dataframe_advanced(self, df: pd.DataFrame, sheet_name: str = 'Sheet1', 
                               start_row: int = 0, start_col: int = 0,
                               style: Optional[ExcelStyle] = None,
                               auto_filter: bool = True,
                               freeze_panes: Optional[Tuple[int, int]] = None) -> Dict[str, Any]:
        """高级DataFrame写入"""
        try:
            if not self.writer:
                return {
                    'success': False,
                    'error': '请先创建工作簿'
                }
            
            # 写入数据
            df.to_excel(self.writer, sheet_name=sheet_name, 
                       startrow=start_row, startcol=start_col, index=False)
            
            # 获取工作表
            worksheet = self.writer.sheets[sheet_name]
            
            # 应用样式
            if style:
                self._apply_style_to_range(worksheet, df, start_row, start_col, style)
            
            # 自动筛选
            if auto_filter:
                worksheet.autofilter(start_row, start_col, 
                                   start_row + len(df), start_col + len(df.columns) - 1)
            
            # 冻结窗格
            if freeze_panes:
                worksheet.freeze_panes(freeze_panes[0], freeze_panes[1])
            
            # 自动调整列宽
            self._auto_adjust_column_width(worksheet, df, start_col)
            
            return {
                'success': True,
                'message': f'数据写入工作表 "{sheet_name}" 成功',
                'rows_written': len(df),
                'columns_written': len(df.columns)
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'写入数据失败: {str(e)}'
            }
    
    def _apply_style_to_range(self, worksheet, df: pd.DataFrame, 
                            start_row: int, start_col: int, style: ExcelStyle):
        """应用样式到范围"""
        # 创建格式
        format_dict = {
            'font_name': style.font_name,
            'font_size': style.font_size,
            'bold': style.font_bold,
            'font_color': style.font_color,
            'align': style.alignment
        }
        
        if style.bg_color:
            format_dict['bg_color'] = style.bg_color
        
        if style.border:
            format_dict['border'] = 1
        
        if style.number_format:
            format_dict['num_format'] = style.number_format
        
        cell_format = self.workbook.add_format(format_dict)
        
        # 应用到标题行
        header_format = self.workbook.add_format({
            **format_dict,
            'bold': True,
            'bg_color': '#D3D3D3' if not style.bg_color else style.bg_color
        })
        
        # 写入标题
        for col, column_name in enumerate(df.columns):
            worksheet.write(start_row, start_col + col, column_name, header_format)
        
        # 写入数据
        for row in range(len(df)):
            for col in range(len(df.columns)):
                worksheet.write(start_row + 1 + row, start_col + col, 
                              df.iloc[row, col], cell_format)
    
    def _auto_adjust_column_width(self, worksheet, df: pd.DataFrame, start_col: int):
        """自动调整列宽"""
        for col, column_name in enumerate(df.columns):
            # 计算列的最大宽度
            max_length = len(str(column_name))
            
            for value in df.iloc[:, col]:
                if pd.notna(value):
                    max_length = max(max_length, len(str(value)))
            
            # 设置列宽（限制最大宽度）
            adjusted_width = min(max_length + 2, 50)
            worksheet.set_column(start_col + col, start_col + col, adjusted_width)
    
    def add_chart(self, df: pd.DataFrame, chart_config: ChartConfig, 
                 data_range: Optional[str] = None) -> Dict[str, Any]:
        """添加图表"""
        try:
            if not self.worksheet:
                return {
                    'success': False,
                    'error': '请先创建工作表'
                }
            
            # 创建图表
            if chart_config.chart_type == 'bar':
                chart = self.workbook.add_chart({'type': 'column'})
            elif chart_config.chart_type == 'line':
                chart = self.workbook.add_chart({'type': 'line'})
            elif chart_config.chart_type == 'pie':
                chart = self.workbook.add_chart({'type': 'pie'})
            else:
                return {
                    'success': False,
                    'error': f'不支持的图表类型: {chart_config.chart_type}'
                }
            
            # 设置数据范围
            if not data_range:
                # 自动确定数据范围
                data_range = f'A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}'
            
            # 添加数据系列
            chart.add_series({
                'categories': f'={data_range}',
                'values': f'={data_range}',
                'name': chart_config.title
            })
            
            # 设置图表属性
            chart.set_title({'name': chart_config.title})
            if chart_config.x_axis_title:
                chart.set_x_axis({'name': chart_config.x_axis_title})
            if chart_config.y_axis_title:
                chart.set_y_axis({'name': chart_config.y_axis_title})
            
            chart.set_size({'width': chart_config.width * 64, 
                          'height': chart_config.height * 20})
            
            # 插入图表
            self.worksheet.insert_chart(chart_config.position, chart)
            
            return {
                'success': True,
                'message': f'{chart_config.chart_type}图表添加成功',
                'chart_position': chart_config.position
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'添加图表失败: {str(e)}'
            }
    
    def add_conditional_formatting(self, range_addr: str, rule_type: str, 
                                 **kwargs) -> Dict[str, Any]:
        """添加条件格式"""
        try:
            if not self.worksheet:
                return {
                    'success': False,
                    'error': '请先创建工作表'
                }
            
            if rule_type == 'color_scale':
                # 颜色刻度
                self.worksheet.conditional_format(range_addr, {
                    'type': '3_color_scale',
                    'min_color': kwargs.get('min_color', '#FF0000'),
                    'mid_color': kwargs.get('mid_color', '#FFFF00'),
                    'max_color': kwargs.get('max_color', '#00FF00')
                })
            elif rule_type == 'data_bar':
                # 数据条
                self.worksheet.conditional_format(range_addr, {
                    'type': 'data_bar',
                    'bar_color': kwargs.get('bar_color', '#0070C0')
                })
            elif rule_type == 'icon_set':
                # 图标集
                self.worksheet.conditional_format(range_addr, {
                    'type': 'icon_set',
                    'icon_style': kwargs.get('icon_style', '3_traffic_lights')
                })
            elif rule_type == 'cell_value':
                # 单元格值
                self.worksheet.conditional_format(range_addr, {
                    'type': 'cell',
                    'criteria': kwargs.get('criteria', '>'),
                    'value': kwargs.get('value', 0),
                    'format': self.workbook.add_format({
                        'bg_color': kwargs.get('bg_color', '#FFC7CE'),
                        'font_color': kwargs.get('font_color', '#9C0006')
                    })
                })
            else:
                return {
                    'success': False,
                    'error': f'不支持的条件格式类型: {rule_type}'
                }
            
            return {
                'success': True,
                'message': f'条件格式 {rule_type} 添加成功',
                'range': range_addr
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'添加条件格式失败: {str(e)}'
            }
    
    def create_pivot_table(self, df: pd.DataFrame, 
                          rows: List[str], 
                          columns: Optional[List[str]] = None,
                          values: Optional[List[str]] = None,
                          aggfunc: str = 'sum') -> Dict[str, Any]:
        """创建数据透视表"""
        try:
            # 创建数据透视表
            pivot_params = {
                'data': df,
                'index': rows,
                'aggfunc': aggfunc
            }
            
            if columns:
                pivot_params['columns'] = columns
            if values:
                pivot_params['values'] = values
            
            pivot_table = pd.pivot_table(**pivot_params)
            
            return {
                'success': True,
                'message': '数据透视表创建成功',
                'pivot_table': pivot_table,
                'shape': pivot_table.shape
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'创建数据透视表失败: {str(e)}'
            }
    
    def add_data_validation(self, range_addr: str, validation_type: str, 
                          **kwargs) -> Dict[str, Any]:
        """添加数据验证"""
        try:
            if not self.worksheet:
                return {
                    'success': False,
                    'error': '请先创建工作表'
                }
            
            validation_dict = {'validate': validation_type}
            
            if validation_type == 'list':
                validation_dict['source'] = kwargs.get('source', [])
            elif validation_type == 'integer':
                validation_dict.update({
                    'criteria': kwargs.get('criteria', 'between'),
                    'minimum': kwargs.get('minimum', 0),
                    'maximum': kwargs.get('maximum', 100)
                })
            elif validation_type == 'decimal':
                validation_dict.update({
                    'criteria': kwargs.get('criteria', 'between'),
                    'minimum': kwargs.get('minimum', 0.0),
                    'maximum': kwargs.get('maximum', 100.0)
                })
            elif validation_type == 'date':
                validation_dict.update({
                    'criteria': kwargs.get('criteria', 'between'),
                    'minimum': kwargs.get('minimum', datetime(2000, 1, 1)),
                    'maximum': kwargs.get('maximum', datetime(2030, 12, 31))
                })
            
            # 添加错误消息
            if 'error_title' in kwargs:
                validation_dict['error_title'] = kwargs['error_title']
            if 'error_message' in kwargs:
                validation_dict['error_message'] = kwargs['error_message']
            
            # 添加输入消息
            if 'input_title' in kwargs:
                validation_dict['input_title'] = kwargs['input_title']
            if 'input_message' in kwargs:
                validation_dict['input_message'] = kwargs['input_message']
            
            self.worksheet.data_validation(range_addr, validation_dict)
            
            return {
                'success': True,
                'message': f'数据验证添加成功',
                'range': range_addr,
                'type': validation_type
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'添加数据验证失败: {str(e)}'
            }
    
    def protect_worksheet(self, password: Optional[str] = None, 
                         options: Optional[Dict[str, bool]] = None) -> Dict[str, Any]:
        """保护工作表"""
        try:
            if not self.worksheet:
                return {
                    'success': False,
                    'error': '请先创建工作表'
                }
            
            protect_options = options or {
                'select_locked_cells': True,
                'select_unlocked_cells': True,
                'format_cells': False,
                'format_columns': False,
                'format_rows': False,
                'insert_columns': False,
                'insert_rows': False,
                'insert_hyperlinks': False,
                'delete_columns': False,
                'delete_rows': False,
                'sort': False,
                'autofilter': False,
                'pivot_tables': False
            }
            
            self.worksheet.protect(password, protect_options)
            
            return {
                'success': True,
                'message': '工作表保护设置成功',
                'password_protected': password is not None
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'保护工作表失败: {str(e)}'
            }
    
    def add_formula(self, cell: str, formula: str) -> Dict[str, Any]:
        """添加公式"""
        try:
            if not self.worksheet:
                return {
                    'success': False,
                    'error': '请先创建工作表'
                }
            
            self.worksheet.write_formula(cell, formula)
            
            return {
                'success': True,
                'message': f'公式添加成功',
                'cell': cell,
                'formula': formula
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'添加公式失败: {str(e)}'
            }
    
    def merge_cells(self, range_addr: str, text: str = '', 
                   style: Optional[ExcelStyle] = None) -> Dict[str, Any]:
        """合并单元格"""
        try:
            if not self.worksheet:
                return {
                    'success': False,
                    'error': '请先创建工作表'
                }
            
            # 解析范围
            start_cell, end_cell = range_addr.split(':')
            
            # 创建格式
            cell_format = None
            if style:
                format_dict = {
                    'font_name': style.font_name,
                    'font_size': style.font_size,
                    'bold': style.font_bold,
                    'font_color': style.font_color,
                    'align': style.alignment,
                    'valign': 'vcenter'
                }
                
                if style.bg_color:
                    format_dict['bg_color'] = style.bg_color
                
                if style.border:
                    format_dict['border'] = 1
                
                cell_format = self.workbook.add_format(format_dict)
            
            # 合并单元格并写入文本
            self.worksheet.merge_range(range_addr, text, cell_format)
            
            return {
                'success': True,
                'message': f'单元格合并成功',
                'range': range_addr,
                'text': text
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'合并单元格失败: {str(e)}'
            }
    
    def save_workbook(self, filename: Optional[str] = None) -> Dict[str, Any]:
        """保存工作簿"""
        try:
            if self.writer:
                self.writer.close()
                return {
                    'success': True,
                    'message': '工作簿保存成功',
                    'filename': filename
                }
            elif self.workbook:
                if filename:
                    self.workbook.close()
                    return {
                        'success': True,
                        'message': '工作簿保存成功',
                        'filename': filename
                    }
                else:
                    # 返回内存中的数据
                    self.workbook.close()
                    return {
                        'success': True,
                        'message': '工作簿创建完成（内存中）'
                    }
            else:
                return {
                    'success': False,
                    'error': '没有可保存的工作簿'
                }
        except Exception as e:
            return {
                'success': False,
                'error': f'保存工作簿失败: {str(e)}'
            }
    
    def read_excel_advanced(self, filename: str, **kwargs) -> Dict[str, Any]:
        """高级Excel读取"""
        try:
            # 默认参数
            read_params = {
                'sheet_name': kwargs.get('sheet_name', 0),
                'header': kwargs.get('header', 0),
                'index_col': kwargs.get('index_col', None),
                'usecols': kwargs.get('usecols', None),
                'skiprows': kwargs.get('skiprows', None),
                'nrows': kwargs.get('nrows', None),
                'na_values': kwargs.get('na_values', None),
                'dtype': kwargs.get('dtype', None),
                'parse_dates': kwargs.get('parse_dates', False),
                'date_parser': kwargs.get('date_parser', None)
            }
            
            # 移除None值
            read_params = {k: v for k, v in read_params.items() if v is not None}
            
            # 读取Excel
            df = pd.read_excel(filename, **read_params)
            
            # 基本信息
            info = {
                'shape': df.shape,
                'columns': df.columns.tolist(),
                'dtypes': df.dtypes.to_dict(),
                'memory_usage': df.memory_usage(deep=True).sum(),
                'null_counts': df.isnull().sum().to_dict()
            }
            
            return {
                'success': True,
                'message': 'Excel文件读取成功',
                'dataframe': df,
                'info': info
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'读取Excel文件失败: {str(e)}'
            }
    
    def get_sheet_names(self, filename: str) -> Dict[str, Any]:
        """获取Excel文件的工作表名称"""
        try:
            excel_file = pd.ExcelFile(filename)
            sheet_names = excel_file.sheet_names
            
            return {
                'success': True,
                'message': '工作表名称获取成功',
                'sheet_names': sheet_names,
                'sheet_count': len(sheet_names)
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'获取工作表名称失败: {str(e)}'
            }

def create_enhanced_excel_tools() -> EnhancedExcelTools:
    """创建增强Excel工具实例"""
    return EnhancedExcelTools()

def create_excel_style(font_name: str = "Arial", font_size: int = 11, 
                      font_bold: bool = False, font_color: str = "000000",
                      bg_color: Optional[str] = None, border: bool = False,
                      alignment: str = "left", number_format: Optional[str] = None) -> ExcelStyle:
    """创建Excel样式"""
    return ExcelStyle(
        font_name=font_name,
        font_size=font_size,
        font_bold=font_bold,
        font_color=font_color,
        bg_color=bg_color,
        border=border,
        alignment=alignment,
        number_format=number_format
    )

def create_chart_config(chart_type: str, title: str, 
                       x_axis_title: Optional[str] = None,
                       y_axis_title: Optional[str] = None,
                       width: int = 15, height: int = 10,
                       position: str = "A1") -> ChartConfig:
    """创建图表配置"""
    return ChartConfig(
        chart_type=chart_type,
        title=title,
        x_axis_title=x_axis_title,
        y_axis_title=y_axis_title,
        width=width,
        height=height,
        position=position
    )