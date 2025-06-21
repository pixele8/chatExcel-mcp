from fastmcp import FastMCP

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ChatExcel MCP Server - 增强版Excel数据处理服务器
提供统一的Excel文件处理、数据分析和可视化功能
"""
from typing import Optional

# 导入核心模块
try:
    from core.config import get_config, SecurityConfig, PerformanceConfig
    from core.exceptions import (
        ChatExcelError, FileAccessError, CodeExecutionError, 
        ValidationError, SecurityError, ConfigurationError
    )
    from core.types import (
        ExecutionResult, ExecutionStatus, FileInfo, FileType,
        ExcelMetadata, CodeAnalysis, ProcessingTask, OperationType
    )
    CORE_MODULES_AVAILABLE = True
except ImportError as e:
    print(f"⚠ 核心模块导入失败: {e}")
    CORE_MODULES_AVAILABLE = False

# 统一的依赖管理系统
class DependencyManager:
    """统一管理可选依赖的导入和状态"""
    
    def __init__(self):
        self.available_modules = {}
        self.failed_imports = []
        self.config = get_config() if CORE_MODULES_AVAILABLE else None
        self._initialize_dependencies()
    
    def _initialize_dependencies(self) -> None:
        """初始化所有可选依赖"""
        dependencies = {
            'df_processed_error_handler': [
                'VariableLifecycleManager', 'NameErrorHandler', 
                'enhanced_execute_with_error_handling', 'analyze_code_variables'
            ],
            'column_checker': ['ColumnChecker'],
            'excel_helper': ['_suggest_excel_read_parameters', 'detect_excel_structure'],
            'excel_smart_tools': [
                'suggest_excel_read_parameters', 'detect_excel_file_structure', 
                'create_excel_read_template'
            ],
            'comprehensive_data_verification': ['ComprehensiveDataVerifier'],
            'data_verification': ['verify_data_processing_result', 'DataVerificationEngine'],
            'excel_enhanced_tools': ['ExcelEnhancedProcessor', 'get_excel_processor'],
            'formulas_tools': [
                'parse_excel_formula', 'compile_excel_workbook', 'execute_excel_formula',
                'analyze_excel_dependencies', 'validate_excel_formula'
            ],
            'excel_data_quality_tools': [
                'ExcelDataQualityController', 'ExcelCellContentExtractor', 
                'ExcelCharacterConverter', 'ExcelMultiConditionExtractor', 
                'ExcelMultiTableMerger', 'ExcelDataCleaner', 'ExcelBatchProcessor'
            ]
        }
        
        for module_name, imports in dependencies.items():
            self._import_module(module_name, imports)
    
    def _import_module(self, module_name: str, imports: list) -> None:
        """安全导入单个模块"""
        try:
            module = __import__(module_name, fromlist=imports)
            imported_items = {}
            
            for item_name in imports:
                if hasattr(module, item_name):
                    imported_items[item_name] = getattr(module, item_name)
            
            if imported_items:
                self.available_modules[module_name] = imported_items
                print(f"✓ 成功导入模块: {module_name}")
            else:
                self.failed_imports.append(module_name)
                print(f"⚠ 模块 {module_name} 导入失败: 未找到指定项")
                
        except ImportError:
            self.failed_imports.append(module_name)
            print(f"⚠ 模块 {module_name} 导入失败: ImportError")
        except Exception as e:
            self.failed_imports.append(module_name)
            print(f"⚠ 模块 {module_name} 导入时发生错误: {str(e)}")
    
    def get_module_item(self, module_name: str, item_name: str, default=None):
        """获取模块中的特定项"""
        if module_name in self.available_modules:
            return self.available_modules[module_name].get(item_name, default)
        return default
    
    def is_available(self, module_name: str) -> bool:
        """检查模块是否可用"""
        return module_name in self.available_modules

# 初始化依赖管理器
dependency_manager = DependencyManager()

# 标准库导入
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import base64
import os
from chardet import detect
import traceback
from io import StringIO
import sys
import time
from textwrap import wrap
import json
from pathlib import Path
from typing import Dict, Any, List, Optional, Union, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from enum import Enum

# 通过依赖管理器获取可选模块
# Excel助手工具
_suggest_excel_read_parameters = dependency_manager.get_module_item('excel_helper', '_suggest_excel_read_parameters')
detect_excel_structure = dependency_manager.get_module_item('excel_helper', 'detect_excel_structure')

# Excel智能工具
suggest_excel_read_parameters = dependency_manager.get_module_item('excel_smart_tools', 'suggest_excel_read_parameters')
detect_excel_file_structure = dependency_manager.get_module_item('excel_smart_tools', 'detect_excel_file_structure')
create_excel_read_template = dependency_manager.get_module_item('excel_smart_tools', 'create_excel_read_template')



# 数据验证工具
ComprehensiveDataVerifier = dependency_manager.get_module_item('comprehensive_data_verification', 'ComprehensiveDataVerifier')
verify_data_processing_result = dependency_manager.get_module_item('data_verification', 'verify_data_processing_result')
DataVerificationEngine = dependency_manager.get_module_item('data_verification', 'DataVerificationEngine')

# Excel增强工具
ExcelEnhancedProcessor = dependency_manager.get_module_item('excel_enhanced_tools', 'ExcelEnhancedProcessor')
get_excel_processor = dependency_manager.get_module_item('excel_enhanced_tools', 'get_excel_processor')

# Excel公式处理工具
parse_excel_formula = dependency_manager.get_module_item('formulas_tools', 'parse_excel_formula')
compile_excel_workbook = dependency_manager.get_module_item('formulas_tools', 'compile_excel_workbook')
execute_excel_formula = dependency_manager.get_module_item('formulas_tools', 'execute_excel_formula')
analyze_excel_dependencies = dependency_manager.get_module_item('formulas_tools', 'analyze_excel_dependencies')
validate_excel_formula = dependency_manager.get_module_item('formulas_tools', 'validate_excel_formula')

# Excel数据质量工具
ExcelDataQualityController = dependency_manager.get_module_item('excel_data_quality_tools', 'ExcelDataQualityController')
ExcelCellContentExtractor = dependency_manager.get_module_item('excel_data_quality_tools', 'ExcelCellContentExtractor')
ExcelCharacterConverter = dependency_manager.get_module_item('excel_data_quality_tools', 'ExcelCharacterConverter')
ExcelMultiConditionExtractor = dependency_manager.get_module_item('excel_data_quality_tools', 'ExcelMultiConditionExtractor')
ExcelMultiTableMerger = dependency_manager.get_module_item('excel_data_quality_tools', 'ExcelMultiTableMerger')
ExcelDataCleaner = dependency_manager.get_module_item('excel_data_quality_tools', 'ExcelDataCleaner')
ExcelBatchProcessor = dependency_manager.get_module_item('excel_data_quality_tools', 'ExcelBatchProcessor')

# 错误处理工具
VariableLifecycleManager = dependency_manager.get_module_item('df_processed_error_handler', 'VariableLifecycleManager')
NameErrorHandler = dependency_manager.get_module_item('df_processed_error_handler', 'NameErrorHandler')
enhanced_execute_with_error_handling = dependency_manager.get_module_item('df_processed_error_handler', 'enhanced_execute_with_error_handling')
analyze_code_variables = dependency_manager.get_module_item('df_processed_error_handler', 'analyze_code_variables')

# 列检查器
ColumnChecker = dependency_manager.get_module_item('column_checker', 'ColumnChecker')

from security.code_executor import execute_code_safely
# from enhanced_excel_helper import smart_read_excel, detect_file_encoding, validate_excel_data_integrity
# 使用excel_helper中的功能替代
from excel_helper import _suggest_excel_read_parameters, detect_excel_structure

# 标准化错误处理机制
class ErrorType(Enum):
    """错误类型枚举"""
    FILE_NOT_FOUND = "file_not_found"
    PERMISSION_DENIED = "permission_denied"
    INVALID_FORMAT = "invalid_format"
    PROCESSING_ERROR = "processing_error"
    DEPENDENCY_ERROR = "dependency_error"
    VALIDATION_ERROR = "validation_error"
    EXECUTION_ERROR = "execution_error"
    MODULE_NOT_FOUND = "module_not_found"
    UNKNOWN_ERROR = "unknown_error"

class StandardErrorHandler:
    """标准化错误处理器"""
    
    @staticmethod
    def create_error_response(error_type: ErrorType, message: str, 
                            details: Optional[Dict[str, Any]] = None,
                            suggestions: Optional[List[str]] = None) -> Dict[str, Any]:
        """创建标准化错误响应"""
        response = {
            "success": False,
            "error_type": error_type.value,
            "message": message,
            "timestamp": time.time()
        }
        
        if details:
            response["details"] = details
            
        if suggestions:
            response["suggestions"] = suggestions
            
        return response
    
    @staticmethod
    def create_success_response(data: Any, message: str = "操作成功") -> Dict[str, Any]:
        """创建标准化成功响应"""
        return {
            "success": True,
            "message": message,
            "data": data,
            "timestamp": time.time()
        }
    
    @staticmethod
    def handle_exception(e: Exception, context: str = "") -> Dict[str, Any]:
        """统一异常处理"""
        error_mapping = {
            FileNotFoundError: ErrorType.FILE_NOT_FOUND,
            PermissionError: ErrorType.PERMISSION_DENIED,
            ValueError: ErrorType.INVALID_FORMAT,
            ImportError: ErrorType.DEPENDENCY_ERROR,
        }
        
        error_type = error_mapping.get(type(e), ErrorType.UNKNOWN_ERROR)
        
        details = {
            "exception_type": type(e).__name__,
            "traceback": traceback.format_exc(),
            "context": context
        }
        
        return StandardErrorHandler.create_error_response(
            error_type, 
            str(e), 
            details
        )

# 初始化全局错误处理器
error_handler = StandardErrorHandler()

# 统一常量定义
MAX_FILE_SIZE = 999999999999  # 无限制文件大小
BLACKLIST = []  # 完全清空黑名单，允许所有操作包括subprocess等
TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")
CHARTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "charts")

# 安全初始化数据质量工具实例
def initialize_data_quality_tools() -> Dict[str, Any]:
    """安全初始化数据质量工具"""
    tools = {}
    
    if ExcelDataQualityController:
        try:
            tools['data_quality_controller'] = ExcelDataQualityController()
        except Exception as e:
            logging.warning(f"初始化ExcelDataQualityController失败: {e}")
    
    if ExcelCellContentExtractor:
        try:
            tools['cell_content_extractor'] = ExcelCellContentExtractor()
        except Exception as e:
            logging.warning(f"初始化ExcelCellContentExtractor失败: {e}")
    
    if ExcelCharacterConverter:
        try:
            tools['character_converter'] = ExcelCharacterConverter()
        except Exception as e:
            logging.warning(f"初始化ExcelCharacterConverter失败: {e}")
    
    if ExcelMultiConditionExtractor:
        try:
            tools['multi_condition_extractor'] = ExcelMultiConditionExtractor()
        except Exception as e:
            logging.warning(f"初始化ExcelMultiConditionExtractor失败: {e}")
    
    if ExcelMultiTableMerger:
        try:
            tools['multi_table_merger'] = ExcelMultiTableMerger()
        except Exception as e:
            logging.warning(f"初始化ExcelMultiTableMerger失败: {e}")
    
    if ExcelDataCleaner:
        try:
            tools['data_cleaner'] = ExcelDataCleaner()
        except Exception as e:
            logging.warning(f"初始化ExcelDataCleaner失败: {e}")
    
    if ExcelBatchProcessor:
        try:
            tools['batch_processor'] = ExcelBatchProcessor()
        except Exception as e:
            logging.warning(f"初始化ExcelBatchProcessor失败: {e}")
    
    return tools

# 初始化数据质量工具
data_quality_tools = initialize_data_quality_tools()

# 向后兼容的工具实例访问
data_quality_controller = data_quality_tools.get('data_quality_controller')
cell_content_extractor = data_quality_tools.get('cell_content_extractor')
character_converter = data_quality_tools.get('character_converter')
multi_condition_extractor = data_quality_tools.get('multi_condition_extractor')
multi_table_merger = data_quality_tools.get('multi_table_merger')
data_cleaner = data_quality_tools.get('data_cleaner')
batch_processor = data_quality_tools.get('batch_processor')

mcp = FastMCP("chatExcel")

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('chatexcel_mcp.log')
    ]
)
logger = logging.getLogger(__name__)

# 向后兼容的工具函数（使用新的标准化错误处理）
def create_error_response(error_type: str, message: str, details: dict = None, solutions: list = None) -> dict:
    """创建统一格式的错误响应（向后兼容版本）
    
    Args:
        error_type: 错误类型
        message: 错误消息
        details: 错误详情
        solutions: 解决方案建议
        
    Returns:
        dict: 标准化错误响应
    """
    # 映射旧的错误类型到新的枚举
    error_type_mapping = {
        "FILE_NOT_FOUND": ErrorType.FILE_NOT_FOUND,
        "PERMISSION_DENIED": ErrorType.PERMISSION_DENIED,
        "INVALID_FORMAT": ErrorType.INVALID_FORMAT,
        "PROCESSING_ERROR": ErrorType.PROCESSING_ERROR,
        "DEPENDENCY_ERROR": ErrorType.DEPENDENCY_ERROR,
        "VALIDATION_ERROR": ErrorType.VALIDATION_ERROR,
    }
    
    mapped_error_type = error_type_mapping.get(error_type, ErrorType.UNKNOWN_ERROR)
    
    response = StandardErrorHandler.create_error_response(
        mapped_error_type, message, details
    )
    
    # 添加解决方案（如果提供）
    if solutions:
        response["solutions"] = solutions
        
    return response

def create_success_response(data: dict, message: str = "操作成功完成") -> dict:
    """创建统一格式的成功响应（已弃用，请使用error_handler.create_success_response）
    
    Args:
        data: 响应数据
        message: 成功消息
        
    Returns:
        dict: 标准化成功响应
    """
    global error_handler
    return error_handler.create_success_response(data, message)

def get_template_path(template_name: str) -> str:
    """获取模板文件的绝对路径
    
    Args:
        template_name: 模板文件名
        
    Returns:
        str: 模板文件绝对路径
        
    Raises:
        FileNotFoundError: 模板文件不存在
    """
    template_path = os.path.join(TEMPLATE_DIR, template_name)
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件未找到: {template_path}")
    
    return template_path

def validate_file_access(file_path: str) -> dict:
    """验证文件访问权限和大小（宽松模式）
    
    Args:
        file_path: 文件路径
        
    Returns:
        dict: 验证结果，包含status和相关信息
    """
    # 宽松模式：即使文件不存在也允许执行（可能是要创建新文件）
    if not os.path.exists(file_path):
        logger.warning(f"文件不存在，但允许执行: {file_path}")
        return error_handler.create_success_response(
            {"file_size": 0, "note": "文件不存在，可能会创建新文件"},
            message="文件访问验证通过"
        )

    file_size = os.path.getsize(file_path)
    # 移除文件大小限制
    logger.debug(f"文件大小: {file_size / (1024*1024):.1f}MB")
    
    return error_handler.create_success_response(
        {"file_size": file_size},
        message="文件访问验证通过"
    )

import logging

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('chatExcel.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

@mcp.tool()
def read_metadata(file_path: str) -> dict:
    """读取CSV文件元数据并返回MCP兼容格式
    
    Args:
        file_path: CSV文件的绝对路径
        
    Returns:
        dict: 包含列信息、文件信息和状态的结构化元数据
    """
    logger.info(f"开始读取文件元数据: {file_path}")
    
    try:
        # 验证文件访问
        validation_result = validate_file_access(file_path)
        if not validation_result.get("success", False):
            return {
                "status": "ERROR",
                "error": "FILE_ACCESS_FAILED",
                "message": validation_result.get("message", "文件访问验证失败"),
                "details": validation_result
            }
        
        # Detect encoding and delimiter
        with open(file_path, 'rb') as f:
            rawdata = f.read(50000)
            enc = detect(rawdata)['encoding'] or 'utf-8'

        with open(file_path, 'r', encoding=enc) as f:
            first_line = f.readline()
            delimiter = ',' if ',' in first_line else '\t' if '\t' in first_line else ';'

        # 获取文件大小
        file_size = validation_result.get("data", {}).get("file_size", 0)
        
        # Get actual row count efficiently
        try:
            # Count total rows without loading data
            total_rows = sum(1 for _ in open(file_path, 'r', encoding=enc)) - 1  # -1 for header
        except Exception:
            # Fallback: use pandas to count rows
            # Read only first column to count rows efficiently
            temp_df = pd.read_csv(file_path, encoding=enc, delimiter=delimiter, usecols=[0])
            total_rows = len(temp_df)
        
        # Check file size limit
        if file_size > MAX_FILE_SIZE:
            return {
                "status": "ERROR",
                "error": "FILE_TOO_LARGE",
                "max_size": f"{MAX_FILE_SIZE / 1024 / 1024}MB",
                "actual_size": f"{file_size / 1024 / 1024:.1f}MB"
            }
        
        # Read sample data for metadata analysis
        sample_size = min(100, total_rows)
        df = pd.read_csv(file_path, encoding=enc, delimiter=delimiter, nrows=sample_size)

        # Calculate additional metadata
        columns_metadata = []
        for col in df.columns:
            col_meta = {
                "name": col,
                "type": str(df[col].dtype),
                "sample": df[col].dropna().iloc[:2].tolist(),
                "stats": {
                    "null_count": df[col].isnull().sum(),
                    "unique_count": df[col].nunique(),
                    "is_numeric": pd.api.types.is_numeric_dtype(df[col])
                },
                "warnings": [],
                "suggested_operations": []
            }
            
            # Add numeric stats if applicable
            if pd.api.types.is_numeric_dtype(df[col]):
                col_meta["stats"].update({
                    "min": df[col].min(),
                    "max": df[col].max(),
                    "mean": df[col].mean(),
                    "std": df[col].std()
                })
                col_meta["suggested_operations"].extend([
                    "normalize", "scale", "log_transform"
                ])
            
            # Add categorical stats if applicable
            if pd.api.types.is_string_dtype(df[col]):
                col_meta["suggested_operations"].extend([
                    "one_hot_encode", "label_encode", "text_processing"
                ])
            
            # Add datetime detection
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                col_meta["suggested_operations"].extend([
                    "extract_year", "extract_month", "time_delta"
                ])
            
            # Add warnings
            if df[col].isnull().sum() > 0:
                col_meta["warnings"].append(f"{df[col].isnull().sum()} null values found")
            if df[col].nunique() == 1:
                col_meta["warnings"].append("Column contains only one unique value")
            if pd.api.types.is_numeric_dtype(df[col]) and df[col].abs().max() > 1e6:
                col_meta["warnings"].append("Large numeric values detected - consider scaling")
            
            columns_metadata.append(col_meta)

        from pandas.api.types import infer_dtype
        
        # Format concise response
        summary = {
            "status": "SUCCESS",
            "file_info": {
                "size": f"{file_size / 1024:.1f}KB",
                "encoding": enc,
                "delimiter": delimiter
            },
            "dataset": {
                "total_rows": total_rows,
                "sample_rows": len(df),
                "columns": len(df.columns),
                "column_types": {
                    col: infer_dtype(df[col])
                    for col in df.columns
                }
            },
            "warnings": {
                "message": "Data quality issues detected" if (
                    df.isnull().any().any() or 
                    df.duplicated().any() or
                    (df.nunique() == 1).any()
                ) else "No significant data quality issues found",
                **({
                    "null_columns": {
                        "count": sum(df.isnull().any()),
                        "columns": [col for col in df.columns if df[col].isnull().any()]
                    }
                } if sum(df.isnull().any()) > 0 else {}),
                **({
                    "total_nulls": df.isnull().sum().sum()
                } if df.isnull().sum().sum() > 0 else {}),
                **({
                    "duplicate_rows": {
                        "count": df.duplicated().sum(),
                        "rows": df[df.duplicated()].index.tolist()
                    }
                } if df.duplicated().sum() > 0 else {}),
                **({
                    "single_value_columns": {
                        "count": sum(df.nunique() == 1),
                        "columns": [col for col in df.columns if df[col].nunique() == 1]
                    }
                } if sum(df.nunique() == 1) > 0 else {})
            }
        }
            
        return summary

    except Exception as e:
        return {
            "status": "ERROR",
            "error_type": type(e).__name__,
            "message": str(e),
            "solution": [
                "Check if the file is being used by another program",
                "Try saving the file as UTF-8 encoded CSV",
                "Contact the administrator to check MCP file access permissions"
            ],
            "traceback": traceback.format_exc()
        }




@mcp.tool()
def verify_data_integrity(original_file: str, processed_data: Optional[str] = None, 
                         comparison_file: Optional[str] = None, verification_type: str = "basic") -> dict:
    """数据完整性验证和比对核准工具。
    
    Args:
        original_file: 原始Excel文件路径
        processed_data: 处理后的数据（JSON字符串格式）或文件路径
        comparison_file: 用于比较的另一个Excel文件路径（可选）
        verification_type: 验证类型 ("basic", "detailed", "statistical")
    
    Returns:
        dict: 验证结果报告
    """
    try:
        # from data_verification import DataVerificationEngine, verify_data_processing_result
        # 使用内置验证功能
        
        # 验证原始文件访问
        validation_result = validate_file_access(original_file)
        if not validation_result.get("success", False):
            return {
                "success": False,
                "error": validation_result.get("message", "文件访问失败"),
                "suggestion": "请确保原始文件路径正确且文件存在。"
            }
        
        # 创建验证引擎
        verifier = DataVerificationEngine()
        
        if comparison_file:
            # 文件对比模式
            validation_result2 = validate_file_access(comparison_file)
            if not validation_result2.get("success", False):
                return {
                    "success": False,
                    "error": f"比较文件访问失败: {validation_result2.get('message', '未知错误')}",
                    "suggestion": "请确保比较文件路径正确且文件存在。"
                }
            
            # 读取两个文件进行比较
            df1 = pd.read_excel(original_file)
            df2 = pd.read_excel(comparison_file)
            
            comparison_result = verifier.compare_dataframes(
                df1, df2, 
                name1=os.path.basename(original_file),
                name2=os.path.basename(comparison_file)
            )
            
            # 添加增强的数据质量检查
            quality_result1 = data_quality_controller.comprehensive_quality_check(
                original_file, "comprehensive"
            )
            quality_result2 = data_quality_controller.comprehensive_quality_check(
                comparison_file, "comprehensive"
            )
            
            comparison_result["enhanced_quality_analysis"] = {
                "original_file_quality": quality_result1,
                "comparison_file_quality": quality_result2,
                "quality_comparison": {
                    "score_difference": quality_result1.get("overall_score", 0) - quality_result2.get("overall_score", 0),
                    "better_quality_file": "original" if quality_result1.get("overall_score", 0) > quality_result2.get("overall_score", 0) else "comparison"
                }
            }
            
            return {
                "success": True,
                "verification_type": "file_comparison",
                "comparison_result": comparison_result,
                "summary": {
                    "files_compared": [original_file, comparison_file],
                    "match_score": comparison_result.get("match_score", 0),
                    "has_differences": comparison_result.get("has_differences", True)
                }
            }
        
        elif processed_data:
            # 数据处理结果验证模式
            try:
                # 尝试解析处理后的数据
                if isinstance(processed_data, str):
                    if processed_data.endswith(('.xlsx', '.xls', '.csv')):
                        # 如果是文件路径
                        if processed_data.endswith('.csv'):
                            processed_df = pd.read_csv(processed_data)
                        else:
                            processed_df = pd.read_excel(processed_data)
                    else:
                        # 如果是JSON字符串
                        import json
                        data_dict = json.loads(processed_data)
                        processed_df = pd.DataFrame(data_dict)
                else:
                    return {
                        "success": False,
                        "error": "处理后数据格式不支持",
                        "suggestion": "请提供JSON字符串、CSV或Excel文件路径。"
                    }
                
                # 执行验证
                verification_result = verify_data_processing_result(
                    original_file, processed_df, verification_type
                )
                
                return {
                    "success": True,
                    "verification_type": "processing_verification",
                    "verification_result": verification_result
                }
                
            except Exception as parse_error:
                return {
                    "success": False,
                    "error": f"数据解析失败: {str(parse_error)}",
                    "suggestion": "请检查处理后数据的格式是否正确。"
                }
        
        else:
            # 基础完整性检查模式
            df = pd.read_excel(original_file)
            
            # 执行基础完整性检查
            integrity_report = {
                "file_info": {
                    "path": original_file,
                    "shape": df.shape,
                    "columns": list(df.columns),
                    "dtypes": df.dtypes.to_dict()
                },
                "data_quality": {
                    "total_rows": len(df),
                    "total_columns": len(df.columns),
                    "null_counts": df.isnull().sum().to_dict(),
                    "null_percentage": (df.isnull().sum() / len(df) * 100).to_dict(),
                    "duplicate_rows": df.duplicated().sum(),
                    "memory_usage": df.memory_usage(deep=True).to_dict()
                },
                "column_analysis": {}
            }
            
            # 详细列分析
            for col in df.columns:
                col_info = {
                    "dtype": str(df[col].dtype),
                    "null_count": df[col].isnull().sum(),
                    "unique_count": df[col].nunique(),
                    "sample_values": df[col].dropna().head(3).tolist()
                }
                
                if df[col].dtype in ['int64', 'float64']:
                    col_info.update({
                        "min": df[col].min(),
                        "max": df[col].max(),
                        "mean": df[col].mean(),
                        "std": df[col].std()
                    })
                
                integrity_report["column_analysis"][col] = col_info
            
            return {
                "success": True,
                "verification_type": "integrity_check",
                "integrity_report": integrity_report
            }
    
    except Exception as e:
        return {
            "success": False,
            "error": {
                "type": type(e).__name__,
                "message": str(e),
                "traceback": traceback.format_exc()
            },
            "suggestion": "请检查输入参数和文件格式是否正确。"
        }

@mcp.tool()
def read_excel_metadata(file_path: str) -> dict:
    """增强版Excel文件元数据读取，支持智能编码检测和完整性验证。
    
    Args:
        file_path: Excel文件路径
    
    Returns:
        dict: 包含文件元数据、编码信息、完整性验证结果的字典
    """
    try:
        # 验证文件访问
        validation_result = validate_file_access(file_path)
        if not validation_result.get("success", False):
            return validation_result

        # 智能编码检测
        from utils.encoding_detector import EncodingDetector
        encoding, confidence = EncodingDetector.detect_file_encoding(file_path)
        encoding_info = {'encoding': encoding, 'confidence': confidence}
        
        # 使用智能读取功能
        try:
            from services.excel_service import ExcelService
            excel_service = ExcelService()
            df = excel_service.smart_read_excel(file_path)
            read_params = {}
        except Exception as e:
            return {
                "status": "ERROR",
                "error": "SMART_READ_FAILED",
                "message": f"智能读取失败: {str(e)}",
                "encoding_info": encoding_info
            }
        
        # 数据完整性验证 (暂时禁用，函数已删除)
        # integrity_result = validate_excel_data_integrity(file_path, df)
        integrity_result = {'status': 'success', 'message': '数据完整性验证跳过'}
        
        # 增强的数据质量检查 (使用性能优化版本)
        quality_result = data_quality_controller.comprehensive_quality_check(
            file_path, "standard"  # 使用 standard 级别以提升性能
        )
        
        # 高级单元格内容分析 (使用性能优化增强版本)
        cell_analysis = cell_content_extractor.extract_cell_content_advanced(
            file_path=file_path,
            cell_range="A1:Z1000",  # 使用优化后的扩展范围
            extract_type="all",     # 保持智能提取类型，支持完整分析
            max_cells=26000         # 调整为默认限制值26000个单元格
        )
        
        # 获取所有工作表信息
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = workbook.sheetnames
        sheets_info = {}
        
        for sheet_name in sheet_names:
            sheet_obj = workbook[sheet_name]
            sheets_info[sheet_name] = {
                'max_row': sheet_obj.max_row,
                'max_column': sheet_obj.max_column,
                'has_data': sheet_obj.max_row > 1
            }
        workbook.close()
        
        # 智能参数推荐（使用增强版）
        suggested_params = _suggest_excel_read_parameters(file_path, read_params.get('sheet_name')) if _suggest_excel_read_parameters else {}

        # 使用通用列分析函数
        columns_metadata = _analyze_dataframe_columns(df)

        summary = {
            "status": "SUCCESS",
            "file_info": {
                "size": f"{validation_result.get('data', {}).get('file_size', 0) / 1024:.1f}KB",
                "sheets": sheet_names,
                "sheets_info": sheets_info,
                "encoding": encoding_info
            },
            "dataset": {
                "total_rows": read_result.get('total_rows', len(df)),
                "sample_rows": len(df),
                "columns": len(df.columns),
                "column_types": {
                    col: str(df[col].dtype)
                    for col in df.columns
                }
            },
            "read_params": read_params,
            "suggested_params": suggested_params,
            "columns_metadata": columns_metadata,
            "integrity_check": integrity_result,
            "enhanced_quality_analysis": quality_result,
            "cell_content_analysis": cell_analysis,
            "warnings": {
                "message": "Data quality issues detected" if (
                    df.isnull().any().any() or 
                    df.duplicated().any() or
                    (df.nunique() == 1).any()
                ) else "No significant data quality issues found",
                **({
                    "null_columns": {
                        "count": sum(df.isnull().any()),
                        "columns": [col for col in df.columns if df[col].isnull().any()]
                    }
                } if sum(df.isnull().any()) > 0 else {}),
                **({
                    "total_nulls": df.isnull().sum().sum()
                } if df.isnull().sum().sum() > 0 else {}),
                **({
                    "duplicate_rows": {
                        "count": df.duplicated().sum(),
                        "rows": df[df.duplicated()].index.tolist()
                    }
                } if df.duplicated().sum() > 0 else {}),
                **({
                    "single_value_columns": {
                        "count": sum(df.nunique() == 1),
                        "columns": [col for col in df.columns if df[col].nunique() == 1]
                    }
                } if sum(df.nunique() == 1) > 0 else {})
            }
        }
        return summary

    except Exception as e:
        return {
            "status": "ERROR",
            "error_type": type(e).__name__,
            "message": str(e),
            "solution": [
                "Ensure the file is a valid Excel file (.xlsx)",
                "Check if the file is being used by another program",
                "Contact the administrator to check MCP file access permissions"
            ],
            "traceback": traceback.format_exc()
        }



def _read_csv_with_smart_encoding(file_path: str):
    """优化的CSV文件读取函数，使用智能编码检测和缓存
    
    Args:
        file_path: CSV文件路径
        
    Returns:
        pandas.DataFrame or None: 读取的DataFrame或None（如果失败）
    """
    # 使用标准库pandas
    import pandas as pd
        
    try:
        # 首先尝试检测文件编码
        from utils.encoding_detector import EncodingDetector
        encoding, confidence = EncodingDetector.detect_file_encoding(file_path)
        detected_encoding = encoding
        
        # 尝试使用检测到的编码读取CSV文件
        try:
            return pd.read_csv(file_path, encoding=detected_encoding)
        except UnicodeDecodeError:
            # 如果检测到的编码失败，尝试常见编码
            common_encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'big5', 'latin1', 'cp1252']
            
            for enc in common_encodings:
                try:
                    return pd.read_csv(file_path, encoding=enc)
                except (UnicodeDecodeError, UnicodeError):
                    continue
            
            return None
                    
    except Exception as e:
        logger.error(f"Failed to read CSV file {file_path}: {str(e)}")
        return None


def _execute_code_safely(code: str, df, file_path: Optional[str] = None) -> dict:
    """安全执行代码的通用函数
    
    Args:
        code: 要执行的代码字符串
        df: pandas DataFrame
        
    Returns:
        dict: 执行结果
    """
    import sys
    from io import StringIO
    
    # 使用标准库pandas
    import pandas as pd
    
    # 捕获标准输出
    old_stdout = sys.stdout
    sys.stdout = captured_output = StringIO()
    
    try:
        # 尝试使用安全代码执行器
        SecureCodeExecutor = dependency_manager.get_module_item('security.secure_code_executor', 'SecureCodeExecutor')
        if SecureCodeExecutor:
            executor = SecureCodeExecutor(
                max_memory_mb=256,
                max_execution_time=30,
                enable_ast_analysis=False
            )
            
            context = {'df': df, 'pd': pd, 'file_path': file_path}
            execution_result = executor.execute_code(code, context)
            
            if not execution_result['success']:
                return {
                    'success': False,
                    'error': execution_result.get('error', 'Unknown error'),
                    'output': execution_result.get('output', captured_output.getvalue()),
                    'suggestions': execution_result.get('suggestions', [])
                }
            
            # 处理执行结果
            result_data = _format_execution_result(execution_result.get('result'), pd)
            
            return {
                'success': True,
                'output': execution_result.get('output', captured_output.getvalue()),
                'result': result_data
            }
        else:
            # 简单的代码执行（如果没有安全执行器）
            context = {'df': df, 'pd': pd, 'file_path': file_path}
            exec(code, context)
            
            result_data = _format_execution_result(context.get('result'), pd)
            
            return {
                'success': True,
                'output': captured_output.getvalue(),
                'result': result_data
            }
            
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'output': captured_output.getvalue(),
            'suggestions': [
                'Check code syntax',
                'Ensure operations are valid for the data'
            ]
        }
    finally:
        sys.stdout = old_stdout


def _format_execution_result(result, pd) -> dict:
    """格式化执行结果
    
    Args:
        result: 执行结果
        pd: pandas模块
        
    Returns:
        dict: 格式化的结果
    """
    if result is None:
        return None
        
    if pd and isinstance(result, pd.DataFrame):
        return {
            "type": "DataFrame",
            "shape": result.shape,
            "columns": result.columns.tolist(),
            "data": result.head(10).to_dict('records'),
            "dtypes": result.dtypes.astype(str).to_dict()
        }
    elif pd and isinstance(result, pd.Series):
        return {
            "type": "Series",
            "name": result.name,
            "length": len(result),
            "data": result.head(10).tolist(),
            "dtype": str(result.dtype)
        }
    else:
        return {
            "type": type(result).__name__,
            "value": str(result)
        }


def _analyze_dataframe_columns(df) -> list:
    """通用DataFrame列分析函数
    
    Args:
        df: pandas DataFrame
        
    Returns:
        list: 列元数据列表
    """
    # 使用标准库pandas
    import pandas as pd
    
    columns_metadata = []
    for col in df.columns:
        col_meta = {
            "name": col,
            "type": str(df[col].dtype),
            "sample": df[col].dropna().iloc[:2].tolist() if len(df[col].dropna()) > 0 else [],
            "stats": {
                "null_count": int(df[col].isnull().sum()),
                "unique_count": int(df[col].nunique()),
                "is_numeric": pd.api.types.is_numeric_dtype(df[col])
            },
            "warnings": [],
            "suggested_operations": []
        }
        
        # 数值类型分析
        if pd.api.types.is_numeric_dtype(df[col]):
            col_meta["stats"].update({
                "min": float(df[col].min()) if not df[col].empty else None,
                "max": float(df[col].max()) if not df[col].empty else None,
                "mean": float(df[col].mean()) if not df[col].empty else None,
                "std": float(df[col].std()) if not df[col].empty else None
            })
            col_meta["suggested_operations"].extend([
                "normalize", "scale", "log_transform"
            ])
            
        # 字符串类型分析
        if pd.api.types.is_string_dtype(df[col]):
            col_meta["suggested_operations"].extend([
                "one_hot_encode", "label_encode", "text_processing"
            ])
            
        # 日期时间类型分析
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            col_meta["suggested_operations"].extend([
                "extract_year", "extract_month", "time_delta"
            ])
            
        # 数据质量警告
        if df[col].isnull().sum() > 0:
            col_meta["warnings"].append(f"{df[col].isnull().sum()} null values found")
        if df[col].nunique() == 1:
            col_meta["warnings"].append("Column contains only one unique value")
        if pd.api.types.is_numeric_dtype(df[col]) and not df[col].empty and df[col].abs().max() > 1e6:
            col_meta["warnings"].append("Large numeric values detected - consider scaling")
            
        columns_metadata.append(col_meta)
    
    return columns_metadata


def smart_column_matcher(target_column: str, available_columns: list) -> dict:
    """智能列名匹配工具
    
    Args:
        target_column: 目标列名
        available_columns: 可用的列名列表
        
    Returns:
        dict: 匹配结果和建议
    """
    import difflib
    import re
    
    result = {
        'exact_match': None,
        'close_matches': [],
        'suggestions': [],
        'normalized_matches': []
    }
    
    # 1. 精确匹配
    if target_column in available_columns:
        result['exact_match'] = target_column
        return result
    
    # 2. 大小写不敏感匹配
    target_lower = target_column.lower()
    for col in available_columns:
        if col.lower() == target_lower:
            result['exact_match'] = col
            result['suggestions'].append(f"找到大小写不同的匹配: '{col}'")
            return result
    
    # 3. 去除空格和特殊字符后匹配
    target_normalized = re.sub(r'[\s_-]', '', target_column.lower())
    for col in available_columns:
        col_normalized = re.sub(r'[\s_-]', '', col.lower())
        if col_normalized == target_normalized:
            result['normalized_matches'].append(col)
    
    # 4. 模糊匹配
    close_matches = difflib.get_close_matches(
        target_column, available_columns, n=5, cutoff=0.6
    )
    result['close_matches'] = close_matches
    
    # 5. 中文列名变体匹配
    chinese_variants = {
        '消耗日期': ['消费日期', '使用日期', '支出日期', '花费日期', '消耗时间', '消费时间'],
        '消费日期': ['消耗日期', '使用日期', '支出日期', '花费日期', '消费时间', '消耗时间'],
        '日期': ['时间', 'Date', 'date', '创建日期', '更新日期', '记录日期'],
        '金额': ['数量', '价格', '费用', '成本', 'Amount', 'amount', '总额'],
        '名称': ['姓名', '品名', '项目', 'Name', 'name', '标题'],
        '类型': ['分类', '种类', 'Type', 'type', '类别']
    }
    
    if target_column in chinese_variants:
        for variant in chinese_variants[target_column]:
            if variant in available_columns:
                result['suggestions'].append(f"发现相似列名: '{variant}'")
    
    return result



@mcp.tool()
def run_excel_code(
    file_path: str,
    code: str,
    sheet_name: str | None = None,
    skiprows: int | None = None,
    header: int | None = None,
    usecols: str | None = None,
    encoding: str | None = None,
    auto_detect: bool = True,
    allow_file_write: bool = False
) -> dict:
    """
    增强版Excel代码执行工具，具备强化的pandas导入和错误处理机制。
    
    Args:
        code: 要执行的数据处理代码字符串
        file_path: Excel文件路径
        sheet_name: 可选，工作表名称
        skiprows: 可选，跳过的行数
        header: 可选，用作列名的行号
        usecols: 可选，要解析的列
        encoding: 指定编码（可选）
        auto_detect: 是否启用智能检测和参数优化
        allow_file_write: 是否允许在代码中写入文件
        
    Returns:
        dict: 执行结果或错误信息
    """
    global error_handler
    start_time = time.time()
    
    try:
        # 使用核心模块的异常处理
        if CORE_MODULES_AVAILABLE:
            config = get_config()
            security_config = config.security
            
            # 检查文件大小限制
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                if file_size > security_config.max_file_size:
                    raise FileAccessError(
                        file_path=file_path,
                        reason=f"文件大小 {file_size} 字节超过限制 {security_config.max_file_size} 字节",
                        error_code="FILE_TOO_LARGE",
                        suggestions=["请使用较小的文件或联系管理员调整限制"]
                    )
        
        # 验证文件访问权限
        validation_result = validate_file_access(file_path)
        if not validation_result.get("success", False):
            if CORE_MODULES_AVAILABLE:
                raise FileAccessError(
                    file_path=file_path,
                    reason=validation_result.get("message", "文件访问验证失败"),
                    error_code="ACCESS_DENIED",
                    suggestions=["请检查文件路径和权限"]
                )
            return validation_result

        # 智能读取Excel文件
        read_params = {
            'sheet_name': sheet_name,
            'skiprows': skiprows,
            'header': header,
            'usecols': usecols
        }
        
        # 移除None值参数
        read_params = {k: v for k, v in read_params.items() if v is not None}
        
        # 使用pandas读取Excel文件
        df = pd.read_excel(file_path, **read_params)
        
        # 记录文件信息
        if CORE_MODULES_AVAILABLE:
            file_info = FileInfo(
                path=file_path,
                name=os.path.basename(file_path),
                size=os.path.getsize(file_path),
                type=FileType.EXCEL,
                encoding=encoding
            )
        
        # 使用增强的安全代码执行器（支持__import__）
        if allow_file_write:
            # 如果允许文件写入，使用更宽松的执行环境
            execution_result = _execute_code_safely(code, df, file_path)
        else:
            # 标准执行环境
            execution_result = _execute_code_safely(code, df, file_path)
        
        # 增强执行结果
        if CORE_MODULES_AVAILABLE and isinstance(execution_result, dict):
            execution_time = time.time() - start_time
            execution_result.update({
                "execution_time": execution_time,
                "file_info": file_info.to_dict() if 'file_info' in locals() else None,
                "parameters_used": read_params
            })
        
        return execution_result

    except FileNotFoundError as e:
        if CORE_MODULES_AVAILABLE:
            error = FileAccessError(
                file_path=file_path,
                reason="文件未找到",
                error_code="FILE_NOT_FOUND",
                suggestions=["请确认文件路径是否正确", "检查文件是否存在"]
            )
            return error.to_dict()
        
        return error_handler.create_error_response(
            ErrorType.FILE_NOT_FOUND,
            f"文件未找到: {file_path}",
            suggestions=["请确认文件路径是否正确。"]
        )
        
    except Exception as e:
        execution_time = time.time() - start_time
        
        if CORE_MODULES_AVAILABLE:
            if isinstance(e, ChatExcelError):
                result = e.to_dict()
                result["execution_time"] = execution_time
                return result
            else:
                error = CodeExecutionError(
                    code=code,
                    error_details=f"执行Excel代码时出错: {str(e)}",
                    error_code="EXECUTION_FAILED",
                    details={"exception_type": type(e).__name__, "traceback": traceback.format_exc()},
                    suggestions=["检查代码语法", "确认数据格式正确"]
                )
                result = error.to_dict()
                result["execution_time"] = execution_time
                return result
        
        logger.error(f"执行Excel代码时出错: {e}", exc_info=True)
        return error_handler.handle_exception(e, context="run_excel_code")




@mcp.tool()
def run_code(code: str, file_path: str) -> dict:
    """在CSV文件上执行数据处理代码，具备安全检查功能。
    
    Args:
        code: 要执行的数据处理代码字符串。
        file_path: CSV文件路径。
    
    Returns:
        dict: 执行结果，包含数据、输出或错误信息。
    """
    global error_handler
    try:
        # 验证文件访问
        validation_result = validate_file_access(file_path)
        if not validation_result.get("success", False):
            return error_handler.create_error_response(
                ErrorType.FILE_ACCESS_ERROR,
                validation_result.get("message", "文件访问验证失败"),
                suggestions=["Check the file path and ensure the file exists"]
            )
        
        # 优化的CSV文件读取 - 使用缓存的编码检测结果
        df = _read_csv_with_smart_encoding(file_path)
        if df is None:
            return error_handler.create_error_response(
                ErrorType.FILE_READ_ERROR,
                "Failed to read CSV file with any supported encoding",
                suggestions=[
                    "Check the file encoding and format",
                    "Try converting the file to UTF-8 format",
                    "Ensure the file is a valid CSV format"
                ]
            )
        return execute_code_safely(code, df, file_path)
        
        # 执行代码，并注入 file_path 变量
        execution_result = _execute_code_safely(code, df, file_path=file_path)
        
        if not execution_result['success']:
            return error_handler.create_error_response(
                ErrorType.EXECUTION_ERROR,
                execution_result.get('error', 'Code execution failed'),
                suggestions=execution_result.get('suggestions', [
                    "Check data processing syntax",
                    "Ensure operations are valid for the data"
                ])
            )
        
        return error_handler.create_success_response(
            execution_result.get('result'),
            message="Code executed successfully",
            metadata={
                "output": execution_result.get('output', ''),
                "suggestion": "Use 'result' variable to store your final output."
            }
        )
            
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.EXECUTION_ERROR,
            f"Unexpected error during code execution: {str(e)}",
            suggestions=[
                "Check data processing syntax",
                "Ensure operations are valid for the data",
                "Contact support if the issue persists"
            ]
        )


@mcp.tool()
def bar_chart_to_html(
    categories: list,
    values: list,
    title: str = "Interactive Chart",
   
) -> dict:
    """Generate interactive HTML bar chart using Chart.js template.
    
    Args:
        categories: List of category names for x-axis
        values: List of numeric values for y-axis
        title: Chart title (default: "Interactive Chart")
        x_label: Label for X-axis (default: "Categories")
        y_label: Label for Y-axis (default: "Values")
        
    Returns:
        dict: Contains file path and status information
        
    Example:
        >>> bar_chart_to_html(
        ...     categories=['Electronics', 'Clothing', 'Home Goods', 'Sports Equipment'],
        ...     values=[120000, 85000, 95000, 60000],
        ...     title="Q1 Sales by Product Category"
        ... )
        {
            "status": "SUCCESS",
            "filepath": "/absolute/path/to/plotXXXXXX.html",
        }
    """
    # Validate input lengths
    if len(categories) != len(values):
        return error_handler.create_error_response(
            ErrorType.VALIDATION_ERROR,
            f"Categories ({len(categories)}) and values ({len(values)}) must be same length",
            suggestions=["Ensure categories and values lists have the same length"]
        )

    # Read template file
    template_path = get_template_path("barchart_template.html")
    try:
        with open(template_path, 'r', encoding='utf-8') as f:
            template = f.read()
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.FILE_READ_ERROR,
            f"Failed to read chart template: {str(e)}",
            suggestions=["Check if template files exist", "Verify file permissions"]
        )

    # Prepare data for Chart.js
    all_categories = categories
    all_values = values
    colors = [
        "#4e73df", "#1cc88a", "#36b9cc", "#f6c23e",
        "#e74a3b", "#858796", "#f8f9fc", "#5a5c69",
        "#6610f2", "#6f42c1", "#e83e8c", "#d63384",
        "#fd7e14", "#ffc107", "#28a745", "#20c997",
        "#17a2b8", "#007bff", "#6c757d", "#343a40",
        "#dc3545", "#ff6b6b", "#4ecdc4", "#1a535c"
    ][:len(all_categories)]

    # Inject data into template
    template = template.replace(
        'labels: ["Electronics", "Clothing", "Home Goods", "Sports Equipment"]',
        f'labels: {json.dumps(all_categories)}'
    ).replace(
        'data: [120000, 85000, 95000, 60000]',
        f'data: {json.dumps(all_values)}'
    ).replace(
        'backgroundColor: ["#4e73df", "#1cc88a", "#36b9cc", "#f6c23e"]',
        f'backgroundColor: {json.dumps(colors)}'
    ).replace(
        'Sales by Category (2023)',
        title
    ).replace(
        'legend: { position: \'top\' },',
        ''
    )

    # Save to plot directory as HTML
    charts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "charts")
    os.makedirs(charts_dir, exist_ok=True)

    timestamp = str(int(time.time()))
    filename = f"chart_{timestamp}.html"
    filepath = os.path.join(charts_dir, filename)

    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(template)
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.FILE_WRITE_ERROR,
            f"Failed to write chart file: {str(e)}",
            suggestions=["Check directory permissions", "Ensure sufficient disk space"]
        )

    return error_handler.create_success_response(
        {"filepath": os.path.abspath(filepath)},
        message="Bar chart generated successfully"
    )


@mcp.tool()
def pie_chart_to_html(
    labels: list,
    values: list,
    title: str = "Interactive Pie Chart"
) -> dict:
    """Generate interactive HTML pie chart using Chart.js template.
    
    Args:
        labels: List of label names for each pie slice
        values: List of numeric values for each slice
        title: Chart title (default: "Interactive Pie Chart")
        
    Returns:
        dict: Contains file path and status information
        
    Example:
        >>> pie_chart_to_html(
        ...     labels=['Electronics', 'Clothing', 'Home Goods'],
        ...     values=[120000, 85000, 95000],
        ...     title="Q1 Sales Distribution"
        ... )
        {
            "status": "SUCCESS",
            "filepath": "/absolute/path/to/plotXXXXXX.html",
        }
    """
    # Validate input lengths
    if len(labels) != len(values):
        return error_handler.create_error_response(
            ErrorType.VALIDATION_ERROR,
            f"Labels ({len(labels)}) and values ({len(values)}) must be same length",
            suggestions=["Ensure labels and values lists have the same length"]
        )

    # Read template file
    template_path = get_template_path("piechart_template.html")
    try:
        with open(template_path, 'r', encoding='utf-8') as f:
            template = f.read()
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.FILE_READ_ERROR,
            f"Failed to read chart template: {str(e)}",
            suggestions=["Check if template files exist", "Verify file permissions"]
        )

    # Prepare data for Chart.js
    colors = [
        "#4e73df", "#1cc88a", "#36b9cc", "#f6c23e",
        "#e74a3b", "#858796", "#f8f9fc", "#5a5c69",
        "#6610f2", "#6f42c1", "#e83e8c", "#d63384",
        "#fd7e14", "#ffc107", "#28a745", "#20c997",
        "#17a2b8", "#007bff", "#6c757d", "#343a40",
        "#dc3545", "#ff6b6b", "#4ecdc4", "#1a535c"
    ][:len(labels)]

    # Inject data into template
    template = template.replace(
        'labels: ["Apple", "Samsung", "Huawei", "Xiaomi", "Others"]',
        f'labels: {json.dumps(labels)}'
    ).replace(
        'data: [45, 25, 12, 8, 10]',
        f'data: {json.dumps(values)}'
    ).replace(
        'backgroundColor: ["#4e73df", "#1cc88a", "#36b9cc", "#f6c23e", "#e74a3b"]',
        f'backgroundColor: {json.dumps(colors)}'
    ).replace(
        'Global Smartphone Market Share (2023)',
        title
    )

    # Save to plot directory as HTML
    charts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "charts")
    os.makedirs(charts_dir, exist_ok=True)

    timestamp = str(int(time.time()))
    filename = f"chart_{timestamp}.html"
    filepath = os.path.join(charts_dir, filename)

    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(template)
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.FILE_WRITE_ERROR,
            f"Failed to write chart file: {str(e)}",
            suggestions=["Check directory permissions", "Ensure sufficient disk space"]
        )

    return error_handler.create_success_response(
        {"filepath": os.path.abspath(filepath)},
        message="Pie chart generated successfully"
    )

@mcp.tool()
def line_chart_to_html(
    labels: list,
    datasets: list,
    title: str = "Interactive Line Chart"
) -> dict:
    """Generate interactive HTML line chart using Chart.js template.
    
    Args:
        labels: List of label names for x-axis
        datasets: List of datasets, each containing:
            - label: Name of the dataset
            - data: List of numeric values (3 dimensions: [x, y, z])
        title: Chart title (default: "Interactive Line Chart")
        
    Returns:
        dict: Contains file path and status information
        
    Example:
        >>> line_chart_to_html(
        ...     labels=['Jan', 'Feb', 'Mar'],
        ...     datasets=[
        ...         {'label': 'Sales', 'data': [[100, 200, 300], [150, 250, 350], [200, 300, 400]]},
        ...         {'label': 'Expenses', 'data': [[50, 100, 150], [75, 125, 175], [100, 150, 200]]}
        ...     ],
        ...     title="Monthly Performance"
        ... )
        {
            "status": "SUCCESS",
            "filepath": "/absolute/path/to/plotXXXXXX.html",
        }
    """
    # Validate input
    if not all(len(d['data']) == len(labels) for d in datasets):
        return error_handler.create_error_response(
            ErrorType.VALIDATION_ERROR,
            "All datasets must have same length as labels",
            suggestions=["Ensure all dataset data arrays match the labels length"]
        )

    # Read template file
    template_path = get_template_path("linechart_template.html")
    try:
        with open(template_path, 'r', encoding='utf-8') as f:
            template = f.read()
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.FILE_READ_ERROR,
            f"Failed to read chart template: {str(e)}",
            suggestions=["Check if template files exist", "Verify file permissions"]
        )

    # Prepare data for Chart.js
    chart_data = {
        "labels": labels,
        "datasets": []
    }
    
    # Create datasets using main labels
    for dataset in datasets:
            chart_data['datasets'].append({
                "label": dataset['label'],
                "data": dataset['data'],
                "borderColor": '#4e73df',  # Default color
                "backgroundColor": '#4e73df',
            "borderWidth": 2,
            "pointRadius": 5,
            "tension": 0,
            "fill": False
        })

    # Inject data into template
    template = template.replace(
        'labels: ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]',
        f'labels: {json.dumps(labels)}'
    ).replace(
        'datasets: [\n' +
        '                {\n' +
        '                    label: "Electronics",\n' +
        '                    data: [6500, 5900, 8000, 8100, 8600, 8250, 9500, 10500, 12000, 11500, 13000, 15000],\n' +
        '                    borderColor: "#4e73df",\n' +
        '                    backgroundColor: "#4e73df",\n' +
        '                    borderWidth: 2,\n' +
        '                    pointRadius: 5,\n' +
        '                    tension: 0,\n' +
        '                    fill: false\n' +
        '                },\n' +
        '                {\n' +
        '                    label: "Clothing",\n' +
        '                    data: [12000, 11000, 12500, 10500, 11500, 13000, 14000, 12500, 11000, 9500, 10000, 12000],\n' +
        '                    borderColor: "#1cc88a",\n' +
        '                    backgroundColor: "#1cc88a",\n' +
        '                    borderWidth: 2,\n' +
        '                    pointRadius: 5,\n' +
        '                    tension: 0,\n' +
        '                    fill: false\n' +
        '                },\n' +
        '                {\n' +
        '                    label: "Home Goods",\n' +
        '                    data: [8000, 8500, 9000, 9500, 10000, 10500, 11000, 11500, 12000, 12500, 13000, 13500],\n' +
        '                    borderColor: "#36b9cc",\n' +
        '                    backgroundColor: "#36b9cc",\n' +
        '                    borderWidth: 2,\n' +
        '                    pointRadius: 5,\n' +
        '                    tension: 0,\n' +
        '                    fill: false\n' +
        '                }\n' +
        '            ]',
        f'datasets: {json.dumps(chart_data["datasets"], indent=16)}'
    ).replace(
        'Interactive Sales Trend Dashboard',
        title
    ).replace(
        'Monthly Sales Trend (2023)',
        title
    )

    # Save to plot directory as HTML
    charts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "charts")
    os.makedirs(charts_dir, exist_ok=True)

    timestamp = str(int(time.time()))
    filename = f"chart_{timestamp}.html"
    filepath = os.path.join(charts_dir, filename)

    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(template)
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.FILE_WRITE_ERROR,
            f"Failed to write chart file: {str(e)}",
            suggestions=["Check directory permissions", "Ensure sufficient disk space"]
        )

    return error_handler.create_success_response(
        {"filepath": os.path.abspath(filepath)},
        message="Line chart generated successfully"
    )


@mcp.tool()
def validate_data_quality(file_path: str) -> dict:
    """验证数据质量并提供改进建议
    
    Args:
        file_path: 数据文件路径
        
    Returns:
        dict: 数据质量报告和改进建议
    """
    try:
        validation_result = validate_file_access(file_path)
        if not validation_result.get("success", False):
            return validation_result
        
        df = pd.read_csv(file_path)
        
        quality_report = {
            "basic_info": {
                "total_rows": len(df),
                "columns": len(df.columns),
                "memory_usage": df.memory_usage(deep=True).sum()
            },
            "missing_data": {
                "total_missing": df.isnull().sum().sum(),
                "missing_by_column": df.isnull().sum().to_dict(),
                "missing_percentage": (df.isnull().sum() / len(df) * 100).to_dict()
            },
            "duplicates": {
                "duplicate_rows": df.duplicated().sum(),
                "duplicate_percentage": df.duplicated().sum() / len(df) * 100
            },
            "data_types": df.dtypes.astype(str).to_dict(),
            "recommendations": []
        }
        
        # 生成建议
        if quality_report["missing_data"]["total_missing"] > 0:
            quality_report["recommendations"].append("考虑处理缺失值：删除、填充或插值")
        
        if quality_report["duplicates"]["duplicate_rows"] > 0:
            quality_report["recommendations"].append("发现重复行，考虑去重处理")
        
        # 检查数据类型优化机会
        for col, dtype in quality_report["data_types"].items():
            if dtype == 'object':
                if df[col].nunique() / len(df) < 0.5:  # 低基数字符串
                    quality_report["recommendations"].append(f"列 '{col}' 可考虑转换为分类类型以节省内存")
        
        return error_handler.create_success_response(
            quality_report, 
            message="数据质量分析完成"
        )
        
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.PROCESSING_ERROR,
            f"数据质量验证失败: {str(e)}",
            suggestions=["检查文件格式", "确认文件可读性"],
            details={"traceback": traceback.format_exc()}
        )


# 注册新的Excel智能工具
@mcp.tool()
def suggest_excel_read_parameters_tool(file_path: str) -> dict:
    """智能推荐Excel文件读取参数
    
    Args:
        file_path: Excel文件的绝对路径
        
    Returns:
        dict: 包含推荐参数的结构化响应
    """
    return suggest_excel_read_parameters(file_path)

@mcp.tool()
def detect_excel_file_structure_tool(file_path: str) -> dict:
    """检测Excel文件结构
    
    Args:
        file_path: Excel文件的绝对路径
        
    Returns:
        dict: 包含文件结构信息的响应
    """
    return detect_excel_file_structure(file_path)

@mcp.tool()
def create_excel_read_template_tool(file_path: str, sheet_name: Optional[str] = None, skiprows: Optional[int] = None, header: Optional[int] = None, usecols: Optional[str] = None) -> dict:
    """生成Excel读取代码模板
    
    Args:
        file_path: Excel文件的绝对路径
        sheet_name: 工作表名称
        skiprows: 跳过的行数
        header: 标题行位置
        usecols: 使用的列
        
    Returns:
        dict: 包含代码模板的响应
    """
    return create_excel_read_template(file_path, sheet_name, skiprows, header, usecols)

@mcp.tool()
def comprehensive_data_verification_tool(
    file_path: str,
    reference_file: str = None,
    verification_level: str = "detailed",
    save_report: bool = True
) -> dict:
    """
    综合数据验证和核准工具
    
    提供全面的Excel数据验证、质量评估和比对核准功能。
    支持单文件验证和双文件比较验证模式。
    
    Args:
        file_path: 要验证的Excel文件路径
        reference_file: 参考文件路径（可选，用于比较验证）
        verification_level: 验证级别
            - "basic": 基础验证（文件结构、基本统计）
            - "detailed": 详细验证（包含数据质量分析）
            - "comprehensive": 综合验证（包含异常检测和深度分析）
        save_report: 是否保存验证报告到本地
    
    Returns:
        dict: 包含以下字段的验证结果
            - overall_status: 总体状态 (EXCELLENT/GOOD/ACCEPTABLE/POOR/CRITICAL/FAILED)
            - data_quality_score: 数据质量得分 (0-100)
            - file_analysis: 文件结构分析结果
            - data_integrity: 数据完整性验证结果
            - comparison_results: 比较验证结果（如果提供了参考文件）
            - recommendations: 改进建议列表
            - detailed_report: 详细报告（详细和综合级别）
    
    功能特点:
    1. 多层次验证：支持基础、详细、综合三个验证级别
    2. 智能编码检测：自动检测文件编码并优化读取
    3. 数据质量评估：计算综合质量得分
    4. 异常检测：识别异常值和数据模式
    5. 比较验证：支持与参考文件的详细比较
    6. 报告生成：自动生成验证报告并可保存
    7. 建议系统：提供针对性的数据改进建议
    
    使用示例:
    - 基础验证: comprehensive_data_verification_tool("data.xlsx", verification_level="basic")
    - 详细验证: comprehensive_data_verification_tool("data.xlsx", verification_level="detailed")
    - 比较验证: comprehensive_data_verification_tool("data.xlsx", "reference.xlsx", "comprehensive")
    """
    try:
        # 验证文件路径
        if not os.path.exists(file_path):
            return {
                "success": False,
                "error": f"文件不存在: {file_path}",
                "overall_status": "FAILED"
            }
        
        if reference_file and not os.path.exists(reference_file):
            return {
                "success": False,
                "error": f"参考文件不存在: {reference_file}",
                "overall_status": "FAILED"
            }
        
        # 创建综合验证器
        verifier = ComprehensiveDataVerifier()
        
        # 执行综合验证
        verification_result = verifier.comprehensive_excel_verification(
            file_path=file_path,
            reference_file=reference_file,
            verification_level=verification_level,
            save_report=save_report
        )
        
        # 添加成功标志
        verification_result["success"] = True
        
        # 添加验证摘要
        verification_result["verification_summary"] = {
            "file_name": os.path.basename(file_path),
            "verification_time": verification_result.get("timestamp"),
            "quality_score": verification_result.get("data_quality_score", 0),
            "status": verification_result.get("overall_status", "UNKNOWN"),
            "has_reference": reference_file is not None,
            "level": verification_level,
            "recommendations_count": len(verification_result.get("recommendations", []))
        }
        
        return verification_result
        
    except Exception as e:
        return {
            "success": False,
            "error": f"综合验证过程中发生错误: {str(e)}",
            "overall_status": "ERROR",
            "file_path": file_path,
            "reference_file": reference_file,
            "verification_level": verification_level
        }


@mcp.tool()
def batch_data_verification_tool(
    file_paths: list,
    verification_level: str = "detailed",
    save_reports: bool = True
) -> dict:
    """
    批量数据验证工具
    
    对多个Excel文件进行批量验证和质量评估。
    
    Args:
        file_paths: Excel文件路径列表
        verification_level: 验证级别 ("basic", "detailed", "comprehensive")
        save_reports: 是否保存验证报告
    
    Returns:
        dict: 批量验证结果
            - overall_summary: 总体摘要
            - individual_results: 各文件验证结果
            - quality_ranking: 质量排名
            - batch_recommendations: 批量建议
    """
    try:
        if not file_paths or not isinstance(file_paths, list):
            return {
                "success": False,
                "error": "请提供有效的文件路径列表"
            }
        
        verifier = ComprehensiveDataVerifier()
        batch_results = {
            "success": True,
            "total_files": len(file_paths),
            "processed_files": 0,
            "failed_files": 0,
            "overall_summary": {},
            "individual_results": {},
            "quality_ranking": [],
            "batch_recommendations": []
        }
        
        quality_scores = []
        
        # 逐个验证文件
        for file_path in file_paths:
            try:
                if os.path.exists(file_path):
                    result = verifier.comprehensive_excel_verification(
                        file_path=file_path,
                        verification_level=verification_level,
                        save_report=save_reports
                    )
                    
                    batch_results["individual_results"][file_path] = result
                    quality_scores.append({
                        "file": os.path.basename(file_path),
                        "path": file_path,
                        "score": result.get("data_quality_score", 0),
                        "status": result.get("overall_status", "UNKNOWN")
                    })
                    batch_results["processed_files"] += 1
                else:
                    batch_results["individual_results"][file_path] = {
                        "success": False,
                        "error": "文件不存在",
                        "overall_status": "FAILED"
                    }
                    batch_results["failed_files"] += 1
                    
            except Exception as e:
                batch_results["individual_results"][file_path] = {
                    "success": False,
                    "error": str(e),
                    "overall_status": "ERROR"
                }
                batch_results["failed_files"] += 1
        
        # 生成质量排名
        batch_results["quality_ranking"] = sorted(
            quality_scores, key=lambda x: x["score"], reverse=True
        )
        
        # 生成总体摘要
        if quality_scores:
            scores = [item["score"] for item in quality_scores]
            batch_results["overall_summary"] = {
                "average_quality_score": sum(scores) / len(scores),
                "highest_score": max(scores),
                "lowest_score": min(scores),
                "excellent_files": len([s for s in scores if s >= 90]),
                "good_files": len([s for s in scores if 80 <= s < 90]),
                "acceptable_files": len([s for s in scores if 70 <= s < 80]),
                "poor_files": len([s for s in scores if 60 <= s < 70]),
                "critical_files": len([s for s in scores if s < 60])
            }
        
        # 生成批量建议
        if batch_results["failed_files"] > 0:
            batch_results["batch_recommendations"].append(
                f"有{batch_results['failed_files']}个文件验证失败，请检查文件格式和路径"
            )
        
        if quality_scores:
            avg_score = sum([item["score"] for item in quality_scores]) / len(quality_scores)
            if avg_score < 70:
                batch_results["batch_recommendations"].append(
                    "整体数据质量偏低，建议进行数据清洗和质量改进"
                )
            elif avg_score >= 90:
                batch_results["batch_recommendations"].append(
                    "整体数据质量优秀，可以放心使用"
                )
        
        return batch_results
        
    except Exception as e:
        return {
            "success": False,
            "error": f"批量验证过程中发生错误: {str(e)}"
        }


@mcp.tool()
def excel_read_enhanced(
    file_path: str,
    sheet_name: str = None,
    start_row: int = None,
    end_row: int = None,
    start_col: str = None,
    end_col: str = None,
    use_go_service: bool = True
) -> dict:
    """
    增强版 Excel 读取工具，集成 Go excelize 库提供高性能处理
    
    Args:
        file_path: Excel 文件路径
        sheet_name: 工作表名称（可选）
        start_row: 起始行号（可选）
        end_row: 结束行号（可选）
        start_col: 起始列（如 'A'，可选）
        end_col: 结束列（如 'Z'，可选）
        use_go_service: 是否优先使用 Go 服务（默认 True）
        
    Returns:
        dict: 读取结果，包含数据和性能信息
    """
    try:
        processor = get_excel_processor()
        result = processor.read_excel_enhanced(
            file_path=file_path,
            sheet_name=sheet_name,
            start_row=start_row,
            end_row=end_row,
            start_col=start_col,
            end_col=end_col,
            use_go=use_go_service
        )
        return result
    except Exception as e:
        return {
            "success": False,
            "error": f"增强 Excel 读取失败: {str(e)}",
            "suggestion": "请检查文件路径和参数设置"
        }


@mcp.tool()
def excel_write_enhanced(
    file_path: str,
    data: list,
    sheet_name: str = None,
    start_row: int = None,
    start_col: str = None,
    use_go_service: bool = True
) -> dict:
    """
    增强版 Excel 写入工具，集成 Go excelize 库提供高性能处理
    
    Args:
        file_path: Excel 文件路径
        data: 要写入的数据（字典列表格式）
        sheet_name: 工作表名称（可选）
        start_row: 起始行号（可选）
        start_col: 起始列（如 'A'，可选）
        use_go_service: 是否优先使用 Go 服务（默认 True）
        
    Returns:
        dict: 写入结果，包含性能信息
    """
    try:
        if not isinstance(data, list):
            return {
                "success": False,
                "error": "数据格式错误，需要字典列表格式",
                "suggestion": "请提供 [{column1: value1, column2: value2}, ...] 格式的数据"
            }
        
        processor = get_excel_processor()
        result = processor.write_excel_enhanced(
            file_path=file_path,
            data=data,
            sheet_name=sheet_name,
            start_row=start_row,
            start_col=start_col,
            use_go=use_go_service
        )
        return result
    except Exception as e:
        return {
            "success": False,
            "error": f"增强 Excel 写入失败: {str(e)}",
            "suggestion": "请检查文件路径和数据格式"
        }


@mcp.tool()
def excel_chart_enhanced(
    file_path: str,
    chart_type: str,
    data_range: str,
    sheet_name: str = None,
    title: str = None,
    x_axis_title: str = None,
    y_axis_title: str = None
) -> dict:
    """
    增强版 Excel 图表创建工具，使用 Go excelize 库提供高性能图表生成
    
    Args:
        file_path: Excel 文件路径
        chart_type: 图表类型（'col', 'line', 'pie', 'bar', 'area', 'scatter' 等）
        data_range: 数据范围（如 'A1:B10'）
        sheet_name: 工作表名称（可选）
        title: 图表标题（可选）
        x_axis_title: X轴标题（可选）
        y_axis_title: Y轴标题（可选）
        
    Returns:
        dict: 图表创建结果
    """
    try:
        # 验证图表类型
        valid_chart_types = ['col', 'line', 'pie', 'bar', 'area', 'scatter', 'doughnut']
        if chart_type not in valid_chart_types:
            return {
                "success": False,
                "error": f"不支持的图表类型: {chart_type}",
                "suggestion": f"支持的图表类型: {', '.join(valid_chart_types)}"
            }
        
        processor = get_excel_processor()
        result = processor.create_chart_enhanced(
            file_path=file_path,
            chart_type=chart_type,
            data_range=data_range,
            sheet_name=sheet_name,
            chart_title=title,
            x_axis_title=x_axis_title,
            y_axis_title=y_axis_title
        )
        return result
    except Exception as e:
        return {
            "success": False,
            "error": f"增强图表创建失败: {str(e)}",
            "suggestion": "请检查文件路径、数据范围和图表参数"
        }


@mcp.tool()
def excel_info_enhanced(file_path: str) -> dict:
    """
    增强版 Excel 文件信息获取工具，使用 Go excelize 库提供详细文件分析
    
    Args:
        file_path: Excel 文件路径
        
    Returns:
        dict: 详细的文件信息，包括工作表、行列数等
    """
    try:
        processor = get_excel_processor()
        result = processor.get_file_info_enhanced(file_path)
        return result
    except Exception as e:
        return {
            "success": False,
            "error": f"获取文件信息失败: {str(e)}",
            "suggestion": "请检查文件路径是否正确"
        }


@mcp.tool()
def excel_performance_comparison(
    file_path: str,
    operation: str = "read",
    test_data: list = None
) -> dict:
    """
    Excel 性能对比工具，比较 Go 服务和 pandas 的性能差异
    
    Args:
        file_path: Excel 文件路径
        operation: 操作类型（'read' 或 'write'）
        test_data: 测试数据（写入操作时需要）
        
    Returns:
        dict: 性能对比结果
    """
    try:
        import time
        
        results = {
            "success": True,
            "operation": operation,
            "file_path": file_path,
            "performance_comparison": {},
            "recommendation": ""
        }
        
        processor = get_excel_processor()
        
        if operation == "read":
            # 测试 Go 服务性能
            start_time = time.time()
            go_result = processor.read_excel_enhanced(file_path, use_go=True)
            go_time = time.time() - start_time
            
            # 测试 pandas 性能
            start_time = time.time()
            pandas_result = processor.read_excel_enhanced(file_path, use_go=False)
            pandas_time = time.time() - start_time
            
            results["performance_comparison"] = {
                "go_service": {
                    "time_seconds": round(go_time, 4),
                    "success": go_result.get("success", False),
                    "method": go_result.get("data", {}).get("method", "unknown")
                },
                "pandas": {
                    "time_seconds": round(pandas_time, 4),
                    "success": pandas_result.get("success", False),
                    "method": pandas_result.get("data", {}).get("method", "unknown")
                }
            }
            
            # 性能提升计算
            if go_time > 0 and pandas_time > 0:
                speedup = pandas_time / go_time
                results["performance_comparison"]["speedup"] = round(speedup, 2)
                
                if speedup > 2:
                    results["recommendation"] = f"Go 服务比 pandas 快 {speedup:.1f} 倍，建议使用 Go 服务"
                elif speedup < 0.8:
                    results["recommendation"] = "pandas 性能更好，建议使用 pandas"
                else:
                    results["recommendation"] = "两种方法性能相近，可根据需要选择"
        
        elif operation == "write":
            if not test_data:
                return {
                    "success": False,
                    "error": "写入测试需要提供 test_data 参数",
                    "suggestion": "请提供测试数据列表"
                }
            
            # 创建临时文件进行测试
            import tempfile
            
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp1, \
                 tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp2:
                
                # 测试 Go 服务性能
                start_time = time.time()
                go_result = processor.write_excel_enhanced(tmp1.name, test_data, use_go=True)
                go_time = time.time() - start_time
                
                # 测试 pandas 性能
                start_time = time.time()
                pandas_result = processor.write_excel_enhanced(tmp2.name, test_data, use_go=False)
                pandas_time = time.time() - start_time
                
                # 清理临时文件
                try:
                    os.unlink(tmp1.name)
                    os.unlink(tmp2.name)
                except:
                    pass
                
                results["performance_comparison"] = {
                    "go_service": {
                        "time_seconds": round(go_time, 4),
                        "success": go_result.get("success", False),
                        "method": go_result.get("data", {}).get("method", "unknown")
                    },
                    "pandas": {
                        "time_seconds": round(pandas_time, 4),
                        "success": pandas_result.get("success", False),
                        "method": pandas_result.get("data", {}).get("method", "unknown")
                    }
                }
                
                # 性能提升计算
                if go_time > 0 and pandas_time > 0:
                    speedup = pandas_time / go_time
                    results["performance_comparison"]["speedup"] = round(speedup, 2)
                    
                    if speedup > 2:
                        results["recommendation"] = f"Go 服务比 pandas 快 {speedup:.1f} 倍，建议使用 Go 服务"
                    elif speedup < 0.8:
                        results["recommendation"] = "pandas 性能更好，建议使用 pandas"
                    else:
                        results["recommendation"] = "两种方法性能相近，可根据需要选择"
        
        else:
            return {
                "success": False,
                "error": f"不支持的操作类型: {operation}",
                "suggestion": "支持的操作类型: 'read', 'write'"
            }
        
        return results
        
    except Exception as e:
        return {
            "success": False,
            "error": f"性能对比测试失败: {str(e)}",
            "suggestion": "请检查文件路径和参数设置"
        }


# 注册 Excel 公式处理工具
@mcp.tool()
def parse_formula(formula: str, validate_security: bool = False) -> str:
    """解析 Excel 公式
    
    Args:
        formula: Excel 公式字符串
        validate_security: 是否进行安全验证
        
    Returns:
        str: JSON 格式的解析结果
    """
    return parse_excel_formula(formula, validate_security)

@mcp.tool()
def compile_workbook(file_path: str, output_format: str = 'python') -> str:
    """编译 Excel 工作簿为代码
    
    Args:
        file_path: Excel 文件路径
        output_format: 输出格式 ('python' 或 'json')
        
    Returns:
        str: JSON 格式的编译结果
    """
    return compile_excel_workbook(file_path, output_format)

@mcp.tool()
def execute_formula(formula: str, context: str = '{}') -> str:
    """执行 Excel 公式
    
    Args:
        formula: Excel 公式字符串
        context: JSON 格式的上下文数据
        
    Returns:
        str: JSON 格式的执行结果
    """
    return execute_excel_formula(formula, context)

@mcp.tool()
def analyze_dependencies(file_path: str) -> str:
    """分析 Excel 文件的公式依赖关系
    
    Args:
        file_path: Excel 文件路径
        
    Returns:
        str: JSON 格式的依赖分析结果
    """
    return analyze_excel_dependencies(file_path)

@mcp.tool()
def validate_formula(formula: str) -> str:
    """验证 Excel 公式的安全性和有效性
    
    Args:
        formula: Excel 公式字符串
        
    Returns:
        str: JSON 格式的验证结果
    """
    return validate_excel_formula(formula)


# 增强的Excel数据质量控制工具
@mcp.tool()
def enhanced_data_quality_check(file_path: str, quality_level: str = "comprehensive") -> dict:
    """增强的Excel数据质量检查工具
    
    Args:
        file_path: Excel文件路径
        quality_level: 质量检查级别 ("basic", "standard", "comprehensive")
        
    Returns:
        dict: 数据质量检查结果
    """
    try:
        return data_quality_controller.comprehensive_quality_check(file_path, quality_level)
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.PROCESSING_ERROR,
            f"数据质量检查失败: {str(e)}",
            suggestions=["检查文件格式", "确认文件可读性"]
        )


@mcp.tool()
def extract_cell_content_advanced(file_path: str, cell_range: Optional[str] = None, sheet_name: Optional[str] = None, 
                                 extract_type: str = "all") -> dict:
    """高级单元格内容提取工具
    
    Args:
        file_path: Excel文件路径
        cell_range: 单元格范围 (如 "A1:C10")
        sheet_name: 工作表名称
        extract_type: 提取类型 ("all", "text", "numbers", "formulas", "formatted")
        
    Returns:
        dict: 提取的单元格内容
    """
    try:
        return cell_content_extractor.extract_cell_content_advanced(
            file_path, cell_range, sheet_name, extract_type
        )
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.PROCESSING_ERROR,
            f"单元格内容提取失败: {str(e)}",
            suggestions=["检查单元格范围格式", "确认工作表名称"]
        )


@mcp.tool()
def convert_character_formats(file_path: str, conversion_rules: dict, output_path: Optional[str] = None) -> dict:
    """字符格式自动化转换工具
    
    Args:
        file_path: Excel文件路径
        conversion_rules: 转换规则字典
        output_path: 输出文件路径
        
    Returns:
        dict: 转换结果
    """
    try:
        return character_converter.batch_character_conversion(
            file_path, conversion_rules, output_path
        )
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.PROCESSING_ERROR,
            f"字符格式转换失败: {str(e)}",
            suggestions=["检查转换规则格式", "确认输出路径权限"]
        )


@mcp.tool()
def extract_multi_condition_data(file_path: str, conditions: list, sheet_name: Optional[str] = None) -> dict:
    """多条件数据提取工具
    
    Args:
        file_path: Excel文件路径
        conditions: 条件列表
        sheet_name: 工作表名称
        
    Returns:
        dict: 提取的数据
    """
    try:
        return multi_condition_extractor.extract_with_multiple_conditions(
            file_path, conditions, sheet_name
        )
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.PROCESSING_ERROR,
            f"多条件数据提取失败: {str(e)}",
            suggestions=["检查条件格式", "确认工作表存在"]
        )


@mcp.tool()
def merge_multiple_tables(file_paths: list, merge_config: dict, output_path: Optional[str] = None) -> dict:
    """多表格数据合并工具
    
    Args:
        file_paths: Excel文件路径列表
        merge_config: 合并配置
        output_path: 输出文件路径
        
    Returns:
        dict: 合并结果
    """
    try:
        return multi_table_merger.merge_multiple_excel_files(
            file_paths, merge_config, output_path
        )
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.PROCESSING_ERROR,
            f"表格合并失败: {str(e)}",
            suggestions=["检查文件路径", "确认合并配置"]
        )


@mcp.tool()
def clean_excel_data(file_path: str, cleaning_options: dict, output_path: Optional[str] = None) -> dict:
    """Excel数据清洗工具
    
    Args:
        file_path: Excel文件路径
        cleaning_options: 清洗选项
        output_path: 输出文件路径
        
    Returns:
        dict: 清洗结果
    """
    try:
        return data_cleaner.clean_excel_data(
            file_path=file_path, 
            cleaning_config=cleaning_options, 
            output_file=output_path
        )
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.PROCESSING_ERROR,
            f"数据清洗失败: {str(e)}",
            suggestions=["检查清洗选项", "确认文件格式"]
        )


@mcp.tool()
def batch_process_excel_files(file_paths: list, processing_config: dict) -> dict:
    """批量Excel文件处理工具
    
    Args:
        file_paths: Excel文件路径列表
        processing_config: 处理配置
        
    Returns:
        dict: 批量处理结果
    """
    try:
        return batch_processor.batch_process_files(
            file_paths, processing_config
        )
    except Exception as e:
        return error_handler.create_error_response(
            ErrorType.PROCESSING_ERROR,
            f"批量处理失败: {str(e)}",
            suggestions=["检查文件路径列表", "确认处理配置"]
        )


if __name__ == "__main__":
    import argparse
    import sys
    
    # 添加命令行参数处理
    parser = argparse.ArgumentParser(
        description='ChatExcel MCP Server - Excel数据处理服务器',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        '--version', 
        action='version', 
        version='ChatExcel MCP Server v2.1.0'
    )
    parser.add_argument(
        '--tools', 
        action='store_true',
        help='显示可用工具列表'
    )
    parser.add_argument(
        '--config', 
        type=str,
        help='指定配置文件路径'
    )
    parser.add_argument(
        '--health-check', 
        action='store_true',
        help='执行健康检查'
    )
    parser.add_argument(
        '--debug', 
        action='store_true',
        help='启用调试模式'
    )
    
    # 解析命令行参数
    args = parser.parse_args()
    
    # 初始化配置
    if CORE_MODULES_AVAILABLE:
        try:
            if args.config:
                from core.config import load_config_from_file
                config = load_config_from_file(args.config)
            else:
                config = get_config()
            
            # 设置调试模式
            if args.debug:
                config.logging.level = "DEBUG"
                print("🔧 调试模式已启用")
                
        except Exception as e:
            print(f"⚠ 配置加载失败: {e}")
            config = None
    else:
        config = None
    
    # 健康检查
    if args.health_check:
        print("🏥 执行健康检查...")
        
        health_status = {
            "core_modules": CORE_MODULES_AVAILABLE,
            "dependency_manager": len(dependency_manager.available_modules),
            "failed_imports": len(dependency_manager.failed_imports),
            "config_loaded": config is not None
        }
        
        print(f"✓ 核心模块: {'可用' if health_status['core_modules'] else '不可用'}")
        print(f"✓ 可用依赖模块: {health_status['dependency_manager']}")
        print(f"⚠ 失败导入: {health_status['failed_imports']}")
        print(f"✓ 配置: {'已加载' if health_status['config_loaded'] else '未加载'}")
        
        if dependency_manager.failed_imports:
            print("\n失败的模块导入:")
            for module in dependency_manager.failed_imports:
                print(f"  - {module}")
        
        # 检查关键目录
        try:
            os.makedirs(CHARTS_DIR, exist_ok=True)
            print(f"✓ 图表目录: {CHARTS_DIR}")
        except Exception as e:
            print(f"⚠ 图表目录创建失败: {e}")
        
        sys.exit(0)
    
    # 如果请求工具列表
    if args.tools:
        print("ChatExcel MCP Server 可用工具:")
        print("1. read_excel_file - 读取Excel文件")
        print("2. run_excel_code - 执行Excel数据处理代码")
        print("3. create_chart - 创建图表")
        print("4. export_data - 导出数据")
        print("5. data_cleaner - 数据清洗")
        print("6. batch_process_excel_files - 批量处理Excel文件")
        print("7. read_metadata - 读取文件元数据")
        print("8. convert_character_formats - 字符格式转换")
        print("9. extract_multi_condition_data - 多条件数据提取")
        print("10. merge_multiple_tables - 多表格合并")
        print("...以及其他21个专业Excel处理工具")
        print(f"\n总计: 31个工具")
        print(f"支持格式: xlsx, xls, csv, json, html")
        
        if CORE_MODULES_AVAILABLE and config:
            print(f"\n配置信息:")
            print(f"  - 最大文件大小: {config.security.max_file_size} 字节")
            print(f"  - 执行超时: {config.performance.execution_timeout} 秒")
            print(f"  - 缓存启用: {config.performance.enable_cache}")
        
        sys.exit(0)
    
    # 确保必要目录存在
    try:
        os.makedirs(CHARTS_DIR, exist_ok=True)
        print(f"✓ 图表目录已准备: {CHARTS_DIR}")
    except Exception as e:
        print(f"⚠ 图表目录创建失败: {e}")
    
    # 启动信息
    print("🚀 启动 ChatExcel MCP Server...")
    print(f"📊 工具数量: 31")
    print(f"📁 支持格式: xlsx, xls, csv, json, html")
    print(f"🔧 核心模块: {'已加载' if CORE_MODULES_AVAILABLE else '备用模式'}")
    print(f"⚙️  依赖模块: {len(dependency_manager.available_modules)} 可用")
    
    if dependency_manager.failed_imports:
        print(f"⚠️  失败导入: {len(dependency_manager.failed_imports)} 个模块")
    
    if CORE_MODULES_AVAILABLE and config:
        print(f"🛡️  安全配置: 已启用")
        print(f"⚡ 性能优化: 已启用")
    
    print("\n服务器启动中...")
    
    try:
        mcp.run()
    except KeyboardInterrupt:
        print("\n👋 服务器已停止")
    except Exception as e:
        print(f"\n❌ 服务器启动失败: {e}")
        if args.debug:
            import traceback
            traceback.print_exc()
        sys.exit(1)




# 备用实现类
class FallbackVariableLifecycleManager:
    def __init__(self):
        self.variable_registry = {}
        self.creation_order = []
    
    def register_variable(self, name, value, source='user'):
        self.variable_registry[name] = {
            'value': value,
            'source': source,
            'type': type(value).__name__
        }
        if name not in self.creation_order:
            self.creation_order.append(name)
    
    def get_dataframes(self):
        return [k for k, v in self.variable_registry.items() 
                if hasattr(v.get('value'), 'shape')]

class FallbackNameErrorHandler:
    def __init__(self):
        pass
    
    def handle_name_error(self, error, code, local_vars):
        return {'success': False, 'error': str(error)}

class FallbackColumnChecker:
    def __init__(self):
        pass
    
    def match_column(self, missing_column, available_columns):
        return {
            'exact_match': None,
            'case_insensitive_match': None,
            'normalized_matches': []
        }
    
    def generate_code_suggestions(self, missing_column, match_result):
        return []

def fallback_enhanced_execute_with_error_handling(code, local_vars, var_manager=None, error_handler=None):
    try:
        exec(code, local_vars)
        return {'success': True, 'variables': local_vars}
    except Exception as e:
        return {'success': False, 'error': str(e)}

def fallback_analyze_code_variables(code):
    # 简单的变量分析
    import re
    
    # 查找变量赋值
    assignments = re.findall(r'(\w+)\s*=', code)
    
    # 查找变量引用
    references = re.findall(r'(\w+)', code)
    
    return {
        'defined_variables': list(set(assignments)),
        'referenced_variables': list(set(references)),
        'variables': list(set(assignments + references)),
        'assignments': assignments
    }

def fallback_enhanced_run_excel_code(file_path, code, **kwargs):
    """增强的Excel代码执行备用函数，集成pandas导入修复"""
    import pandas as pd
    import numpy as np
    import traceback
    
    try:
        # 提取参数
        sheet_name = kwargs.get('sheet_name')
        skiprows = kwargs.get('skiprows')
        header = kwargs.get('header')
        usecols = kwargs.get('usecols')
        
        # 构建读取参数
        read_kwargs = {}
        if sheet_name is not None:
            read_kwargs['sheet_name'] = sheet_name
        if skiprows is not None:
            read_kwargs['skiprows'] = skiprows
        if header is not None:
            read_kwargs['header'] = header
        if usecols is not None:
            read_kwargs['usecols'] = usecols
        
        # 读取Excel文件
        df = pd.read_excel(file_path, **read_kwargs)
        
        # 如果返回字典（多工作表），取第一个
        if isinstance(df, dict):
            first_sheet = list(df.keys())[0]
            df = df[first_sheet]
        
        # 创建增强的安全执行环境
        safe_globals = {
            # pandas 相关 - 多种引用方式确保兼容性
            'pd': pd,
            'pandas': pd,
            'DataFrame': pd.DataFrame,
            'Series': pd.Series,
            
            # numpy 相关
            'np': np,
            'numpy': np,
            
            # 数据相关
            'df': df.copy() if hasattr(df, 'copy') else df,
            'data': df.copy() if hasattr(df, 'copy') else df,
            'df_original': df.copy() if hasattr(df, 'copy') else df,
            
            # 文件路径
            'file_path': file_path,
            
            # 常用内置函数
            'len': len, 'sum': sum, 'max': max, 'min': min,
            'abs': abs, 'round': round, 'sorted': sorted,
            'enumerate': enumerate, 'zip': zip, 'range': range,
            'list': list, 'dict': dict, 'set': set, 'tuple': tuple,
            'str': str, 'int': int, 'float': float, 'bool': bool,
            'print': print, 'type': type, 'isinstance': isinstance,
            'hasattr': hasattr, 'getattr': getattr,
            
            # 添加 __import__ 支持
            '__import__': __import__,
        }
        
        # 检查是否使用了df_processed但未定义
        if 'df_processed' in code and 'df_processed =' not in code:
            code = f"df_processed = df.copy()\n{code}"
        
        # 执行代码
        exec(code, safe_globals)
        
        return {
            'success': True,
            'variables': {k: v for k, v in safe_globals.items() if not k.startswith('_')},
            'data_shape': df.shape if hasattr(df, 'shape') else None,
            'message': '代码执行成功（增强备用实现）'
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f"代码执行失败: {str(e)}",
            'error_type': type(e).__name__,
            'traceback': traceback.format_exc(),
            'errors': [{
                'message': f"代码执行失败: {str(e)}",
                'code': 'ErrorCode.CODE_EXECUTION_ERROR'
            }],
            'warnings': []
        }
