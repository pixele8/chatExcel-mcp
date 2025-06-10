"""Excel文件智能分析和参数推荐辅助函数"""
import pandas as pd
import openpyxl
from typing import Dict, Any, List, Optional

def _suggest_excel_read_parameters(file_path: str, sheet_name: str = None) -> Dict[str, Any]:
    """
    分析Excel文件结构并建议最佳的读取参数
    使用增强的多级列头检测器
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        
    Returns:
        包含建议参数的字典
    """
    try:
        # 导入增强检测器
        from .enhanced_multiheader_detector import enhanced_suggest_excel_read_parameters
        
        # 使用增强检测器
        result = enhanced_suggest_excel_read_parameters(file_path, sheet_name)
        
        # 如果增强检测器成功，直接返回结果
        if "error" not in result.get("analysis", {}):
            return result
        
        # 如果增强检测器失败，回退到简化版本
        return _fallback_suggest_parameters(file_path, sheet_name)
        
    except ImportError:
        # 如果无法导入增强检测器，使用回退方案
        return _fallback_suggest_parameters(file_path, sheet_name)
    except Exception as e:
        # 其他错误也使用回退方案
        return _fallback_suggest_parameters(file_path, sheet_name)

def _fallback_suggest_parameters(file_path: str, sheet_name: str = None) -> Dict[str, Any]:
    """
    回退的参数建议方案（简化版本）
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        
    Returns:
        包含建议参数的字典
    """
    try:
        # 读取前10行数据进行分析
        sample_df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=10)
        
        suggestions = {
            "recommended_params": {},
            "analysis": {
                "empty_rows": [],
                "potential_headers": [],
                "multi_level_header_detected": False,
                "fallback_mode": True
            },
            "warnings": ["使用简化检测模式"],
            "tips": []
        }
        
        # 检测空行
        empty_rows = []
        for i, row in sample_df.iterrows():
            if row.isna().all():
                empty_rows.append(i)
        
        suggestions["analysis"]["empty_rows"] = empty_rows
        
        # 检测潜在的标题行
        potential_headers = []
        for i, row in sample_df.iterrows():
            if not row.isna().all():
                # 检查是否主要包含文本（可能是标题）
                text_count = sum(1 for val in row if pd.notna(val) and isinstance(val, str))
                total_count = sum(1 for val in row if pd.notna(val))
                if total_count > 0 and text_count / total_count > 0.5:
                    potential_headers.append(i)
        
        suggestions["analysis"]["potential_headers"] = potential_headers
        
        # 处理skiprows
        if empty_rows and empty_rows[0] == 0:
            consecutive_empty = 0
            for i in empty_rows:
                if i == consecutive_empty:
                    consecutive_empty += 1
                else:
                    break
            if consecutive_empty > 0:
                suggestions["recommended_params"]["skiprows"] = consecutive_empty
                suggestions["tips"].append(f"检测到前{consecutive_empty}行为空行，建议跳过")
        
        # 简化的单级列头处理
        if potential_headers:
            header_row = potential_headers[-1]
            skiprows = suggestions["recommended_params"].get("skiprows", 0)
            adjusted_header = max(0, header_row - skiprows)
            suggestions["recommended_params"]["header"] = adjusted_header
            suggestions["tips"].append(f"使用第{header_row + 1}行作为列头")
        else:
            suggestions["recommended_params"]["header"] = 0
            suggestions["tips"].append("使用第1行作为列头")
        
        return suggestions
        
    except Exception as e:
        # 如果分析失败，返回默认建议
        return {
            "recommended_params": {"header": 0},
            "analysis": {
                "multi_level_header_detected": False,
                "error": str(e),
                "fallback_mode": True
            },
            "warnings": [f"文件分析失败，使用默认参数: {str(e)}"],
            "tips": ["建议手动检查文件格式"]
        }

def detect_excel_structure(file_path: str, sheet_name: str = None) -> Dict[str, Any]:
    """检测Excel文件的详细结构信息
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        
    Returns:
        dict: 详细的结构信息
    """
    structure_info = {
        "sheets": [],
        "merged_cells": [],
        "data_range": {},
        "formatting_info": {}
    }
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        
        # 获取所有工作表信息
        for sheet in workbook.worksheets:
            sheet_info = {
                "name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "merged_cells": [str(merged_range) for merged_range in sheet.merged_cells.ranges]
            }
            structure_info["sheets"].append(sheet_info)
        
        # 如果指定了工作表，获取详细信息
        if sheet_name and sheet_name in [s.title for s in workbook.worksheets]:
            target_sheet = workbook[sheet_name]
            structure_info["merged_cells"] = [str(merged_range) for merged_range in target_sheet.merged_cells.ranges]
            structure_info["data_range"] = {
                "min_row": target_sheet.min_row,
                "max_row": target_sheet.max_row,
                "min_column": target_sheet.min_column,
                "max_column": target_sheet.max_column
            }
        
        workbook.close()
        
    except Exception as e:
        structure_info["error"] = str(e)
    
    return structure_info