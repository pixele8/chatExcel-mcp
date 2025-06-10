#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel复杂格式处理能力测试脚本
测试项目对Excel文件的高级格式处理能力
"""

import pandas as pd
import openpyxl
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
import os
import tempfile
import json

def create_complex_excel_test_file():
    """创建包含复杂格式的Excel测试文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "复杂格式测试"
    
    # 创建多级列头
    # 第一行：主分类
    ws['A1'] = '基本信息'
    ws['D1'] = '销售数据'
    ws['G1'] = '财务指标'
    
    # 第二行：子分类
    ws['A2'] = '姓名'
    ws['B2'] = '部门'
    ws['C2'] = '职位'
    ws['D2'] = 'Q1销售额'
    ws['E2'] = 'Q2销售额'
    ws['F2'] = 'Q3销售额'
    ws['G2'] = '利润率'
    ws['H2'] = '成本率'
    
    # 合并单元格（模拟多级列头）
    ws.merge_cells('A1:C1')  # 基本信息
    ws.merge_cells('D1:F1')  # 销售数据
    ws.merge_cells('G1:H1')  # 财务指标
    
    # 添加样式
    for cell in ['A1', 'D1', 'G1']:
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal='center')
    
    # 添加测试数据
    test_data = [
        ['张三', '销售部', '经理', 120000, 135000, 142000, 0.15, 0.65],
        ['李四', '技术部', '工程师', 0, 0, 0, 0, 0.8],
        ['王五', '市场部', '专员', 85000, 92000, 88000, 0.12, 0.7],
        ['赵六', '财务部', '会计', 0, 0, 0, 0, 0.75],
        ['钱七', '销售部', '代表', 65000, 71000, 69000, 0.1, 0.72]
    ]
    
    for i, row_data in enumerate(test_data, start=3):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)
    
    # 保存文件
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(temp_file.name)
    temp_file.close()
    return temp_file.name

def test_basic_excel_reading(file_path):
    """测试基本Excel读取功能"""
    print("\n=== 基本Excel读取测试 ===")
    try:
        # 默认读取
        df_default = pd.read_excel(file_path)
        print(f"默认读取 - 形状: {df_default.shape}")
        print(f"列名: {list(df_default.columns)}")
        print(f"前3行数据:\n{df_default.head(3)}")
        return True
    except Exception as e:
        print(f"基本读取失败: {e}")
        return False

def test_skiprows_functionality(file_path):
    """测试跳过行功能（处理多级列头）"""
    print("\n=== 跳过行功能测试 ===")
    try:
        # 跳过第一行（主分类行）
        df_skip1 = pd.read_excel(file_path, skiprows=1)
        print(f"跳过1行 - 形状: {df_skip1.shape}")
        print(f"列名: {list(df_skip1.columns)}")
        
        # 跳过前两行，从数据行开始
        df_skip2 = pd.read_excel(file_path, skiprows=2)
        print(f"跳过2行 - 形状: {df_skip2.shape}")
        print(f"列名: {list(df_skip2.columns)}")
        
        return True
    except Exception as e:
        print(f"跳过行功能测试失败: {e}")
        return False

def test_header_specification(file_path):
    """测试指定列头行功能"""
    print("\n=== 指定列头行测试 ===")
    try:
        # 指定第2行为列头
        df_header1 = pd.read_excel(file_path, header=1)
        print(f"第2行为列头 - 形状: {df_header1.shape}")
        print(f"列名: {list(df_header1.columns)}")
        
        # 多级列头处理
        df_multiheader = pd.read_excel(file_path, header=[0, 1])
        print(f"多级列头 - 形状: {df_multiheader.shape}")
        print(f"列名结构: {df_multiheader.columns}")
        
        return True
    except Exception as e:
        print(f"指定列头行测试失败: {e}")
        return False

def test_usecols_functionality(file_path):
    """测试选择特定列功能"""
    print("\n=== 选择特定列测试 ===")
    try:
        # 按列索引选择
        df_cols_idx = pd.read_excel(file_path, usecols=[0, 1, 3, 4], skiprows=1)
        print(f"按索引选择列 - 形状: {df_cols_idx.shape}")
        print(f"列名: {list(df_cols_idx.columns)}")
        
        # 按列名选择（需要先知道列名）
        df_temp = pd.read_excel(file_path, skiprows=1)
        selected_cols = [df_temp.columns[0], df_temp.columns[1], df_temp.columns[3]]
        df_cols_name = pd.read_excel(file_path, usecols=selected_cols, skiprows=1)
        print(f"按列名选择 - 形状: {df_cols_name.shape}")
        print(f"列名: {list(df_cols_name.columns)}")
        
        return True
    except Exception as e:
        print(f"选择特定列测试失败: {e}")
        return False

def test_encoding_detection(file_path):
    """测试编码检测功能"""
    print("\n=== 编码检测测试 ===")
    try:
        # 使用openpyxl引擎
        df_openpyxl = pd.read_excel(file_path, engine='openpyxl')
        print(f"openpyxl引擎 - 形状: {df_openpyxl.shape}")
        
        # 检查中文字符处理
        chinese_cols = [col for col in df_openpyxl.columns if any('\u4e00' <= char <= '\u9fff' for char in str(col))]
        print(f"包含中文的列: {chinese_cols}")
        
        return True
    except Exception as e:
        print(f"编码检测测试失败: {e}")
        return False

def test_data_type_detection(file_path):
    """测试数据类型检测"""
    print("\n=== 数据类型检测测试 ===")
    try:
        df = pd.read_excel(file_path, skiprows=1)
        print("数据类型信息:")
        for col in df.columns:
            dtype = df[col].dtype
            null_count = df[col].isnull().sum()
            unique_count = df[col].nunique()
            print(f"  {col}: {dtype}, 空值: {null_count}, 唯一值: {unique_count}")
        
        return True
    except Exception as e:
        print(f"数据类型检测失败: {e}")
        return False

def generate_enhancement_recommendations():
    """生成Excel处理能力增强建议"""
    recommendations = {
        "当前支持的功能": [
            "基本Excel文件读取(.xlsx格式)",
            "openpyxl引擎支持",
            "基本数据类型检测",
            "文件大小限制检查",
            "编码处理(UTF-8)"
        ],
        "需要增强的功能": [
            "多级列头自动识别和处理",
            "合并单元格智能解析",
            "skiprows参数动态配置",
            "header参数多级支持",
            "usecols高级选择功能",
            "xlrd引擎支持(.xls格式)",
            "数据验证和清洗建议",
            "复杂表格结构自动识别"
        ],
        "建议的代码增强": {
            "read_excel_metadata函数": [
                "添加skiprows参数支持",
                "添加header参数支持",
                "添加usecols参数支持",
                "增加多级列头检测逻辑",
                "增加合并单元格检测"
            ],
            "run_excel_code函数": [
                "支持更多pandas.read_excel参数",
                "添加数据预处理建议",
                "增加复杂格式处理示例"
            ],
            "新增功能建议": [
                "create_excel_template工具",
                "detect_excel_structure工具",
                "suggest_read_parameters工具"
            ]
        },
        "依赖库状态": {
            "pandas": "✓ 已安装，版本充足",
            "openpyxl": "✓ 已安装，版本充足",
            "xlrd": "⚠ 未明确安装，建议添加",
            "xlsxwriter": "⚠ 未安装，建议添加用于写入"
        }
    }
    return recommendations

def main():
    """主测试函数"""
    print("Excel复杂格式处理能力测试")
    print("=" * 50)
    
    # 创建测试文件
    test_file = create_complex_excel_test_file()
    print(f"测试文件已创建: {test_file}")
    
    try:
        # 执行各项测试
        tests = [
            test_basic_excel_reading,
            test_skiprows_functionality,
            test_header_specification,
            test_usecols_functionality,
            test_encoding_detection,
            test_data_type_detection
        ]
        
        results = {}
        for test_func in tests:
            test_name = test_func.__name__
            try:
                result = test_func(test_file)
                results[test_name] = "通过" if result else "失败"
            except Exception as e:
                results[test_name] = f"错误: {e}"
        
        # 输出测试结果
        print("\n" + "=" * 50)
        print("测试结果汇总:")
        for test_name, result in results.items():
            status = "✓" if result == "通过" else "✗"
            print(f"  {status} {test_name}: {result}")
        
        # 生成增强建议
        print("\n" + "=" * 50)
        print("Excel处理能力分析和建议:")
        recommendations = generate_enhancement_recommendations()
        print(json.dumps(recommendations, ensure_ascii=False, indent=2))
        
    finally:
        # 清理测试文件
        if os.path.exists(test_file):
            os.unlink(test_file)
            print(f"\n测试文件已清理: {test_file}")

if __name__ == "__main__":
    main()