#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
增强多级列头检测系统的全面测试
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys
from typing import Dict, Any

# 添加当前目录到Python路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from excel_helper import _suggest_excel_read_parameters
from enhanced_excel_helper import smart_read_excel
from server import run_excel_code

def create_test_files():
    """创建各种类型的测试Excel文件"""
    
    # 1. 简单单级列头文件
    print("创建简单单级列头文件...")
    df_simple = pd.DataFrame({
        '姓名': ['张三', '李四', '王五'],
        '年龄': [25, 30, 35],
        '城市': ['北京', '上海', '广州']
    })
    df_simple.to_excel('test_simple_header.xlsx', index=False)
    
    # 2. 真正的多级列头文件（手动创建）
    print("创建真正的多级列头文件...")
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 第一级列头（合并单元格）
    ws.merge_cells('A1:B1')
    ws['A1'] = '个人信息'
    ws.merge_cells('C1:D1')
    ws['C1'] = '工作信息'
    
    # 第二级列头
    ws['A2'] = '姓名'
    ws['B2'] = '年龄'
    ws['C2'] = '公司'
    ws['D2'] = '职位'
    
    # 数据行
    data = [
        ['张三', 25, 'ABC公司', '工程师'],
        ['李四', 30, 'XYZ公司', '经理'],
        ['王五', 35, 'DEF公司', '总监']
    ]
    
    for i, row in enumerate(data, start=3):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    
    wb.save('test_true_multiheader.xlsx')
    wb.close()
    
    # 3. 复杂格式文件（标题+空行+多级列头）
    print("创建复杂格式文件...")
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 标题行
    ws['A1'] = '员工信息统计表'
    # 空行
    # 第一级列头
    ws.merge_cells('A3:B3')
    ws['A3'] = '基本信息'
    ws.merge_cells('C3:E3')
    ws['C3'] = '详细信息'
    
    # 第二级列头
    ws['A4'] = '姓名'
    ws['B4'] = '性别'
    ws['C4'] = '部门'
    ws['D4'] = '薪资'
    ws['E4'] = '入职日期'
    
    # 数据
    complex_data = [
        ['张三', '男', '技术部', 8000, '2020-01-01'],
        ['李四', '女', '销售部', 7000, '2020-02-01'],
        ['王五', '男', '人事部', 6000, '2020-03-01']
    ]
    
    for i, row in enumerate(complex_data, start=5):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    
    wb.save('test_complex_format.xlsx')
    wb.close()
    
    # 4. 伪多级列头文件（看起来像多级但实际不是）
    print("创建伪多级列头文件...")
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 第一行：看起来像标题但实际是独立的列头
    ws['A1'] = '产品名称'
    ws['B1'] = '价格'
    ws['C1'] = '库存'
    
    # 第二行：另一组独立的列头
    ws['A2'] = '供应商'
    ws['B2'] = '联系电话'
    ws['C2'] = '地址'
    
    # 数据（两个不同的数据集）
    ws['A3'] = '苹果'
    ws['B3'] = 5.0
    ws['C3'] = 100
    
    ws['A4'] = 'ABC供应商'
    ws['B4'] = '13800138000'
    ws['C4'] = '北京市朝阳区'
    
    wb.save('test_pseudo_multiheader.xlsx')
    wb.close()
    
    # 5. 极端情况：大量空行和不规则结构
    print("创建极端情况文件...")
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 前5行空行
    # 第6行：标题
    ws['A6'] = '数据报告'
    # 第7-8行空行
    # 第9行：列头
    ws['A9'] = 'ID'
    ws['B9'] = '名称'
    ws['C9'] = '数值'
    
    # 数据
    extreme_data = [
        [1, '项目A', 100],
        [2, '项目B', 200],
        [3, '项目C', 300]
    ]
    
    for i, row in enumerate(extreme_data, start=10):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    
    wb.save('test_extreme_case.xlsx')
    wb.close()
    
    print("所有测试文件创建完成！")

def test_parameter_suggestion(file_path: str, expected_type: str) -> Dict[str, Any]:
    """测试参数建议功能"""
    print(f"\n测试文件: {file_path}")
    print(f"期望类型: {expected_type}")
    
    try:
        suggestions = _suggest_excel_read_parameters(file_path)
        
        print(f"推荐参数: {suggestions['recommended_params']}")
        print(f"分析结果: {suggestions['analysis']}")
        print(f"警告: {suggestions['warnings']}")
        print(f"提示: {suggestions['tips']}")
        
        return {
            'success': True,
            'suggestions': suggestions,
            'file': file_path,
            'expected_type': expected_type
        }
        
    except Exception as e:
        print(f"参数建议失败: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'file': file_path,
            'expected_type': expected_type
        }

def test_smart_read(file_path: str, expected_type: str) -> Dict[str, Any]:
    """测试智能读取功能"""
    print(f"\n智能读取测试: {file_path}")
    
    try:
        result = smart_read_excel(file_path)
        
        if result['success'] and result['dataframe'] is not None:
            df = result['dataframe']
            print(f"数据形状: {df.shape}")
            print(f"列名: {list(df.columns)}")
            print(f"列名类型: {type(df.columns)}")
            
            # 检查是否为MultiIndex
            is_multi_index = hasattr(df.columns, 'levels')
            print(f"是否为MultiIndex: {is_multi_index}")
            
            if df.shape[0] > 0:
                print(f"前几行数据:\n{df.head(2)}")
            
            if result['warnings']:
                print(f"警告: {result['warnings']}")
            
            return {
                'success': True,
                'shape': df.shape,
                'columns': list(df.columns),
                'is_multi_index': is_multi_index,
                'file': file_path,
                'expected_type': expected_type
            }
        else:
            print(f"智能读取失败: {result['errors']}")
            return {
                'success': False,
                'error': result['errors'],
                'file': file_path,
                'expected_type': expected_type
            }
        
    except Exception as e:
        print(f"智能读取失败: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'file': file_path,
            'expected_type': expected_type
        }

def test_run_excel_code(file_path: str, expected_type: str) -> Dict[str, Any]:
    """测试run_excel_code函数"""
    print(f"\nrun_excel_code测试: {file_path}")
    
    try:
        # 测试基本的pandas操作
        code = f"""
import pandas as pd
df = pd.read_excel('{file_path}')
print(f"数据形状: {{df.shape}}")
print(f"列名: {{list(df.columns)}}")
print(f"前3行数据:")
print(df.head(3))
result = df.shape
"""
        
        result = run_excel_code(file_path, code)
        
        if 'output' in result:
            print(f"执行成功")
            print(f"输出: {result['output'][:500]}...")  # 只显示前500字符
            return {
                'success': True,
                'output': result['output'],
                'file': file_path,
                'expected_type': expected_type
            }
        else:
            print(f"执行失败: {result.get('error', '未知错误')}")
            return {
                'success': False,
                'error': result.get('error', '未知错误'),
                'file': file_path,
                'expected_type': expected_type
            }
        
    except Exception as e:
        print(f"run_excel_code测试失败: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'file': file_path,
            'expected_type': expected_type
        }

def analyze_results(results: Dict[str, list]):
    """分析测试结果"""
    print("\n" + "="*60)
    print("测试结果分析")
    print("="*60)
    
    for test_type, test_results in results.items():
        print(f"\n{test_type}:")
        success_count = sum(1 for r in test_results if r['success'])
        total_count = len(test_results)
        print(f"  成功率: {success_count}/{total_count} ({success_count/total_count*100:.1f}%)")
        
        # 按文件类型分析
        by_type = {}
        for result in test_results:
            file_type = result['expected_type']
            if file_type not in by_type:
                by_type[file_type] = {'success': 0, 'total': 0}
            by_type[file_type]['total'] += 1
            if result['success']:
                by_type[file_type]['success'] += 1
        
        for file_type, stats in by_type.items():
            success_rate = stats['success'] / stats['total'] * 100
            print(f"    {file_type}: {stats['success']}/{stats['total']} ({success_rate:.1f}%)")
        
        # 显示失败的测试
        failed_tests = [r for r in test_results if not r['success']]
        if failed_tests:
            print(f"  失败的测试:")
            for failed in failed_tests:
                print(f"    - {failed['file']}: {failed.get('error', '未知错误')}")

def cleanup_test_files():
    """清理测试文件"""
    test_files = [
        'test_simple_header.xlsx',
        'test_true_multiheader.xlsx',
        'test_complex_format.xlsx',
        'test_pseudo_multiheader.xlsx',
        'test_extreme_case.xlsx'
    ]
    
    for file in test_files:
        try:
            if os.path.exists(file):
                os.remove(file)
                print(f"已删除: {file}")
        except Exception as e:
            print(f"删除文件失败 {file}: {str(e)}")

def main():
    """主测试函数"""
    print("增强多级列头检测系统 - 全面测试")
    print("="*60)
    
    # 创建测试文件
    create_test_files()
    
    # 定义测试用例
    test_cases = [
        ('test_simple_header.xlsx', '简单单级列头'),
        ('test_true_multiheader.xlsx', '真正多级列头'),
        ('test_complex_format.xlsx', '复杂格式'),
        ('test_pseudo_multiheader.xlsx', '伪多级列头'),
        ('test_extreme_case.xlsx', '极端情况')
    ]
    
    # 存储所有测试结果
    results = {
        '参数建议测试': [],
        '智能读取测试': [],
        'run_excel_code测试': []
    }
    
    # 运行所有测试
    for file_path, expected_type in test_cases:
        if os.path.exists(file_path):
            # 测试参数建议
            param_result = test_parameter_suggestion(file_path, expected_type)
            results['参数建议测试'].append(param_result)
            
            # 测试智能读取
            smart_result = test_smart_read(file_path, expected_type)
            results['智能读取测试'].append(smart_result)
            
            # 测试run_excel_code
            code_result = test_run_excel_code(file_path, expected_type)
            results['run_excel_code测试'].append(code_result)
        else:
            print(f"警告: 测试文件 {file_path} 不存在")
    
    # 分析结果
    analyze_results(results)
    
    # 清理测试文件
    print("\n清理测试文件...")
    cleanup_test_files()
    
    print("\n测试完成！")

if __name__ == "__main__":
    main()