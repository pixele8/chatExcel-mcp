#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试多级列头处理能力
"""

import pandas as pd
import os
from server import run_excel_code
from excel_helper import _suggest_excel_read_parameters
from enhanced_excel_helper import smart_read_excel

def create_simple_header_file():
    """创建简单单级列头文件"""
    data = {
        '姓名': ['张三', '李四', '王五'],
        '年龄': [25, 30, 35],
        '薪资': [8000, 12000, 15000]
    }
    df = pd.DataFrame(data)
    file_path = 'simple_header.xlsx'
    df.to_excel(file_path, index=False)
    return file_path

def create_multiheader_file():
    """创建多级列头文件（手动方式）"""
    import openpyxl
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    
    # 创建多级列头结构
    # 第一级标题
    ws['A1'] = '基本信息'
    ws['C1'] = '工作信息'
    ws['E1'] = '联系方式'
    
    # 第二级标题
    ws['A2'] = '姓名'
    ws['B2'] = '年龄'
    ws['C2'] = '部门'
    ws['D2'] = '薪资'
    ws['E2'] = '电话'
    ws['F2'] = '邮箱'
    
    # 合并单元格
    ws.merge_cells('A1:B1')  # 基本信息
    ws.merge_cells('C1:D1')  # 工作信息
    ws.merge_cells('E1:F1')  # 联系方式
    
    # 添加数据
    data = [
        ['张三', 25, '技术部', 8000, '13800138000', 'zhangsan@example.com'],
        ['李四', 30, '销售部', 12000, '13900139000', 'lisi@example.com'],
        ['王五', 35, '管理部', 15000, '13700137000', 'wangwu@example.com']
    ]
    
    for i, row in enumerate(data, start=3):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    
    file_path = 'multiheader.xlsx'
    wb.save(file_path)
    return file_path

def create_complex_header_file():
    """创建复杂格式文件（有标题行和空行）"""
    import openpyxl
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    
    # 添加标题
    ws['A1'] = '员工信息统计表'
    ws['A2'] = '2024年度'
    # 空行
    # 列头在第4行
    ws['A4'] = '姓名'
    ws['B4'] = '年龄'
    ws['C4'] = '薪资'
    
    # 数据从第5行开始
    data = [
        ['张三', 25, 8000],
        ['李四', 30, 12000],
        ['王五', 35, 15000]
    ]
    
    for i, row in enumerate(data, start=5):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    
    file_path = 'complex_header.xlsx'
    wb.save(file_path)
    return file_path

def test_parameter_suggestion(file_path, file_type):
    """测试参数建议功能"""
    print(f"\n📋 测试 {file_type} 的参数建议:")
    try:
        suggestions = _suggest_excel_read_parameters(file_path)
        recommended_params = suggestions.get('recommended_params', {})
        is_multi_level = suggestions.get('analysis', {}).get('multi_level_header_detected', False)
        
        print(f"  建议参数: {recommended_params}")
        print(f"  多级列头检测: {is_multi_level}")
        
        return recommended_params, is_multi_level
    except Exception as e:
        print(f"  ❌ 参数建议失败: {e}")
        return {}, False

def test_smart_read(file_path, file_type):
    """测试智能读取功能"""
    print(f"\n📖 测试 {file_type} 的智能读取:")
    try:
        # 测试自动检测模式
        result1 = smart_read_excel(file_path, auto_detect_params=True)
        if result1['success']:
            df1 = result1['dataframe']
            print(f"  自动检测: ✅ 形状={df1.shape}, 列名={list(df1.columns)[:3]}...")
        else:
            print(f"  自动检测: ❌ {result1.get('errors', [])}")
        
        # 测试手动header=0
        result2 = smart_read_excel(file_path, auto_detect_params=False, header=0)
        if result2['success']:
            df2 = result2['dataframe']
            print(f"  header=0: ✅ 形状={df2.shape}, 列名={list(df2.columns)[:3]}...")
        else:
            print(f"  header=0: ❌ {result2.get('errors', [])}")
            
        return result1['success'] or result2['success']
    except Exception as e:
        print(f"  ❌ 智能读取失败: {e}")
        return False

def test_run_excel_code(file_path, file_type):
    """测试run_excel_code功能"""
    print(f"\n🔧 测试 {file_type} 的run_excel_code:")
    try:
        code = '''
print(f"数据形状: {df.shape}")
print(f"列名: {list(df.columns)}")
if len(df) > 0:
    print(f"第一行数据: {df.iloc[0].to_dict()}")
    
result = {
    'shape': df.shape,
    'columns': list(df.columns),
    'has_data': len(df) > 0
}
'''
        
        response = run_excel_code(file_path, code)
        
        if 'output' in response and response['output']:
            print(f"  ✅ 执行成功")
            print(f"  输出: {response['output'][:100]}...")
            if 'result' in response:
                print(f"  结果: {response['result']}")
            return True
        else:
            print(f"  ❌ 执行失败")
            if 'error' in response:
                print(f"  错误: {response['error']}")
            return False
            
    except Exception as e:
        print(f"  ❌ run_excel_code失败: {e}")
        return False

def main():
    """主函数"""
    print("🧪 多级列头处理能力测试")
    print("=" * 60)
    
    test_files = [
        (create_simple_header_file, "简单单级列头"),
        (create_multiheader_file, "真正多级列头"),
        (create_complex_header_file, "复杂格式文件")
    ]
    
    results = []
    
    for create_func, file_type in test_files:
        print(f"\n{'='*20} {file_type} {'='*20}")
        
        file_path = create_func()
        
        try:
            # 测试参数建议
            params, is_multi = test_parameter_suggestion(file_path, file_type)
            
            # 测试智能读取
            smart_read_ok = test_smart_read(file_path, file_type)
            
            # 测试run_excel_code
            run_code_ok = test_run_excel_code(file_path, file_type)
            
            results.append({
                'type': file_type,
                'params': params,
                'is_multi': is_multi,
                'smart_read': smart_read_ok,
                'run_code': run_code_ok,
                'overall': smart_read_ok and run_code_ok
            })
            
        finally:
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"\n🧹 已清理: {file_path}")
    
    # 总结结果
    print(f"\n{'='*60}")
    print("📊 测试结果总结:")
    
    for result in results:
        status = "✅" if result['overall'] else "❌"
        print(f"  {status} {result['type']}:")
        print(f"    - 多级列头检测: {result['is_multi']}")
        print(f"    - 智能读取: {'✅' if result['smart_read'] else '❌'}")
        print(f"    - 代码执行: {'✅' if result['run_code'] else '❌'}")
    
    success_count = sum(1 for r in results if r['overall'])
    total_count = len(results)
    
    print(f"\n🎯 总体结果: {success_count}/{total_count} 通过")
    
    if success_count == total_count:
        print("🎉 所有测试通过！多级列头处理能力正常")
    else:
        print("⚠️ 部分测试失败，多级列头处理仍有问题")
        print("\n🔍 分析:")
        print("  - 简单列头应该能正确处理")
        print("  - 多级列头检测可能需要优化")
        print("  - 复杂格式文件需要更智能的参数选择")

if __name__ == "__main__":
    main()