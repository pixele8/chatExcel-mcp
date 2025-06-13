#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
简单测试脚本
"""

import pandas as pd
import openpyxl
from pathlib import Path
import sys
import os

# 添加当前目录到Python路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from enhanced_excel_helper import smart_read_excel
from server import run_excel_code

def create_simple_test_excel():
    """创建一个简单的测试Excel文件"""
    test_file = "simple_test.xlsx"
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试数据"
    
    # 添加简单的列头和数据
    ws['A1'] = '姓名'
    ws['B1'] = '部门'
    ws['C1'] = '工资'
    
    # 添加数据
    data = [
        ['张三', '技术部', 8000],
        ['李四', '销售部', 6000],
        ['王五', '技术部', 9000]
    ]
    
    for i, row in enumerate(data, start=2):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    
    wb.save(test_file)
    print(f"创建测试文件: {test_file}")
    return test_file

def test_smart_read(file_path):
    """测试智能读取"""
    print("\n=== 测试智能读取 ===")
    
    try:
        result = smart_read_excel(file_path, auto_detect=True)
        if result['success']:
            df = result['dataframe']
            print(f"读取成功，数据形状: {df.shape}")
            print(f"列名: {df.columns.tolist()}")
            print(f"列名类型: {[type(col) for col in df.columns]}")
            print("\n数据内容:")
            print(df)
            
            # 测试分组操作
            print("\n=== 测试分组操作 ===")
            try:
                grouped = df.groupby('部门').size()
                print("分组成功!")
                print(grouped)
                return True
            except Exception as e:
                print(f"分组失败: {e}")
                return False
        else:
            print(f"读取失败: {result['errors']}")
            return False
    except Exception as e:
        print(f"智能读取异常: {e}")
        return False

def test_run_excel_code(file_path):
    """测试run_excel_code函数"""
    print("\n=== 测试run_excel_code函数 ===")
    
    code = """
# 检查DataFrame信息
print(f"DataFrame形状: {df.shape}")
print(f"列名: {df.columns.tolist()}")
print(f"列名类型: {[type(col) for col in df.columns]}")
print("\n数据内容:")
print(df)

# 测试分组操作
try:
    result = df.groupby('部门').size()
    print("\n分组结果:")
    print(result)
except Exception as e:
    print(f"分组失败: {e}")
"""
    
    try:
        response = run_excel_code(file_path, code, auto_detect=True)
        print(f"执行成功: {response.get('success', False)}")
        if 'result' in response:
            print("执行结果:")
            print(response['result'])
        if 'error' in response:
            print(f"执行错误: {response['error']}")
        return response.get('success', False)
    except Exception as e:
        print(f"run_excel_code执行失败: {e}")
        return False

def main():
    """主函数"""
    print("开始简单测试...")
    
    # 创建测试文件
    test_file = create_simple_test_excel()
    
    try:
        # 测试智能读取
        smart_read_success = test_smart_read(test_file)
        
        # 测试run_excel_code
        run_code_success = test_run_excel_code(test_file)
        
        print(f"\n=== 测试结果 ===")
        print(f"智能读取测试: {'通过' if smart_read_success else '失败'}")
        print(f"run_excel_code测试: {'通过' if run_code_success else '失败'}")
        
    finally:
        # 清理测试文件
        if Path(test_file).exists():
            Path(test_file).unlink()
            print(f"\n清理测试文件: {test_file}")

if __name__ == "__main__":
    main()