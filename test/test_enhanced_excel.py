#!/usr/bin/env python3
"""测试增强的Excel处理功能"""

import pandas as pd
import openpyxl
import os
from excel_helper import _suggest_excel_read_parameters, detect_excel_structure

def create_test_excel():
    """创建测试用的复杂Excel文件"""
    # 创建一个包含多级列头和复杂格式的Excel文件
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试数据"
    
    # 添加一些空行
    ws['A1'] = None
    ws['A2'] = None
    
    # 添加标题行
    ws['A3'] = "销售数据报表"
    ws.merge_cells('A3:E3')
    
    # 添加多级列头
    # 第一级列头
    ws['A4'] = "基本信息"
    ws.merge_cells('A4:B4')
    ws['C4'] = "销售数据"
    ws.merge_cells('C4:E4')
    
    # 第二级列头
    ws['A5'] = "产品名称"
    ws['B5'] = "产品类别"
    ws['C5'] = "Q1销售额"
    ws['D5'] = "Q2销售额"
    ws['E5'] = "总销售额"
    
    # 添加数据
    data = [
        ["笔记本电脑", "电子产品", 120000, 135000, 255000],
        ["智能手机", "电子产品", 200000, 180000, 380000],
        ["运动鞋", "服装", 80000, 95000, 175000],
        ["咖啡机", "家电", 45000, 52000, 97000],
        ["书籍", "文具", 25000, 30000, 55000]
    ]
    
    for i, row in enumerate(data, start=6):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    
    # 保存文件
    test_file = "test_complex_excel.xlsx"
    wb.save(test_file)
    wb.close()
    
    return os.path.abspath(test_file)

def test_parameter_suggestion():
    """测试参数推荐功能"""
    print("=== 测试Excel参数推荐功能 ===")
    
    # 创建测试文件
    test_file = create_test_excel()
    print(f"创建测试文件: {test_file}")
    
    try:
        # 测试参数推荐
        suggestions = _suggest_excel_read_parameters(test_file)
        print("\n推荐参数:")
        print(f"  推荐的读取参数: {suggestions['recommended_params']}")
        print(f"  分析结果: {suggestions['analysis']}")
        print(f"  警告信息: {suggestions['warnings']}")
        print(f"  提示信息: {suggestions['tips']}")
        
        # 测试结构检测
        print("\n=== 测试Excel结构检测 ===")
        structure = detect_excel_structure(test_file)
        print(f"工作表信息: {structure['sheets']}")
        print(f"合并单元格: {structure['merged_cells']}")
        
        # 测试使用推荐参数读取
        print("\n=== 测试使用推荐参数读取 ===")
        recommended_params = suggestions['recommended_params']
        
        # 基础读取
        print("\n1. 基础读取（无参数）:")
        try:
            df_basic = pd.read_excel(test_file)
            print(f"   形状: {df_basic.shape}")
            print(f"   列名: {df_basic.columns.tolist()}")
            print(f"   前3行:\n{df_basic.head(3)}")
        except Exception as e:
            print(f"   错误: {e}")
        
        # 使用推荐参数读取
        print("\n2. 使用推荐参数读取:")
        try:
            df_recommended = pd.read_excel(test_file, **recommended_params)
            print(f"   形状: {df_recommended.shape}")
            print(f"   列名: {df_recommended.columns.tolist()}")
            print(f"   前3行:\n{df_recommended.head(3)}")
        except Exception as e:
            print(f"   错误: {e}")
        
        # 测试多级列头处理
        print("\n3. 多级列头处理:")
        try:
            df_multiheader = pd.read_excel(test_file, skiprows=3, header=[0, 1])
            print(f"   形状: {df_multiheader.shape}")
            print(f"   列名: {df_multiheader.columns.tolist()}")
            print(f"   前3行:\n{df_multiheader.head(3)}")
        except Exception as e:
            print(f"   错误: {e}")
        
        # 测试选择特定列
        print("\n4. 选择特定列:")
        try:
            df_selected = pd.read_excel(test_file, skiprows=4, usecols=[0, 1, 4])
            print(f"   形状: {df_selected.shape}")
            print(f"   列名: {df_selected.columns.tolist()}")
            print(f"   前3行:\n{df_selected.head(3)}")
        except Exception as e:
            print(f"   错误: {e}")
        
    finally:
        # 清理测试文件
        if os.path.exists(test_file):
            os.remove(test_file)
            print(f"\n清理测试文件: {test_file}")

def test_enhanced_functions():
    """测试增强的函数功能"""
    print("\n=== 测试增强的函数功能 ===")
    
    # 这里可以添加对server.py中增强函数的测试
    # 由于需要MCP环境，这里主要测试核心逻辑
    
    test_file = create_test_excel()
    
    try:
        # 模拟read_excel_metadata的增强功能
        print("\n测试增强的元数据读取功能:")
        
        # 测试不同参数组合
        test_cases = [
            {"skiprows": None, "header": None, "usecols": None},
            {"skiprows": 4, "header": 0, "usecols": None},
            {"skiprows": 3, "header": [0, 1], "usecols": None},
            {"skiprows": 4, "header": 0, "usecols": [0, 1, 4]}
        ]
        
        for i, params in enumerate(test_cases, 1):
            print(f"\n  测试案例 {i}: {params}")
            try:
                read_params = {k: v for k, v in params.items() if v is not None}
                df = pd.read_excel(test_file, **read_params)
                print(f"    成功读取，形状: {df.shape}")
                print(f"    列名: {df.columns.tolist()[:3]}...")
            except Exception as e:
                print(f"    读取失败: {e}")
    
    finally:
        if os.path.exists(test_file):
            os.remove(test_file)

if __name__ == "__main__":
    print("开始测试增强的Excel处理功能...")
    
    # 检查依赖
    print(f"pandas版本: {pd.__version__}")
    print(f"openpyxl版本: {openpyxl.__version__}")
    
    try:
        import xlrd
        print(f"xlrd版本: {xlrd.__version__}")
    except ImportError:
        print("xlrd未安装")
    
    try:
        import xlsxwriter
        print(f"xlsxwriter版本: {xlsxwriter.__version__}")
    except ImportError:
        print("xlsxwriter未安装")
    
    print("\n" + "="*50)
    
    # 运行测试
    test_parameter_suggestion()
    test_enhanced_functions()
    
    print("\n测试完成！")