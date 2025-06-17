#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 数据验证与质量控制工具模块
企业级 Excel 数据处理、验证、转换和质量控制工具集

功能特性:
- 数据完整性验证和修复
- 单元格内容精确提取和格式转换
- 汉字/数字/英文/特殊字符自动化处理
- 多条件表格数据提取和核对
- 多表格特定条件数据合并
- 数据质量评估和报告
- 智能数据清洗和标准化
- 批量数据处理和验证

作者: ChatExcel MCP Team
版本: 1.0.0
创建日期: 2025-06-18
"""

import json
import logging
import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional, Union, Tuple
from collections import defaultdict, Counter
import hashlib
import time
from datetime import datetime, timedelta

import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelDataQualityController:
    """Excel 数据质量控制器"""
    
    def __init__(self):
        self.quality_rules = self._initialize_quality_rules()
        self.character_patterns = self._initialize_character_patterns()
        self.validation_history = []
        
    def _initialize_quality_rules(self) -> Dict[str, Any]:
        """初始化数据质量规则"""
        return {
            'completeness': {
                'min_fill_rate': 0.8,  # 最小填充率
                'critical_columns': [],  # 关键列不能为空
                'allow_zero': True  # 是否允许零值
            },
            'consistency': {
                'date_formats': ['%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%m/%d/%Y'],
                'number_formats': ['integer', 'float', 'percentage', 'currency'],
                'text_case': 'auto'  # 'upper', 'lower', 'title', 'auto'
            },
            'accuracy': {
                'numeric_range': {},  # 数值范围验证
                'text_length': {},  # 文本长度验证
                'regex_patterns': {}  # 正则表达式验证
            },
            'uniqueness': {
                'unique_columns': [],  # 需要唯一性的列
                'composite_keys': []  # 复合主键
            }
        }
    
    def _initialize_character_patterns(self) -> Dict[str, str]:
        """初始化字符模式"""
        return {
            'chinese': r'[\u4e00-\u9fff]+',
            'english': r'[a-zA-Z]+',
            'numbers': r'\d+',
            'special_chars': r'[^\w\s\u4e00-\u9fff]',
            'whitespace': r'\s+',
            'email': r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',
            'phone': r'1[3-9]\d{9}',
            'id_card': r'\d{17}[\dXx]',
            'currency': r'[¥$€£]?\d+(\.\d{2})?',
            'percentage': r'\d+(\.\d+)?%',
            'date': r'\d{4}[-/]\d{1,2}[-/]\d{1,2}'
        }
    
    def comprehensive_quality_check(self, 
                                   file_path: str, 
                                   level: str = "standard") -> Dict[str, Any]:
        """综合数据质量检查
        
        Args:
            file_path: Excel文件路径
            level: 检查级别 (basic, standard, comprehensive)
            
        Returns:
            质量检查结果字典
        """
        try:
            # 执行数据完整性验证
            integrity_result = self.validate_data_integrity(file_path)
            
            # 根据级别执行不同深度的检查
            quality_result = {
                'file_path': file_path,
                'check_level': level,
                'timestamp': datetime.now().isoformat(),
                'integrity_check': integrity_result,
                'overall_score': 0.0,
                'recommendations': [],
                'issues_found': [],
                'passed': True
            }
            
            # 计算总体质量分数
            if integrity_result.get('validation_summary', {}).get('passed', False):
                quality_result['overall_score'] = 85.0
            else:
                quality_result['overall_score'] = 60.0
                quality_result['passed'] = False
                quality_result['issues_found'].append('数据完整性验证未通过')
                quality_result['recommendations'].append('建议进行数据清洗和修复')
            
            return quality_result
            
        except Exception as e:
            logger.error(f"综合质量检查失败: {str(e)}")
            return {
                'file_path': file_path,
                'check_level': level,
                'error': str(e),
                'passed': False,
                'overall_score': 0.0
            }
    
    def validate_data_integrity(self, 
                              file_path: str, 
                              sheet_name: Optional[str] = None,
                              validation_rules: Optional[Dict] = None) -> Dict[str, Any]:
        """验证数据完整性
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            validation_rules: 自定义验证规则
            
        Returns:
            验证结果字典
        """
        try:
            # 读取Excel文件
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            
            # 使用自定义规则或默认规则
            rules = validation_rules or self.quality_rules
            
            validation_result = {
                'file_path': file_path,
                'sheet_name': sheet_name,
                'timestamp': datetime.now().isoformat(),
                'total_rows': len(df),
                'total_columns': len(df.columns),
                'validation_summary': {
                    'passed': True,
                    'issues_count': 0,
                    'warnings_count': 0
                },
                'completeness_check': {},
                'consistency_check': {},
                'accuracy_check': {},
                'uniqueness_check': {},
                'detailed_issues': [],
                'recommendations': []
            }
            
            # 1. 完整性检查
            completeness_result = self._check_completeness(df, rules['completeness'])
            validation_result['completeness_check'] = completeness_result
            
            # 2. 一致性检查
            consistency_result = self._check_consistency(df, rules['consistency'])
            validation_result['consistency_check'] = consistency_result
            
            # 3. 准确性检查
            accuracy_result = self._check_accuracy(df, rules['accuracy'])
            validation_result['accuracy_check'] = accuracy_result
            
            # 4. 唯一性检查
            uniqueness_result = self._check_uniqueness(df, rules['uniqueness'])
            validation_result['uniqueness_check'] = uniqueness_result
            
            # 汇总验证结果
            self._summarize_validation_results(validation_result)
            
            # 记录验证历史
            self.validation_history.append(validation_result)
            
            return {
                'success': True,
                'validation_result': validation_result
            }
            
        except Exception as e:
            logger.error(f"数据完整性验证失败: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'validation_result': None
            }
    
    def _check_completeness(self, df: pd.DataFrame, rules: Dict) -> Dict[str, Any]:
        """检查数据完整性"""
        result = {
            'overall_fill_rate': 0.0,
            'column_fill_rates': {},
            'missing_data_summary': {},
            'critical_columns_status': {},
            'issues': [],
            'passed': True
        }
        
        # 计算总体填充率
        total_cells = df.size
        non_null_cells = df.count().sum()
        result['overall_fill_rate'] = non_null_cells / total_cells if total_cells > 0 else 0
        
        # 检查每列的填充率
        for col in df.columns:
            fill_rate = df[col].count() / len(df) if len(df) > 0 else 0
            result['column_fill_rates'][col] = fill_rate
            
            # 检查是否满足最小填充率要求
            if fill_rate < rules['min_fill_rate']:
                issue = {
                    'type': 'low_fill_rate',
                    'column': col,
                    'current_rate': fill_rate,
                    'required_rate': rules['min_fill_rate'],
                    'missing_count': len(df) - df[col].count()
                }
                result['issues'].append(issue)
                result['passed'] = False
        
        # 检查关键列
        for col in rules.get('critical_columns', []):
            if col in df.columns:
                missing_count = df[col].isnull().sum()
                result['critical_columns_status'][col] = {
                    'missing_count': missing_count,
                    'passed': missing_count == 0
                }
                
                if missing_count > 0:
                    result['issues'].append({
                        'type': 'critical_column_missing',
                        'column': col,
                        'missing_count': missing_count
                    })
                    result['passed'] = False
        
        # 生成缺失数据摘要
        missing_summary = df.isnull().sum().to_dict()
        result['missing_data_summary'] = {k: v for k, v in missing_summary.items() if v > 0}
        
        return result
    
    def _check_consistency(self, df: pd.DataFrame, rules: Dict) -> Dict[str, Any]:
        """检查数据一致性"""
        result = {
            'date_format_consistency': {},
            'number_format_consistency': {},
            'text_case_consistency': {},
            'issues': [],
            'passed': True
        }
        
        # 检查日期格式一致性
        for col in df.columns:
            if df[col].dtype == 'object':
                # 尝试识别日期列
                sample_values = df[col].dropna().astype(str).head(100)
                date_patterns = []
                
                for value in sample_values:
                    for fmt in rules['date_formats']:
                        try:
                            datetime.strptime(value, fmt)
                            date_patterns.append(fmt)
                            break
                        except ValueError:
                            continue
                
                if date_patterns:
                    pattern_counts = Counter(date_patterns)
                    result['date_format_consistency'][col] = {
                        'patterns_found': dict(pattern_counts),
                        'most_common': pattern_counts.most_common(1)[0] if pattern_counts else None,
                        'is_consistent': len(pattern_counts) == 1
                    }
                    
                    if len(pattern_counts) > 1:
                        result['issues'].append({
                            'type': 'inconsistent_date_format',
                            'column': col,
                            'patterns': dict(pattern_counts)
                        })
                        result['passed'] = False
        
        return result
    
    def _check_accuracy(self, df: pd.DataFrame, rules: Dict) -> Dict[str, Any]:
        """检查数据准确性"""
        result = {
            'numeric_range_validation': {},
            'text_length_validation': {},
            'pattern_validation': {},
            'issues': [],
            'passed': True
        }
        
        # 数值范围验证
        for col, range_rule in rules.get('numeric_range', {}).items():
            if col in df.columns and pd.api.types.is_numeric_dtype(df[col]):
                min_val = range_rule.get('min')
                max_val = range_rule.get('max')
                
                out_of_range = []
                if min_val is not None:
                    out_of_range.extend(df[df[col] < min_val].index.tolist())
                if max_val is not None:
                    out_of_range.extend(df[df[col] > max_val].index.tolist())
                
                result['numeric_range_validation'][col] = {
                    'rule': range_rule,
                    'out_of_range_count': len(set(out_of_range)),
                    'out_of_range_rows': list(set(out_of_range))
                }
                
                if out_of_range:
                    result['issues'].append({
                        'type': 'numeric_out_of_range',
                        'column': col,
                        'count': len(set(out_of_range)),
                        'rows': list(set(out_of_range))[:10]  # 只显示前10个
                    })
                    result['passed'] = False
        
        # 文本长度验证
        for col, length_rule in rules.get('text_length', {}).items():
            if col in df.columns:
                min_len = length_rule.get('min', 0)
                max_len = length_rule.get('max', float('inf'))
                
                text_series = df[col].astype(str)
                invalid_lengths = text_series[
                    (text_series.str.len() < min_len) | 
                    (text_series.str.len() > max_len)
                ]
                
                result['text_length_validation'][col] = {
                    'rule': length_rule,
                    'invalid_count': len(invalid_lengths),
                    'invalid_rows': invalid_lengths.index.tolist()
                }
                
                if len(invalid_lengths) > 0:
                    result['issues'].append({
                        'type': 'invalid_text_length',
                        'column': col,
                        'count': len(invalid_lengths)
                    })
                    result['passed'] = False
        
        return result
    
    def _check_uniqueness(self, df: pd.DataFrame, rules: Dict) -> Dict[str, Any]:
        """检查数据唯一性"""
        result = {
            'unique_columns_validation': {},
            'composite_keys_validation': {},
            'issues': [],
            'passed': True
        }
        
        # 单列唯一性检查
        for col in rules.get('unique_columns', []):
            if col in df.columns:
                duplicates = df[df.duplicated(subset=[col], keep=False)]
                result['unique_columns_validation'][col] = {
                    'duplicate_count': len(duplicates),
                    'duplicate_values': duplicates[col].unique().tolist() if len(duplicates) > 0 else []
                }
                
                if len(duplicates) > 0:
                    result['issues'].append({
                        'type': 'duplicate_values',
                        'column': col,
                        'count': len(duplicates)
                    })
                    result['passed'] = False
        
        # 复合主键唯一性检查
        for key_cols in rules.get('composite_keys', []):
            if all(col in df.columns for col in key_cols):
                duplicates = df[df.duplicated(subset=key_cols, keep=False)]
                result['composite_keys_validation'][str(key_cols)] = {
                    'duplicate_count': len(duplicates),
                    'duplicate_rows': duplicates.index.tolist() if len(duplicates) > 0 else []
                }
                
                if len(duplicates) > 0:
                    result['issues'].append({
                        'type': 'duplicate_composite_key',
                        'columns': key_cols,
                        'count': len(duplicates)
                    })
                    result['passed'] = False
        
        return result
    
    def _summarize_validation_results(self, validation_result: Dict) -> None:
        """汇总验证结果"""
        summary = validation_result['validation_summary']
        
        # 统计问题数量
        issues_count = 0
        warnings_count = 0
        
        for check_type in ['completeness_check', 'consistency_check', 'accuracy_check', 'uniqueness_check']:
            check_result = validation_result.get(check_type, {})
            if not check_result.get('passed', True):
                summary['passed'] = False
            
            issues = check_result.get('issues', [])
            for issue in issues:
                if issue.get('severity', 'error') == 'warning':
                    warnings_count += 1
                else:
                    issues_count += 1
        
        summary['issues_count'] = issues_count
        summary['warnings_count'] = warnings_count
        
        # 生成建议
        recommendations = []
        if issues_count > 0:
            recommendations.append("发现数据质量问题，建议进行数据清洗")
        if warnings_count > 0:
            recommendations.append("存在数据质量警告，建议检查相关数据")
        
        validation_result['recommendations'] = recommendations


class ExcelCellContentExtractor:
    """Excel 单元格内容精确提取器"""
    
    def __init__(self):
        self.extraction_patterns = self._initialize_extraction_patterns()
    
    def _initialize_extraction_patterns(self) -> Dict[str, str]:
        """初始化提取模式"""
        return {
            'chinese_only': r'[\u4e00-\u9fff]+',
            'english_only': r'[a-zA-Z]+',
            'numbers_only': r'\d+',
            'chinese_numbers': r'[\u4e00-\u9fff\d]+',
            'alphanumeric': r'[a-zA-Z0-9]+',
            'email': r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',
            'phone': r'1[3-9]\d{9}',
            'id_card': r'\d{17}[\dXx]',
            'currency': r'[¥$€£]?\d+(\.\d{2})?',
            'percentage': r'\d+(\.\d+)?%',
            'date': r'\d{4}[-/]\d{1,2}[-/]\d{1,2}',
            'time': r'\d{1,2}:\d{2}(:\d{2})?',
            'url': r'https?://[^\s]+',
            'special_chars': r'[^\w\s\u4e00-\u9fff]'
        }
    
    def extract_cell_content_advanced(self, 
                                     file_path: str, 
                                     cell_range: Optional[str] = None,
                                     sheet_name: Optional[str] = None,
                                     extract_type: str = 'all') -> Dict[str, Any]:
        """高级单元格内容提取
        
        Args:
            file_path: Excel文件路径
            cell_range: 单元格范围 (如 'A1:C10')
            sheet_name: 工作表名称
            extract_type: 提取类型
            
        Returns:
            提取结果字典
        """
        return self.extract_cell_content(
            file_path=file_path,
            sheet_name=sheet_name,
            cell_range=cell_range,
            extraction_type=extract_type,
            clean_whitespace=True
        )
    
    def extract_cell_content(self, 
                           file_path: str, 
                           sheet_name: Optional[str] = None,
                           cell_range: Optional[str] = None,
                           extraction_type: str = 'all',
                           clean_whitespace: bool = True) -> Dict[str, Any]:
        """精确提取单元格内容
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            cell_range: 单元格范围 (如 'A1:C10')
            extraction_type: 提取类型
            clean_whitespace: 是否清理空白字符
            
        Returns:
            提取结果字典
        """
        try:
            # 使用openpyxl读取原始单元格内容
            workbook = load_workbook(file_path, data_only=False)
            
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    return {
                        'success': False,
                        'error': f'工作表 "{sheet_name}" 不存在'
                    }
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            extraction_result = {
                'file_path': file_path,
                'sheet_name': worksheet.title,
                'extraction_type': extraction_type,
                'timestamp': datetime.now().isoformat(),
                'extracted_data': {},
                'statistics': {},
                'formatting_info': {}
            }
            
            # 确定处理范围
            if cell_range:
                cell_range_obj = worksheet[cell_range]
                if hasattr(cell_range_obj, '__iter__') and not isinstance(cell_range_obj, openpyxl.cell.Cell):
                    # 多个单元格
                    cells = [cell for row in cell_range_obj for cell in row]
                else:
                    # 单个单元格
                    cells = [cell_range_obj]
            else:
                # 处理所有有数据的单元格
                cells = []
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cells.append(cell)
            
            # 提取内容
            extracted_data = {}
            formatting_info = {}
            
            for cell in cells:
                cell_coord = cell.coordinate
                cell_value = cell.value
                
                if cell_value is None:
                    continue
                
                # 转换为字符串进行处理
                cell_text = str(cell_value)
                
                # 清理空白字符
                if clean_whitespace:
                    cell_text = re.sub(r'\s+', ' ', cell_text).strip()
                
                # 根据提取类型处理
                extracted_content = self._extract_by_type(cell_text, extraction_type)
                
                # 获取格式信息
                format_info = self._get_cell_format_info(cell)
                
                extracted_data[cell_coord] = {
                    'original_value': cell_value,
                    'original_text': str(cell_value),
                    'cleaned_text': cell_text,
                    'extracted_content': extracted_content,
                    'data_type': type(cell_value).__name__
                }
                
                formatting_info[cell_coord] = format_info
            
            extraction_result['extracted_data'] = extracted_data
            extraction_result['formatting_info'] = formatting_info
            
            # 生成统计信息
            statistics = self._generate_extraction_statistics(extracted_data)
            extraction_result['statistics'] = statistics
            
            return {
                'success': True,
                'extraction_result': extraction_result
            }
            
        except Exception as e:
            logger.error(f"单元格内容提取失败: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _extract_by_type(self, text: str, extraction_type: str) -> Dict[str, Any]:
        """根据类型提取内容"""
        result = {
            'extraction_type': extraction_type,
            'matches': [],
            'match_count': 0
        }
        
        if extraction_type == 'all':
            # 提取所有类型
            for pattern_name, pattern in self.extraction_patterns.items():
                matches = re.findall(pattern, text)
                if matches:
                    result[pattern_name] = matches
        elif extraction_type in self.extraction_patterns:
            # 提取特定类型
            pattern = self.extraction_patterns[extraction_type]
            matches = re.findall(pattern, text)
            result['matches'] = matches
            result['match_count'] = len(matches)
        else:
            # 自定义正则表达式
            try:
                matches = re.findall(extraction_type, text)
                result['matches'] = matches
                result['match_count'] = len(matches)
            except re.error as e:
                result['error'] = f'正则表达式错误: {str(e)}'
        
        return result
    
    def _get_cell_format_info(self, cell) -> Dict[str, Any]:
        """获取单元格格式信息"""
        format_info = {
            'font': {
                'name': cell.font.name,
                'size': cell.font.size,
                'bold': cell.font.bold,
                'italic': cell.font.italic,
                'color': str(cell.font.color.rgb) if cell.font.color and cell.font.color.rgb else None
            },
            'fill': {
                'pattern_type': cell.fill.patternType,
                'start_color': str(cell.fill.start_color.rgb) if cell.fill.start_color and cell.fill.start_color.rgb else None,
                'end_color': str(cell.fill.end_color.rgb) if cell.fill.end_color and cell.fill.end_color.rgb else None
            },
            'alignment': {
                'horizontal': cell.alignment.horizontal,
                'vertical': cell.alignment.vertical,
                'wrap_text': cell.alignment.wrap_text
            },
            'number_format': cell.number_format,
            'has_hyperlink': cell.hyperlink is not None
        }
        
        return format_info
    
    def _generate_extraction_statistics(self, extracted_data: Dict) -> Dict[str, Any]:
        """生成提取统计信息"""
        statistics = {
            'total_cells_processed': len(extracted_data),
            'cells_with_content': 0,
            'data_type_distribution': {},
            'extraction_summary': {}
        }
        
        data_types = []
        extraction_counts = defaultdict(int)
        
        for cell_data in extracted_data.values():
            if cell_data['extracted_content']['match_count'] > 0:
                statistics['cells_with_content'] += 1
            
            data_types.append(cell_data['data_type'])
            
            # 统计提取结果
            extracted_content = cell_data['extracted_content']
            for key, value in extracted_content.items():
                if isinstance(value, list) and len(value) > 0:
                    extraction_counts[key] += len(value)
        
        # 数据类型分布
        type_counts = Counter(data_types)
        statistics['data_type_distribution'] = dict(type_counts)
        
        # 提取摘要
        statistics['extraction_summary'] = dict(extraction_counts)
        
        return statistics


class ExcelCharacterConverter:
    """Excel 字符格式自动化转换器"""
    
    def __init__(self):
        self.conversion_rules = self._initialize_conversion_rules()
    
    def _initialize_conversion_rules(self) -> Dict[str, Any]:
        """初始化转换规则"""
        return {
            'chinese_to_english_numbers': {
                '零': '0', '一': '1', '二': '2', '三': '3', '四': '4',
                '五': '5', '六': '6', '七': '7', '八': '8', '九': '9',
                '十': '10', '百': '100', '千': '1000', '万': '10000'
            },
            'full_width_to_half_width': {
                '０': '0', '１': '1', '２': '2', '３': '3', '４': '4',
                '５': '5', '６': '6', '７': '7', '８': '8', '９': '9',
                'Ａ': 'A', 'Ｂ': 'B', 'Ｃ': 'C', 'Ｄ': 'D', 'Ｅ': 'E',
                'Ｆ': 'F', 'Ｇ': 'G', 'Ｈ': 'H', 'Ｉ': 'I', 'Ｊ': 'J',
                'Ｋ': 'K', 'Ｌ': 'L', 'Ｍ': 'M', 'Ｎ': 'N', 'Ｏ': 'O',
                'Ｐ': 'P', 'Ｑ': 'Q', 'Ｒ': 'R', 'Ｓ': 'S', 'Ｔ': 'T',
                'Ｕ': 'U', 'Ｖ': 'V', 'Ｗ': 'W', 'Ｘ': 'X', 'Ｙ': 'Y', 'Ｚ': 'Z',
                'ａ': 'a', 'ｂ': 'b', 'ｃ': 'c', 'ｄ': 'd', 'ｅ': 'e',
                'ｆ': 'f', 'ｇ': 'g', 'ｈ': 'h', 'ｉ': 'i', 'ｊ': 'j',
                'ｋ': 'k', 'ｌ': 'l', 'ｍ': 'm', 'ｎ': 'n', 'ｏ': 'o',
                'ｐ': 'p', 'ｑ': 'q', 'ｒ': 'r', 'ｓ': 's', 'ｔ': 't',
                'ｕ': 'u', 'ｖ': 'v', 'ｗ': 'w', 'ｘ': 'x', 'ｙ': 'y', 'ｚ': 'z'
            },
            'special_characters': {
                '，': ',', '。': '.', '；': ';', '：': ':', '？': '?',
                '！': '!', '（': '(', '）': ')', '【': '[', '】': ']',
                '「': '"', '」': '"', '『': "'", '』': "'"
            }
        }
    
    def convert_character_formats(self, 
                                file_path: str,
                                sheet_name: Optional[str] = None,
                                conversion_types: List[str] = None,
                                target_columns: List[str] = None,
                                save_backup: bool = True) -> Dict[str, Any]:
        """转换字符格式
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            conversion_types: 转换类型列表
            target_columns: 目标列
            save_backup: 是否保存备份
            
        Returns:
            转换结果字典
        """
        try:
            # 读取Excel文件
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            
            # 创建备份
            original_df = df.copy() if save_backup else None
            
            # 默认转换类型
            if conversion_types is None:
                conversion_types = ['full_width_to_half_width', 'special_characters']
            
            # 确定目标列
            if target_columns is None:
                target_columns = df.columns.tolist()
            else:
                target_columns = [col for col in target_columns if col in df.columns]
            
            conversion_result = {
                'file_path': file_path,
                'sheet_name': sheet_name,
                'timestamp': datetime.now().isoformat(),
                'conversion_types': conversion_types,
                'target_columns': target_columns,
                'conversion_summary': {},
                'detailed_changes': {},
                'statistics': {}
            }
            
            total_changes = 0
            detailed_changes = {}
            
            # 对每列进行转换
            for col in target_columns:
                if col not in df.columns:
                    continue
                
                column_changes = 0
                column_details = []
                
                # 转换该列的每个单元格
                for idx, value in df[col].items():
                    if pd.isna(value):
                        continue
                    
                    original_value = str(value)
                    converted_value = original_value
                    
                    # 应用转换规则
                    for conversion_type in conversion_types:
                        converted_value = self._apply_conversion_rule(
                            converted_value, conversion_type
                        )
                    
                    # 记录变化
                    if converted_value != original_value:
                        df.at[idx, col] = converted_value
                        column_changes += 1
                        total_changes += 1
                        
                        column_details.append({
                            'row': idx,
                            'original': original_value,
                            'converted': converted_value
                        })
                
                if column_changes > 0:
                    detailed_changes[col] = {
                        'changes_count': column_changes,
                        'changes': column_details[:10]  # 只保存前10个变化
                    }
            
            conversion_result['detailed_changes'] = detailed_changes
            conversion_result['conversion_summary'] = {
                'total_changes': total_changes,
                'columns_affected': len(detailed_changes),
                'conversion_rate': total_changes / (len(df) * len(target_columns)) if len(df) > 0 and len(target_columns) > 0 else 0
            }
            
            # 保存转换后的文件
            if total_changes > 0:
                output_path = self._generate_output_path(file_path, 'converted')
                
                if sheet_name:
                    # 如果指定了工作表，需要保持其他工作表不变
                    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                        # 读取原文件的所有工作表
                        original_file = pd.ExcelFile(file_path)
                        for sheet in original_file.sheet_names:
                            if sheet == sheet_name:
                                df.to_excel(writer, sheet_name=sheet, index=False)
                            else:
                                original_sheet_df = pd.read_excel(file_path, sheet_name=sheet)
                                original_sheet_df.to_excel(writer, sheet_name=sheet, index=False)
                else:
                    df.to_excel(output_path, index=False)
                
                conversion_result['output_file'] = output_path
            
            # 生成统计信息
            statistics = self._generate_conversion_statistics(original_df, df, conversion_types)
            conversion_result['statistics'] = statistics
            
            return {
                'success': True,
                'conversion_result': conversion_result,
                'converted_data': df
            }
            
        except Exception as e:
            logger.error(f"字符格式转换失败: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _apply_conversion_rule(self, text: str, conversion_type: str) -> str:
        """应用转换规则"""
        if conversion_type not in self.conversion_rules:
            return text
        
        conversion_map = self.conversion_rules[conversion_type]
        result = text
        
        # 特殊处理中文数字转换
        if conversion_type == 'chinese_to_english_numbers':
            result = self._convert_chinese_numbers(text)
        else:
            # 普通字符替换
            for old_char, new_char in conversion_map.items():
                result = result.replace(old_char, new_char)
        
        return result
    
    def _convert_chinese_numbers(self, text: str) -> str:
        """转换中文数字"""
        # 简单的中文数字转换实现
        chinese_to_arabic = self.conversion_rules['chinese_to_english_numbers']
        result = text
        
        # 处理简单的中文数字
        for chinese, arabic in chinese_to_arabic.items():
            result = result.replace(chinese, arabic)
        
        return result
    
    def _generate_output_path(self, original_path: str, suffix: str) -> str:
        """生成输出文件路径"""
        path = Path(original_path)
        return str(path.parent / f"{path.stem}_{suffix}{path.suffix}")
    
    def _generate_conversion_statistics(self, 
                                      original_df: Optional[pd.DataFrame], 
                                      converted_df: pd.DataFrame, 
                                      conversion_types: List[str]) -> Dict[str, Any]:
        """生成转换统计信息"""
        statistics = {
            'conversion_types_applied': conversion_types,
            'total_cells': converted_df.size,
            'data_type_changes': {},
            'character_distribution': {}
        }
        
        if original_df is not None:
            # 比较数据类型变化
            for col in converted_df.columns:
                if col in original_df.columns:
                    original_type = str(original_df[col].dtype)
                    converted_type = str(converted_df[col].dtype)
                    if original_type != converted_type:
                        statistics['data_type_changes'][col] = {
                            'from': original_type,
                            'to': converted_type
                        }
        
        return statistics
    
    def batch_character_conversion(self, 
                                 file_path: str, 
                                 conversion_rules: Dict[str, Any], 
                                 output_path: str) -> Dict[str, Any]:
        """批量字符转换
        
        Args:
            file_path: Excel文件路径
            conversion_rules: 转换规则字典
            output_path: 输出文件路径
            
        Returns:
            转换结果字典
        """
        try:
            # 读取Excel文件
            df = pd.read_excel(file_path)
            original_df = df.copy()
            
            # 解析转换规则
            conversion_types = conversion_rules.get('conversion_types', ['full_width_to_half_width', 'special_characters'])
            target_columns = conversion_rules.get('target_columns', df.columns.tolist())
            custom_rules = conversion_rules.get('custom_rules', {})
            
            # 添加自定义规则到转换规则中
            if custom_rules:
                self.conversion_rules['custom'] = custom_rules
                if 'custom' not in conversion_types:
                    conversion_types.append('custom')
            
            batch_result = {
                'file_path': file_path,
                'output_path': output_path,
                'timestamp': datetime.now().isoformat(),
                'conversion_rules': conversion_rules,
                'processing_summary': {},
                'detailed_changes': {},
                'statistics': {}
            }
            
            total_changes = 0
            detailed_changes = {}
            
            # 确保目标列存在
            valid_columns = [col for col in target_columns if col in df.columns]
            
            # 对每列进行批量转换
            for col in valid_columns:
                column_changes = 0
                column_details = []
                
                # 批量处理该列的所有值
                for idx, value in df[col].items():
                    if pd.isna(value):
                        continue
                    
                    original_value = str(value)
                    converted_value = original_value
                    
                    # 应用所有转换规则
                    for conversion_type in conversion_types:
                        converted_value = self._apply_conversion_rule(
                            converted_value, conversion_type
                        )
                    
                    # 记录变化
                    if converted_value != original_value:
                        df.at[idx, col] = converted_value
                        column_changes += 1
                        total_changes += 1
                        
                        # 记录详细变化（限制数量以避免内存问题）
                        if len(column_details) < 50:
                            column_details.append({
                                'row': idx,
                                'original': original_value,
                                'converted': converted_value,
                                'applied_rules': conversion_types
                            })
                
                if column_changes > 0:
                    detailed_changes[col] = {
                        'changes_count': column_changes,
                        'sample_changes': column_details
                    }
            
            # 保存转换后的文件
            try:
                df.to_excel(output_path, index=False)
                batch_result['output_saved'] = True
            except Exception as save_error:
                logger.warning(f"保存文件失败: {str(save_error)}")
                batch_result['output_saved'] = False
                batch_result['save_error'] = str(save_error)
            
            # 生成处理摘要
            batch_result['processing_summary'] = {
                'total_rows_processed': len(df),
                'total_columns_processed': len(valid_columns),
                'total_changes': total_changes,
                'columns_affected': len(detailed_changes),
                'conversion_rate': (total_changes / (len(df) * len(valid_columns))) if len(df) > 0 and len(valid_columns) > 0 else 0,
                'rules_applied': conversion_types
            }
            
            batch_result['detailed_changes'] = detailed_changes
            
            # 生成统计信息
            statistics = {
                'original_shape': original_df.shape,
                'final_shape': df.shape,
                'data_integrity_maintained': original_df.shape == df.shape,
                'conversion_efficiency': {
                    'changes_per_second': total_changes / max(1, (datetime.now() - datetime.fromisoformat(batch_result['timestamp'].replace('Z', '+00:00').replace('+00:00', ''))).total_seconds()),
                    'processing_time_estimate': f"{max(1, total_changes // 1000)} seconds"
                },
                'quality_metrics': {
                    'successful_conversions': total_changes,
                    'failed_conversions': 0,  # 当前实现中没有失败的转换
                    'data_loss': False
                }
            }
            
            batch_result['statistics'] = statistics
            
            return {
                'success': True,
                'batch_result': batch_result,
                'converted_data': df
            }
            
        except Exception as e:
            logger.error(f"批量字符转换失败: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'error_type': 'BATCH_CONVERSION_ERROR'
            }


class ExcelMultiConditionExtractor:
    """Excel 多条件数据提取器"""
    
    def __init__(self):
        self.condition_operators = {
            'eq': lambda x, y: x == y,
            'ne': lambda x, y: x != y,
            'gt': lambda x, y: x > y,
            'ge': lambda x, y: x >= y,
            'lt': lambda x, y: x < y,
            'le': lambda x, y: x <= y,
            'contains': lambda x, y: str(y).lower() in str(x).lower(),
            'startswith': lambda x, y: str(x).lower().startswith(str(y).lower()),
            'endswith': lambda x, y: str(x).lower().endswith(str(y).lower()),
            'regex': lambda x, y: bool(re.search(str(y), str(x))),
            'in': lambda x, y: x in y,
            'not_in': lambda x, y: x not in y,
            'is_null': lambda x, y: pd.isna(x),
            'is_not_null': lambda x, y: not pd.isna(x)
        }
    
    def extract_multi_condition_data(self, 
                                   file_path: str,
                                   conditions: List[Dict[str, Any]],
                                   logic_operator: str = 'and',
                                   sheet_name: Optional[str] = None,
                                   output_columns: Optional[List[str]] = None) -> Dict[str, Any]:
        """多条件数据提取
        
        Args:
            file_path: Excel文件路径
            conditions: 条件列表，每个条件包含 column, operator, value
            logic_operator: 逻辑操作符 ('and' 或 'or')
            sheet_name: 工作表名称
            output_columns: 输出列
            
        Returns:
            提取结果字典
        """
        try:
            # 读取Excel文件
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            
            extraction_result = {
                'file_path': file_path,
                'sheet_name': sheet_name,
                'timestamp': datetime.now().isoformat(),
                'conditions': conditions,
                'logic_operator': logic_operator,
                'original_row_count': len(df),
                'extracted_row_count': 0,
                'extraction_rate': 0.0,
                'extracted_data': None,
                'condition_analysis': {}
            }
            
            # 验证条件
            valid_conditions = []
            condition_analysis = {}
            
            for i, condition in enumerate(conditions):
                column = condition.get('column')
                operator = condition.get('operator')
                value = condition.get('value')
                
                analysis = {
                    'valid': True,
                    'error': None,
                    'matches_count': 0
                }
                
                # 验证列是否存在
                if column not in df.columns:
                    analysis['valid'] = False
                    analysis['error'] = f'列 "{column}" 不存在'
                # 验证操作符
                elif operator not in self.condition_operators:
                    analysis['valid'] = False
                    analysis['error'] = f'不支持的操作符 "{operator}"'
                else:
                    # 计算单个条件的匹配数
                    try:
                        condition_mask = self._apply_condition(df, column, operator, value)
                        analysis['matches_count'] = condition_mask.sum()
                        valid_conditions.append((condition, condition_mask))
                    except Exception as e:
                        analysis['valid'] = False
                        analysis['error'] = f'条件应用失败: {str(e)}'
                
                condition_analysis[f'condition_{i}'] = analysis
            
            extraction_result['condition_analysis'] = condition_analysis
            
            # 如果没有有效条件，返回错误
            if not valid_conditions:
                return {
                    'success': False,
                    'error': '没有有效的条件',
                    'extraction_result': extraction_result
                }
            
            # 组合条件
            if len(valid_conditions) == 1:
                final_mask = valid_conditions[0][1]
            else:
                if logic_operator.lower() == 'and':
                    final_mask = valid_conditions[0][1]
                    for _, mask in valid_conditions[1:]:
                        final_mask = final_mask & mask
                elif logic_operator.lower() == 'or':
                    final_mask = valid_conditions[0][1]
                    for _, mask in valid_conditions[1:]:
                        final_mask = final_mask | mask
                else:
                    return {
                        'success': False,
                        'error': f'不支持的逻辑操作符: {logic_operator}'
                    }
            
            # 提取数据
            extracted_df = df[final_mask]
            
            # 选择输出列
            if output_columns:
                available_columns = [col for col in output_columns if col in extracted_df.columns]
                if available_columns:
                    extracted_df = extracted_df[available_columns]
            
            extraction_result['extracted_row_count'] = len(extracted_df)
            extraction_result['extraction_rate'] = len(extracted_df) / len(df) if len(df) > 0 else 0
            extraction_result['extracted_data'] = extracted_df
            
            return {
                'success': True,
                'extraction_result': extraction_result
            }
            
        except Exception as e:
            logger.error(f"多条件数据提取失败: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _apply_condition(self, df: pd.DataFrame, column: str, operator: str, value: Any) -> pd.Series:
        """应用单个条件"""
        condition_func = self.condition_operators[operator]
        
        if operator in ['is_null', 'is_not_null']:
            # 空值检查不需要value参数
            return df[column].apply(lambda x: condition_func(x, None))
        elif operator == 'regex':
            # 正则表达式匹配
            return df[column].apply(lambda x: condition_func(x, value))
        elif operator in ['in', 'not_in']:
            # 列表包含检查
            if not isinstance(value, (list, tuple, set)):
                value = [value]
            return df[column].apply(lambda x: condition_func(x, value))
        else:
            # 普通比较操作
            return df[column].apply(lambda x: condition_func(x, value))
    
    def extract_with_multiple_conditions(self, 
                                       file_path: str, 
                                       conditions: List[Dict[str, Any]], 
                                       sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """多条件数据提取（简化版本，与server.py调用兼容）
        
        Args:
            file_path: Excel文件路径
            conditions: 条件列表
            sheet_name: 工作表名称
            
        Returns:
            提取结果字典
        """
        try:
            # 解析条件格式
            processed_conditions = []
            logic_operator = 'and'  # 默认使用AND逻辑
            
            # 处理条件格式兼容性
            for condition in conditions:
                if isinstance(condition, dict):
                    # 标准格式：{'column': 'A', 'operator': 'eq', 'value': 'test'}
                    if all(key in condition for key in ['column', 'operator', 'value']):
                        processed_conditions.append(condition)
                    # 简化格式：{'A': 'test'} 或 {'A': {'operator': 'eq', 'value': 'test'}}
                    else:
                        for col, val in condition.items():
                            if isinstance(val, dict) and 'operator' in val:
                                processed_conditions.append({
                                    'column': col,
                                    'operator': val.get('operator', 'eq'),
                                    'value': val.get('value')
                                })
                            else:
                                processed_conditions.append({
                                    'column': col,
                                    'operator': 'eq',
                                    'value': val
                                })
                elif isinstance(condition, str):
                    # 字符串格式的逻辑操作符
                    if condition.lower() in ['and', 'or']:
                        logic_operator = condition.lower()
            
            # 如果没有有效条件，返回错误
            if not processed_conditions:
                return {
                    'success': False,
                    'error': '没有有效的提取条件',
                    'conditions_received': conditions
                }
            
            # 调用主要的提取方法
            result = self.extract_multi_condition_data(
                file_path=file_path,
                conditions=processed_conditions,
                logic_operator=logic_operator,
                sheet_name=sheet_name
            )
            
            # 格式化返回结果以匹配预期格式
            if result.get('success'):
                extraction_result = result['extraction_result']
                return {
                    'success': True,
                    'file_path': file_path,
                    'sheet_name': sheet_name,
                    'conditions_applied': processed_conditions,
                    'logic_operator': logic_operator,
                    'original_rows': extraction_result['original_row_count'],
                    'extracted_rows': extraction_result['extracted_row_count'],
                    'extraction_rate': extraction_result['extraction_rate'],
                    'extracted_data': extraction_result['extracted_data'],
                    'condition_analysis': extraction_result['condition_analysis'],
                    'timestamp': extraction_result['timestamp']
                }
            else:
                return result
                
        except Exception as e:
            logger.error(f"多条件数据提取失败: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'error_type': 'MULTI_CONDITION_EXTRACTION_ERROR'
            }
    
    def cross_validate_data(self, 
                          file_path: str,
                          reference_file: str,
                          key_columns: List[str],
                          compare_columns: Optional[List[str]] = None,
                          tolerance: float = 0.001) -> Dict[str, Any]:
        """交叉验证数据
        
        Args:
            file_path: 主文件路径
            reference_file: 参考文件路径
            key_columns: 关键列（用于匹配）
            compare_columns: 比较列
            tolerance: 数值比较容差
            
        Returns:
            验证结果字典
        """
        try:
            # 读取两个文件
            df1 = pd.read_excel(file_path)
            df2 = pd.read_excel(reference_file)
            
            validation_result = {
                'file1': file_path,
                'file2': reference_file,
                'timestamp': datetime.now().isoformat(),
                'key_columns': key_columns,
                'compare_columns': compare_columns,
                'tolerance': tolerance,
                'file1_rows': len(df1),
                'file2_rows': len(df2),
                'matching_analysis': {},
                'differences': {},
                'summary': {}
            }
            
            # 验证关键列是否存在
            missing_keys_df1 = [col for col in key_columns if col not in df1.columns]
            missing_keys_df2 = [col for col in key_columns if col not in df2.columns]
            
            if missing_keys_df1 or missing_keys_df2:
                return {
                    'success': False,
                    'error': f'关键列缺失 - 文件1: {missing_keys_df1}, 文件2: {missing_keys_df2}'
                }
            
            # 基于关键列进行匹配
            merged_df = pd.merge(
                df1, df2, 
                on=key_columns, 
                how='outer', 
                suffixes=('_file1', '_file2'),
                indicator=True
            )
            
            # 分析匹配情况
            matching_analysis = {
                'both_files': (merged_df['_merge'] == 'both').sum(),
                'only_file1': (merged_df['_merge'] == 'left_only').sum(),
                'only_file2': (merged_df['_merge'] == 'right_only').sum(),
                'total_unique_keys': len(merged_df)
            }
            
            validation_result['matching_analysis'] = matching_analysis
            
            # 比较指定列的差异
            if compare_columns:
                differences = {}
                
                # 只比较两个文件都有的记录
                both_df = merged_df[merged_df['_merge'] == 'both']
                
                for col in compare_columns:
                    col1 = f'{col}_file1'
                    col2 = f'{col}_file2'
                    
                    if col1 in both_df.columns and col2 in both_df.columns:
                        # 数值比较
                        if pd.api.types.is_numeric_dtype(both_df[col1]) and pd.api.types.is_numeric_dtype(both_df[col2]):
                            diff_mask = abs(both_df[col1] - both_df[col2]) > tolerance
                        else:
                            # 文本比较
                            diff_mask = both_df[col1].astype(str) != both_df[col2].astype(str)
                        
                        differences[col] = {
                            'different_count': diff_mask.sum(),
                            'different_rate': diff_mask.sum() / len(both_df) if len(both_df) > 0 else 0,
                            'different_rows': both_df[diff_mask].index.tolist()[:10]  # 前10个不同的行
                        }
                
                validation_result['differences'] = differences
            
            # 生成摘要
            summary = {
                'match_rate': matching_analysis['both_files'] / max(len(df1), len(df2)) if max(len(df1), len(df2)) > 0 else 0,
                'data_consistency': 'high' if matching_analysis['both_files'] / max(len(df1), len(df2)) > 0.9 else 'medium' if matching_analysis['both_files'] / max(len(df1), len(df2)) > 0.7 else 'low',
                'recommendations': []
            }
            
            if matching_analysis['only_file1'] > 0:
                summary['recommendations'].append(f"文件1中有 {matching_analysis['only_file1']} 条记录在文件2中不存在")
            
            if matching_analysis['only_file2'] > 0:
                summary['recommendations'].append(f"文件2中有 {matching_analysis['only_file2']} 条记录在文件1中不存在")
            
            validation_result['summary'] = summary
            
            return {
                'success': True,
                'validation_result': validation_result,
                'merged_data': merged_df
            }
            
        except Exception as e:
            logger.error(f"交叉验证失败: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }


class ExcelMultiTableMerger:
    """Excel 多表格数据合并器"""
    
    def __init__(self):
        self.merge_strategies = {
            'inner': 'inner',
            'outer': 'outer', 
            'left': 'left',
            'right': 'right'
        }
    
    def merge_multiple_tables(self, 
                            table_configs: List[Dict[str, Any]],
                            merge_strategy: str = 'outer',
                            key_columns: Optional[List[str]] = None,
                            output_file: Optional[str] = None) -> Dict[str, Any]:
        """合并多个表格
        
        Args:
            table_configs: 表格配置列表，每个配置包含文件路径、工作表等信息
            merge_strategy: 合并策略
            key_columns: 关键列
            output_file: 输出文件路径
            
        Returns:
            合并结果字典
        """
        try:
            merge_result = {
                'timestamp': datetime.now().isoformat(),
                'table_configs': table_configs,
                'merge_strategy': merge_strategy,
                'key_columns': key_columns,
                'tables_processed': 0,
                'merge_summary': {},
                'merged_data': None,
                'output_file': output_file
            }
            
            # 读取所有表格
            dataframes = []
            table_info = []
            
            for i, config in enumerate(table_configs):
                try:
                    file_path = config['file_path']
                    sheet_name = config.get('sheet_name')
                    table_name = config.get('name', f'table_{i}')
                    
                    # 读取表格
                    if sheet_name:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                    else:
                        df = pd.read_excel(file_path)
                    
                    # 添加表格来源标识
                    df['_source_table'] = table_name
                    
                    dataframes.append(df)
                    table_info.append({
                        'name': table_name,
                        'file_path': file_path,
                        'sheet_name': sheet_name,
                        'rows': len(df),
                        'columns': len(df.columns)
                    })
                    
                except Exception as e:
                    logger.warning(f"读取表格 {i} 失败: {str(e)}")
                    continue
            
            merge_result['tables_processed'] = len(dataframes)
            merge_result['table_info'] = table_info
            
            if len(dataframes) == 0:
                return {
                    'success': False,
                    'error': '没有成功读取任何表格',
                    'merge_result': merge_result
                }
            
            if len(dataframes) == 1:
                merged_df = dataframes[0]
            else:
                # 执行合并
                merged_df = self._merge_dataframes(
                    dataframes, merge_strategy, key_columns
                )
            
            # 生成合并摘要
            merge_summary = {
                'original_total_rows': sum(len(df) for df in dataframes),
                'merged_rows': len(merged_df),
                'original_total_columns': sum(len(df.columns) for df in dataframes),
                'merged_columns': len(merged_df.columns),
                'data_reduction_rate': 1 - (len(merged_df) / sum(len(df) for df in dataframes)) if sum(len(df) for df in dataframes) > 0 else 0
            }
            
            merge_result['merge_summary'] = merge_summary
            merge_result['merged_data'] = merged_df
            
            # 保存合并结果
            if output_file:
                merged_df.to_excel(output_file, index=False)
                merge_result['output_file'] = output_file
            
            return {
                'success': True,
                'merge_result': merge_result
            }
            
        except Exception as e:
            logger.error(f"多表格合并失败: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _merge_dataframes(self, 
                         dataframes: List[pd.DataFrame], 
                         strategy: str, 
                         key_columns: Optional[List[str]]) -> pd.DataFrame:
        """合并数据框"""
        if len(dataframes) == 1:
            return dataframes[0]
        
        # 如果没有指定关键列，尝试找到共同列
        if key_columns is None:
            common_columns = set(dataframes[0].columns)
            for df in dataframes[1:]:
                common_columns = common_columns.intersection(set(df.columns))
            
            # 排除来源标识列
            common_columns = common_columns - {'_source_table'}
            
            if common_columns:
                key_columns = list(common_columns)
            else:
                # 如果没有共同列，使用concat进行纵向合并
                return pd.concat(dataframes, ignore_index=True, sort=False)
        
        # 逐步合并
        result_df = dataframes[0]
        
        for df in dataframes[1:]:
            # 检查关键列是否存在
            available_keys = [col for col in key_columns if col in result_df.columns and col in df.columns]
            
            if available_keys:
                result_df = pd.merge(
                    result_df, df, 
                    on=available_keys, 
                    how=strategy,
                    suffixes=('', '_dup')
                )
                
                # 处理重复列
                dup_columns = [col for col in result_df.columns if col.endswith('_dup')]
                for dup_col in dup_columns:
                    original_col = dup_col.replace('_dup', '')
                    # 合并重复列的值（优先使用非空值）
                    if original_col in result_df.columns:
                        result_df[original_col] = result_df[original_col].fillna(result_df[dup_col])
                    result_df = result_df.drop(columns=[dup_col])
            else:
                # 如果没有可用的关键列，进行纵向合并
                result_df = pd.concat([result_df, df], ignore_index=True, sort=False)
        
        return result_df
    
    def merge_multiple_excel_files(self, 
                                 file_paths: List[str], 
                                 merge_config: Dict[str, Any], 
                                 output_path: Optional[str] = None) -> Dict[str, Any]:
        """合并多个Excel文件（与server.py调用兼容）
        
        Args:
            file_paths: Excel文件路径列表
            merge_config: 合并配置字典
            output_path: 输出文件路径
            
        Returns:
            合并结果字典
        """
        try:
            # 解析合并配置
            merge_strategy = merge_config.get('strategy', 'outer')
            key_columns = merge_config.get('key_columns')
            sheet_names = merge_config.get('sheet_names', {})
            table_names = merge_config.get('table_names', {})
            
            # 构建表格配置列表
            table_configs = []
            for i, file_path in enumerate(file_paths):
                config = {
                    'file_path': file_path,
                    'name': table_names.get(str(i), f'file_{i}'),
                    'sheet_name': sheet_names.get(str(i))
                }
                table_configs.append(config)
            
            # 调用主要的合并方法
            result = self.merge_multiple_tables(
                table_configs=table_configs,
                merge_strategy=merge_strategy,
                key_columns=key_columns,
                output_file=output_path
            )
            
            # 格式化返回结果以匹配预期格式
            if result.get('success'):
                merge_result = result['merge_result']
                return {
                    'success': True,
                    'file_paths': file_paths,
                    'merge_config': merge_config,
                    'output_path': output_path,
                    'files_processed': len(file_paths),
                    'tables_merged': merge_result['tables_processed'],
                    'merge_strategy': merge_strategy,
                    'key_columns': key_columns,
                    'original_total_rows': merge_result['merge_summary']['original_total_rows'],
                    'merged_rows': merge_result['merge_summary']['merged_rows'],
                    'merged_columns': merge_result['merge_summary']['merged_columns'],
                    'data_reduction_rate': merge_result['merge_summary']['data_reduction_rate'],
                    'merged_data': merge_result['merged_data'],
                    'table_info': merge_result['table_info'],
                    'timestamp': merge_result['timestamp'],
                    'output_saved': output_path is not None and merge_result.get('output_file') is not None
                }
            else:
                return result
                
        except Exception as e:
            logger.error(f"多Excel文件合并失败: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'error_type': 'MULTI_FILE_MERGE_ERROR',
                'file_paths': file_paths,
                'merge_config': merge_config
            }
    
    def merge_with_conditions(self, 
                            primary_file: str,
                            secondary_files: List[str],
                            merge_conditions: List[Dict[str, Any]],
                            output_file: Optional[str] = None) -> Dict[str, Any]:
        """基于条件合并表格
        
        Args:
            primary_file: 主文件路径
            secondary_files: 次要文件路径列表
            merge_conditions: 合并条件列表
            output_file: 输出文件路径
            
        Returns:
            合并结果字典
        """
        try:
            # 读取主文件
            primary_df = pd.read_excel(primary_file)
            
            merge_result = {
                'primary_file': primary_file,
                'secondary_files': secondary_files,
                'merge_conditions': merge_conditions,
                'timestamp': datetime.now().isoformat(),
                'merge_steps': [],
                'final_result': None
            }
            
            result_df = primary_df.copy()
            
            # 逐个处理次要文件
            for i, secondary_file in enumerate(secondary_files):
                try:
                    secondary_df = pd.read_excel(secondary_file)
                    
                    # 获取对应的合并条件
                    if i < len(merge_conditions):
                        condition = merge_conditions[i]
                    else:
                        condition = merge_conditions[-1]  # 使用最后一个条件
                    
                    # 执行条件合并
                    step_result = self._merge_with_condition(
                        result_df, secondary_df, condition, f"step_{i+1}"
                    )
                    
                    result_df = step_result['merged_data']
                    merge_result['merge_steps'].append(step_result)
                    
                except Exception as e:
                    logger.warning(f"合并文件 {secondary_file} 失败: {str(e)}")
                    continue
            
            merge_result['final_result'] = {
                'rows': len(result_df),
                'columns': len(result_df.columns),
                'merged_data': result_df
            }
            
            # 保存结果
            if output_file:
                result_df.to_excel(output_file, index=False)
                merge_result['output_file'] = output_file
            
            return {
                'success': True,
                'merge_result': merge_result
            }
            
        except Exception as e:
            logger.error(f"条件合并失败: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _merge_with_condition(self, 
                            primary_df: pd.DataFrame, 
                            secondary_df: pd.DataFrame, 
                            condition: Dict[str, Any],
                            step_name: str) -> Dict[str, Any]:
        """执行单步条件合并"""
        merge_type = condition.get('type', 'inner')
        key_columns = condition.get('key_columns', [])
        filter_condition = condition.get('filter')
        
        step_result = {
            'step_name': step_name,
            'merge_type': merge_type,
            'key_columns': key_columns,
            'primary_rows': len(primary_df),
            'secondary_rows': len(secondary_df),
            'merged_rows': 0,
            'merged_data': None
        }
        
        # 应用过滤条件
        filtered_secondary = secondary_df.copy()
        if filter_condition:
            # 这里可以扩展更复杂的过滤逻辑
            pass
        
        # 执行合并
        if key_columns and all(col in primary_df.columns and col in filtered_secondary.columns for col in key_columns):
            merged_df = pd.merge(
                primary_df, filtered_secondary,
                on=key_columns,
                how=merge_type,
                suffixes=('', '_secondary')
            )
        else:
            # 如果没有关键列，进行纵向合并
            merged_df = pd.concat([primary_df, filtered_secondary], ignore_index=True, sort=False)
        
        step_result['merged_rows'] = len(merged_df)
        step_result['merged_data'] = merged_df
        
        return step_result


class ExcelDataCleaner:
    """Excel 数据清洗器"""
    
    def __init__(self):
        self.cleaning_rules = self._initialize_cleaning_rules()
    
    def _initialize_cleaning_rules(self) -> Dict[str, Any]:
        """初始化清洗规则"""
        return {
            'whitespace': {
                'trim': True,
                'normalize_spaces': True,
                'remove_tabs': True,
                'remove_newlines': True
            },
            'duplicates': {
                'remove_exact_duplicates': True,
                'remove_subset_duplicates': False,
                'keep': 'first'  # 'first', 'last', False
            },
            'missing_values': {
                'fill_strategy': 'none',  # 'none', 'forward', 'backward', 'mean', 'median', 'mode'
                'drop_threshold': 0.5  # 缺失值比例超过此值的行/列将被删除
            },
            'outliers': {
                'method': 'iqr',  # 'iqr', 'zscore', 'isolation_forest'
                'threshold': 3.0,
                'action': 'flag'  # 'remove', 'flag', 'cap'
            },
            'data_types': {
                'auto_convert': True,
                'date_columns': [],
                'numeric_columns': [],
                'categorical_columns': []
            }
        }
    
    def clean_excel_data(self, 
                        file_path: str,
                        sheet_name: Optional[str] = None,
                        cleaning_config: Optional[Dict] = None,
                        output_file: Optional[str] = None) -> Dict[str, Any]:
        """清洗Excel数据
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            cleaning_config: 清洗配置
            output_file: 输出文件路径
            
        Returns:
            清洗结果字典
        """
        try:
            # 读取数据
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            
            original_df = df.copy()
            config = cleaning_config or self.cleaning_rules
            
            cleaning_result = {
                'file_path': file_path,
                'sheet_name': sheet_name,
                'timestamp': datetime.now().isoformat(),
                'original_shape': df.shape,
                'cleaning_steps': [],
                'final_shape': None,
                'cleaned_data': None,
                'cleaning_summary': {}
            }
            
            # 1. 清理空白字符
            if config.get('whitespace', {}).get('trim', True):
                step_result = self._clean_whitespace(df, config['whitespace'])
                cleaning_result['cleaning_steps'].append(step_result)
                df = step_result['cleaned_data']
            
            # 2. 处理重复数据
            if config.get('duplicates', {}).get('remove_exact_duplicates', True):
                step_result = self._remove_duplicates(df, config['duplicates'])
                cleaning_result['cleaning_steps'].append(step_result)
                df = step_result['cleaned_data']
            
            # 3. 处理缺失值
            missing_config = config.get('missing_values', {})
            if missing_config.get('fill_strategy', 'none') != 'none':
                step_result = self._handle_missing_values(df, missing_config)
                cleaning_result['cleaning_steps'].append(step_result)
                df = step_result['cleaned_data']
            
            # 4. 处理异常值
            outlier_config = config.get('outliers', {})
            if outlier_config.get('method'):
                step_result = self._handle_outliers(df, outlier_config)
                cleaning_result['cleaning_steps'].append(step_result)
                df = step_result['cleaned_data']
            
            # 5. 数据类型转换
            if config.get('data_types', {}).get('auto_convert', True):
                step_result = self._convert_data_types(df, config['data_types'])
                cleaning_result['cleaning_steps'].append(step_result)
                df = step_result['cleaned_data']
            
            cleaning_result['final_shape'] = df.shape
            cleaning_result['cleaned_data'] = df
            
            # 生成清洗摘要
            summary = self._generate_cleaning_summary(original_df, df, cleaning_result['cleaning_steps'])
            cleaning_result['cleaning_summary'] = summary
            
            # 保存清洗后的数据
            if output_file:
                df.to_excel(output_file, index=False)
                cleaning_result['output_file'] = output_file
            
            return {
                'success': True,
                'cleaning_result': cleaning_result
            }
            
        except Exception as e:
            logger.error(f"数据清洗失败: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _clean_whitespace(self, df: pd.DataFrame, config: Dict) -> Dict[str, Any]:
        """清理空白字符"""
        cleaned_df = df.copy()
        changes_count = 0
        
        for col in cleaned_df.columns:
            if cleaned_df[col].dtype == 'object':
                original_values = cleaned_df[col].astype(str)
                
                if config.get('trim', True):
                    cleaned_df[col] = cleaned_df[col].astype(str).str.strip()
                
                if config.get('normalize_spaces', True):
                    cleaned_df[col] = cleaned_df[col].astype(str).str.replace(r'\s+', ' ', regex=True)
                
                if config.get('remove_tabs', True):
                    cleaned_df[col] = cleaned_df[col].astype(str).str.replace('\t', ' ', regex=False)
                
                if config.get('remove_newlines', True):
                    cleaned_df[col] = cleaned_df[col].astype(str).str.replace('\n', ' ', regex=False)
                    cleaned_df[col] = cleaned_df[col].astype(str).str.replace('\r', ' ', regex=False)
                
                # 统计变化
                changes_count += (original_values != cleaned_df[col].astype(str)).sum()
        
        return {
            'step_name': 'whitespace_cleaning',
            'changes_count': changes_count,
            'cleaned_data': cleaned_df
        }
    
    def _remove_duplicates(self, df: pd.DataFrame, config: Dict) -> Dict[str, Any]:
        """移除重复数据"""
        original_count = len(df)
        
        if config.get('remove_exact_duplicates', True):
            cleaned_df = df.drop_duplicates(keep=config.get('keep', 'first'))
        else:
            cleaned_df = df.copy()
        
        removed_count = original_count - len(cleaned_df)
        
        return {
            'step_name': 'duplicate_removal',
            'removed_count': removed_count,
            'cleaned_data': cleaned_df
        }
    
    def _handle_missing_values(self, df: pd.DataFrame, config: Dict) -> Dict[str, Any]:
        """处理缺失值"""
        cleaned_df = df.copy()
        strategy = config.get('fill_strategy', 'none')
        
        filled_count = 0
        
        if strategy == 'forward':
            filled_count = cleaned_df.isnull().sum().sum()
            cleaned_df = cleaned_df.fillna(method='ffill')
        elif strategy == 'backward':
            filled_count = cleaned_df.isnull().sum().sum()
            cleaned_df = cleaned_df.fillna(method='bfill')
        elif strategy in ['mean', 'median']:
            for col in cleaned_df.columns:
                if pd.api.types.is_numeric_dtype(cleaned_df[col]):
                    missing_count = cleaned_df[col].isnull().sum()
                    if missing_count > 0:
                        if strategy == 'mean':
                            fill_value = cleaned_df[col].mean()
                        else:
                            fill_value = cleaned_df[col].median()
                        cleaned_df[col] = cleaned_df[col].fillna(fill_value)
                        filled_count += missing_count
        elif strategy == 'mode':
            for col in cleaned_df.columns:
                missing_count = cleaned_df[col].isnull().sum()
                if missing_count > 0:
                    mode_value = cleaned_df[col].mode()
                    if len(mode_value) > 0:
                        cleaned_df[col] = cleaned_df[col].fillna(mode_value[0])
                        filled_count += missing_count
        
        return {
            'step_name': 'missing_value_handling',
            'filled_count': filled_count,
            'strategy': strategy,
            'cleaned_data': cleaned_df
        }
    
    def _handle_outliers(self, df: pd.DataFrame, config: Dict) -> Dict[str, Any]:
        """处理异常值"""
        cleaned_df = df.copy()
        method = config.get('method', 'iqr')
        action = config.get('action', 'flag')
        threshold = config.get('threshold', 3.0)
        
        outliers_info = {}
        total_outliers = 0
        
        for col in cleaned_df.columns:
            if pd.api.types.is_numeric_dtype(cleaned_df[col]):
                outliers_mask = self._detect_outliers(cleaned_df[col], method, threshold)
                outliers_count = outliers_mask.sum()
                
                if outliers_count > 0:
                    outliers_info[col] = {
                        'count': outliers_count,
                        'indices': cleaned_df[outliers_mask].index.tolist()
                    }
                    total_outliers += outliers_count
                    
                    if action == 'remove':
                        cleaned_df = cleaned_df[~outliers_mask]
                    elif action == 'cap':
                        # 使用IQR方法进行截断
                        Q1 = cleaned_df[col].quantile(0.25)
                        Q3 = cleaned_df[col].quantile(0.75)
                        IQR = Q3 - Q1
                        lower_bound = Q1 - 1.5 * IQR
                        upper_bound = Q3 + 1.5 * IQR
                        
                        cleaned_df[col] = cleaned_df[col].clip(lower=lower_bound, upper=upper_bound)
        
        return {
            'step_name': 'outlier_handling',
            'method': method,
            'action': action,
            'total_outliers': total_outliers,
            'outliers_info': outliers_info,
            'cleaned_data': cleaned_df
        }
    
    def _detect_outliers(self, series: pd.Series, method: str, threshold: float) -> pd.Series:
        """检测异常值"""
        if method == 'iqr':
            Q1 = series.quantile(0.25)
            Q3 = series.quantile(0.75)
            IQR = Q3 - Q1
            lower_bound = Q1 - 1.5 * IQR
            upper_bound = Q3 + 1.5 * IQR
            return (series < lower_bound) | (series > upper_bound)
        
        elif method == 'zscore':
            z_scores = np.abs((series - series.mean()) / series.std())
            return z_scores > threshold
        
        else:
            # 默认使用IQR方法
            return self._detect_outliers(series, 'iqr', threshold)
    
    def _convert_data_types(self, df: pd.DataFrame, config: Dict) -> Dict[str, Any]:
        """转换数据类型"""
        cleaned_df = df.copy()
        conversions = {}
        
        # 自动转换
        if config.get('auto_convert', True):
            for col in cleaned_df.columns:
                original_type = str(cleaned_df[col].dtype)
                
                # 尝试转换为数值类型
                if cleaned_df[col].dtype == 'object':
                    try:
                        # 尝试转换为数值
                        numeric_series = pd.to_numeric(cleaned_df[col], errors='coerce')
                        if not numeric_series.isnull().all():
                            cleaned_df[col] = numeric_series
                            conversions[col] = {'from': original_type, 'to': str(cleaned_df[col].dtype)}
                    except:
                        pass
        
        # 指定列的转换
        for col in config.get('date_columns', []):
            if col in cleaned_df.columns:
                try:
                    cleaned_df[col] = pd.to_datetime(cleaned_df[col])
                    conversions[col] = {'from': 'object', 'to': 'datetime64[ns]'}
                except:
                    pass
        
        for col in config.get('numeric_columns', []):
            if col in cleaned_df.columns:
                try:
                    cleaned_df[col] = pd.to_numeric(cleaned_df[col])
                    conversions[col] = {'from': 'object', 'to': str(cleaned_df[col].dtype)}
                except:
                    pass
        
        for col in config.get('categorical_columns', []):
            if col in cleaned_df.columns:
                try:
                    cleaned_df[col] = cleaned_df[col].astype('category')
                    conversions[col] = {'from': 'object', 'to': 'category'}
                except:
                    pass
        
        return {
            'step_name': 'data_type_conversion',
            'conversions': conversions,
            'cleaned_data': cleaned_df
        }
    
    def _generate_cleaning_summary(self, 
                                 original_df: pd.DataFrame, 
                                 cleaned_df: pd.DataFrame, 
                                 steps: List[Dict]) -> Dict[str, Any]:
        """生成清洗摘要"""
        summary = {
            'original_shape': original_df.shape,
            'final_shape': cleaned_df.shape,
            'rows_removed': original_df.shape[0] - cleaned_df.shape[0],
            'columns_affected': 0,
            'total_changes': 0,
            'steps_summary': {}
        }
        
        for step in steps:
            step_name = step['step_name']
            summary['steps_summary'][step_name] = {
                'changes': step.get('changes_count', 0) + step.get('filled_count', 0) + step.get('removed_count', 0),
                'details': {k: v for k, v in step.items() if k not in ['step_name', 'cleaned_data']}
            }
            summary['total_changes'] += summary['steps_summary'][step_name]['changes']
        
        return summary
    
    def suggest_cleaning_operations(self, df: pd.DataFrame) -> Dict[str, Any]:
        """建议数据清洗操作
        
        Args:
            df: 待分析的DataFrame
            
        Returns:
            清洗建议字典
        """
        suggestions = {
            'whitespace_issues': [],
            'duplicate_issues': [],
            'missing_value_issues': [],
            'outlier_issues': [],
            'data_type_issues': [],
            'overall_recommendations': []
        }
        
        try:
            # 检查空白字符问题
            for col in df.columns:
                if df[col].dtype == 'object':
                    # 检查前后空格
                    has_leading_spaces = df[col].astype(str).str.startswith(' ').any()
                    has_trailing_spaces = df[col].astype(str).str.endswith(' ').any()
                    has_multiple_spaces = df[col].astype(str).str.contains(r'\s{2,}').any()
                    
                    if has_leading_spaces or has_trailing_spaces or has_multiple_spaces:
                        suggestions['whitespace_issues'].append({
                            'column': col,
                            'issues': {
                                'leading_spaces': has_leading_spaces,
                                'trailing_spaces': has_trailing_spaces,
                                'multiple_spaces': has_multiple_spaces
                            },
                            'recommendation': '建议清理空白字符'
                        })
            
            # 检查重复数据
            duplicate_count = df.duplicated().sum()
            if duplicate_count > 0:
                suggestions['duplicate_issues'].append({
                    'total_duplicates': duplicate_count,
                    'percentage': (duplicate_count / len(df)) * 100,
                    'recommendation': f'发现{duplicate_count}行重复数据，建议删除'
                })
            
            # 检查缺失值
            missing_stats = df.isnull().sum()
            for col, missing_count in missing_stats.items():
                if missing_count > 0:
                    missing_percentage = (missing_count / len(df)) * 100
                    suggestions['missing_value_issues'].append({
                        'column': col,
                        'missing_count': missing_count,
                        'missing_percentage': missing_percentage,
                        'recommendation': self._get_missing_value_recommendation(missing_percentage)
                    })
            
            # 检查异常值
            for col in df.columns:
                if pd.api.types.is_numeric_dtype(df[col]):
                    outliers_mask = self._detect_outliers(df[col], 'iqr', 3.0)
                    outliers_count = outliers_mask.sum()
                    if outliers_count > 0:
                        suggestions['outlier_issues'].append({
                            'column': col,
                            'outliers_count': outliers_count,
                            'outliers_percentage': (outliers_count / len(df)) * 100,
                            'recommendation': '建议检查异常值并决定处理方式'
                        })
            
            # 检查数据类型问题
            for col in df.columns:
                if df[col].dtype == 'object':
                    # 检查是否可以转换为数值类型
                    try:
                        numeric_series = pd.to_numeric(df[col], errors='coerce')
                        if not numeric_series.isnull().all() and numeric_series.isnull().sum() < len(df) * 0.1:
                            suggestions['data_type_issues'].append({
                                'column': col,
                                'current_type': 'object',
                                'suggested_type': 'numeric',
                                'recommendation': '建议转换为数值类型'
                            })
                    except:
                        pass
                    
                    # 检查是否可以转换为日期类型
                    try:
                        date_series = pd.to_datetime(df[col], errors='coerce')
                        if not date_series.isnull().all() and date_series.isnull().sum() < len(df) * 0.1:
                            suggestions['data_type_issues'].append({
                                'column': col,
                                'current_type': 'object',
                                'suggested_type': 'datetime',
                                'recommendation': '建议转换为日期类型'
                            })
                    except:
                        pass
            
            # 生成总体建议
            total_issues = (len(suggestions['whitespace_issues']) + 
                          len(suggestions['duplicate_issues']) + 
                          len(suggestions['missing_value_issues']) + 
                          len(suggestions['outlier_issues']) + 
                          len(suggestions['data_type_issues']))
            
            if total_issues == 0:
                suggestions['overall_recommendations'].append('数据质量良好，无需特殊清洗操作')
            else:
                suggestions['overall_recommendations'].append(f'发现{total_issues}个数据质量问题，建议进行相应清洗操作')
                
                if suggestions['whitespace_issues']:
                    suggestions['overall_recommendations'].append('优先处理空白字符问题')
                if suggestions['duplicate_issues']:
                    suggestions['overall_recommendations'].append('删除重复数据以减少数据冗余')
                if suggestions['missing_value_issues']:
                    suggestions['overall_recommendations'].append('根据业务需求处理缺失值')
                if suggestions['data_type_issues']:
                    suggestions['overall_recommendations'].append('优化数据类型以提高处理效率')
            
            return suggestions
            
        except Exception as e:
            logger.error(f"生成清洗建议失败: {str(e)}")
            return {
                'error': str(e),
                'overall_recommendations': ['数据分析失败，请检查数据格式']
            }
    
    def _get_missing_value_recommendation(self, missing_percentage: float) -> str:
        """根据缺失值比例获取处理建议"""
        if missing_percentage < 5:
            return '缺失值较少，可考虑删除或填充'
        elif missing_percentage < 20:
            return '缺失值适中，建议根据业务逻辑填充'
        elif missing_percentage < 50:
            return '缺失值较多，需要仔细考虑处理策略'
        else:
            return '缺失值过多，建议考虑删除该列或重新收集数据'


class ExcelBatchProcessor:
    """Excel 批量处理器"""
    
    def __init__(self):
        self.quality_controller = ExcelDataQualityController()
        self.content_extractor = ExcelCellContentExtractor()
        self.character_converter = ExcelCharacterConverter()
        self.condition_extractor = ExcelMultiConditionExtractor()
        self.table_merger = ExcelMultiTableMerger()
        self.data_cleaner = ExcelDataCleaner()
    
    def batch_process_files(self, 
                          file_paths: List[str],
                          processing_config: Dict[str, Any],
                          output_directory: Optional[str] = None) -> Dict[str, Any]:
        """批量处理Excel文件
        
        Args:
            file_paths: 文件路径列表
            processing_config: 处理配置
            output_directory: 输出目录
            
        Returns:
            批量处理结果字典
        """
        try:
            batch_result = {
                'timestamp': datetime.now().isoformat(),
                'total_files': len(file_paths),
                'processed_files': 0,
                'failed_files': 0,
                'processing_config': processing_config,
                'file_results': {},
                'summary': {}
            }
            
            # 创建输出目录
            if output_directory:
                Path(output_directory).mkdir(parents=True, exist_ok=True)
            
            # 处理每个文件
            for file_path in file_paths:
                try:
                    file_result = self._process_single_file(
                        file_path, processing_config, output_directory
                    )
                    
                    batch_result['file_results'][file_path] = file_result
                    
                    if file_result['success']:
                        batch_result['processed_files'] += 1
                    else:
                        batch_result['failed_files'] += 1
                        
                except Exception as e:
                    logger.error(f"处理文件 {file_path} 失败: {str(e)}")
                    batch_result['file_results'][file_path] = {
                        'success': False,
                        'error': str(e)
                    }
                    batch_result['failed_files'] += 1
            
            # 生成摘要
            summary = self._generate_batch_summary(batch_result)
            batch_result['summary'] = summary
            
            return {
                'success': True,
                'batch_result': batch_result
            }
            
        except Exception as e:
            logger.error(f"批量处理失败: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _process_single_file(self, 
                           file_path: str, 
                           config: Dict[str, Any], 
                           output_dir: Optional[str]) -> Dict[str, Any]:
        """处理单个文件"""
        file_result = {
            'file_path': file_path,
            'timestamp': datetime.now().isoformat(),
            'operations': [],
            'success': True,
            'final_output': None
        }
        
        current_data = None
        
        # 按配置执行操作
        operations = config.get('operations', [])
        
        for operation in operations:
            op_type = operation.get('type')
            op_config = operation.get('config', {})
            
            try:
                if op_type == 'validate':
                    result = self.quality_controller.validate_data_integrity(
                        file_path, **op_config
                    )
                elif op_type == 'extract_content':
                    result = self.content_extractor.extract_cell_content(
                        file_path, **op_config
                    )
                elif op_type == 'convert_characters':
                    result = self.character_converter.convert_character_formats(
                        file_path, **op_config
                    )
                elif op_type == 'extract_conditions':
                    result = self.condition_extractor.extract_multi_condition_data(
                        file_path, **op_config
                    )
                elif op_type == 'clean_data':
                    result = self.data_cleaner.clean_excel_data(
                        file_path, **op_config
                    )
                else:
                    result = {
                        'success': False,
                        'error': f'不支持的操作类型: {op_type}'
                    }
                
                file_result['operations'].append({
                    'type': op_type,
                    'success': result['success'],
                    'result': result
                })
                
                if not result['success']:
                    file_result['success'] = False
                    break
                
                # 更新当前数据
                if 'converted_data' in result:
                    current_data = result['converted_data']
                elif 'cleaned_data' in result.get('cleaning_result', {}):
                    current_data = result['cleaning_result']['cleaned_data']
                
            except Exception as e:
                file_result['operations'].append({
                    'type': op_type,
                    'success': False,
                    'error': str(e)
                })
                file_result['success'] = False
                break
        
        # 保存最终结果
        if file_result['success'] and current_data is not None and output_dir:
            output_path = Path(output_dir) / f"processed_{Path(file_path).name}"
            current_data.to_excel(output_path, index=False)
            file_result['final_output'] = str(output_path)
        
        return file_result
    
    def _generate_batch_summary(self, batch_result: Dict[str, Any]) -> Dict[str, Any]:
        """生成批量处理摘要"""
        summary = {
            'success_rate': batch_result['processed_files'] / batch_result['total_files'] if batch_result['total_files'] > 0 else 0,
            'operation_statistics': {},
            'common_errors': {},
            'processing_time': 'N/A'  # 可以添加时间统计
        }
        
        # 统计操作类型
        operation_counts = defaultdict(int)
        error_counts = defaultdict(int)
        
        for file_result in batch_result['file_results'].values():
            if isinstance(file_result, dict) and 'operations' in file_result:
                for operation in file_result['operations']:
                    op_type = operation['type']
                    operation_counts[op_type] += 1
                    
                    if not operation['success']:
                        error = operation.get('error', 'Unknown error')
                        error_counts[error] += 1
        
        summary['operation_statistics'] = dict(operation_counts)
        summary['common_errors'] = dict(error_counts)
        
        return summary


# 导出主要类
__all__ = [
    'ExcelDataQualityController',
    'ExcelCellContentExtractor', 
    'ExcelCharacterConverter',
    'ExcelMultiConditionExtractor',
    'ExcelMultiTableMerger',
    'ExcelDataCleaner',
    'ExcelBatchProcessor'
]